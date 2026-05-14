"""
matching.py
companies_db.csv × notices_db.csv 매칭
→ send_history.csv로 중복 발송 방지
→ 매칭결과_YYYYMMDD.xlsx 저장
"""
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re, os

WORK_DIR = r'C:\Users\fbwlg\Desktop\26년도\2. 원스톱\7. 정보 전달 체계 구축'

REALM_CODE = {
    "금융":"01","기술개발":"02","인력":"03","수출":"04",
    "내수":"05","창업":"06","경영":"07","기타":"09",
}
HIGH = ["혁신제품","혁신조달","G-PASS","혁신기업","해외조달","공공구매","조달청"]
MID  = ["해외판로","수출바우처","수출지원","해외진출","글로벌","스케일업","판로개척","해외마케팅"]


def score_notice(notice, row, already_sent):
    pid = notice.get('pblancId','')

    # ① 이미 발송한 공고 제외
    if (row['기업명'], pid) in already_sent:
        return None

    # ② 마감일 지난 공고 제외
    deadline = notice.get('마감일','')
    if deadline and deadline < datetime.today().strftime("%Y-%m-%d"):
        return None

    # ③ 수출실적 없으면 수출 공고 제외
    if str(row.get('수출실적','')) == '아니오' and '수출' in str(notice.get('분야','')):
        return None

    text = " ".join([
        str(notice.get('공고명','')),
        str(notice.get('사업개요','')),
        str(notice.get('해시태그','')),
        str(notice.get('주관기관','')),
        str(notice.get('지원대상','')),
    ])

    matched_high = [kw for kw in HIGH if kw in text]
    matched_mid  = [kw for kw in MID  if kw in text]
    sys_score    = len(matched_high)*3 + len(matched_mid)*2

    raw_kw      = str(row.get('기술키워드','')) + ',' + str(row.get('제품분야',''))
    company_kws = [k.strip() for k in raw_kw.split(',') if k.strip() and k.strip()!='nan']
    matched_co  = [kw for kw in company_kws if kw in text]
    co_score    = len(matched_co)*2

    country       = str(row.get('수출국가',''))
    country_score = 2 if (country and country!='nan' and country in text) else 0

    total = sys_score + co_score + country_score
    if not (matched_high or matched_mid or matched_co):
        return None

    if matched_high or total >= 6: stars = "★★★"
    elif matched_mid or matched_co: stars = "★★"
    else: return None

    return {
        "기업명":         row['기업명'],
        "관련도":         stars,
        "점수":           total,
        "공고ID":         pid,
        "공고명":         notice.get('공고명',''),
        "주관기관":       notice.get('주관기관',''),
        "접수기간":       notice.get('접수기간',''),
        "마감일":         deadline,
        "사업개요":       str(notice.get('사업개요',''))[:150] + "...",
        "시스템매칭":     ", ".join(matched_high+matched_mid),
        "기업키워드매칭": ", ".join(matched_co),
        "국가매칭":       country if country_score else "",
        "공고링크":       notice.get('공고링크',''),
        "담당자검토":     "",
        "검토의견":       "",
    }


def save_excel(results):
    wb = Workbook()
    ws = wb.active; ws.title = "매칭결과"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:O1")
    t=ws["A1"]
    t.value=f"원스톱 스케일업 — 공고 매칭 결과  |  생성일: {datetime.today().strftime('%Y.%m.%d')}"
    t.fill=PatternFill("solid",start_color="1F4E79",end_color="1F4E79")
    t.font=Font(name="맑은 고딕",bold=True,size=12,color="FFFFFF")
    t.alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[1].height=30

    headers=["기업명","관련도","점수","공고명","주관기관","접수기간","마감일",
             "사업개요","시스템매칭","기업키워드매칭","국가매칭",
             "공고링크","담당자검토\n(○/✕)","검토의견","공고ID"]
    widths =[18,8,6,38,20,22,12,38,25,25,10,45,12,28,22]

    s=Side(style="thin",color="BFBFBF")
    bdr=Border(left=s,right=s,top=s,bottom=s)

    for i,(h,w) in enumerate(zip(headers,widths),1):
        ws.column_dimensions[get_column_letter(i)].width=w
        c=ws.cell(row=2,column=i,value=h)
        c.fill=PatternFill("solid",start_color="2E75B6",end_color="2E75B6")
        c.font=Font(name="맑은 고딕",bold=True,size=10,color="FFFFFF")
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bdr
    ws.row_dimensions[2].height=30

    star_colors={"★★★":"FFF2CC","★★":"E2EFDA"}
    for r_i,row in enumerate(results,3):
        bg=star_colors.get(row.get("관련도",""),"FFFFFF")
        ws.row_dimensions[r_i].height=36
        keys=["기업명","관련도","점수","공고명","주관기관","접수기간","마감일",
              "사업개요","시스템매칭","기업키워드매칭","국가매칭",
              "공고링크","담당자검토","검토의견","공고ID"]
        for c_i,key in enumerate(keys,1):
            cell_bg="D5E8D4" if c_i in [13,14] else bg
            c=ws.cell(row=r_i,column=c_i,value=row.get(key,""))
            c.fill=PatternFill("solid",start_color=cell_bg,end_color=cell_bg)
            c.font=Font(name="맑은 고딕",size=9)
            c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=False)
            c.border=bdr

    note_r=len(results)+4
    ws.merge_cells(f"A{note_r}:O{note_r}")
    n=ws.cell(row=note_r,column=1,
        value="※ 담당자검토란 ○(승인)/✕(제외) 입력 후 send.py 실행 → ○ 건만 메일+캘린더 처리 / 이미 발송한 공고는 자동 제외됨")
    n.font=Font(name="맑은 고딕",size=9,color="C00000",italic=True)

    today=datetime.today().strftime("%Y%m%d")
    path=f"매칭결과_{today}.xlsx"
    wb.save(path); return path


def main():
    os.chdir(WORK_DIR)

    # DB 로드
    df_companies = pd.read_excel(
        'TEST_(WALLA) 2026년도 스케일업 프로그램 선정기업.xlsx')
    col_map={
        '기업명':       '(기본정보) 기업명을 입력해 주시기 바랍니다.',
        '이메일':       '(기본정보) 담당자 이메일을 입력해 주시기 바랍니다.',
        '관심사업분야': '(수요파악) 관심있는 정부 사업 분야를 선택하여 주시기 바랍니다.(최대 2개 선택 가능)',
        '기술키워드':   '(수요파악) 귀사 제품의 주요 기술/분야 키워드를 입력하여 주시기 바랍니다.',
        '제품분야':     '(기본정보) 귀사의 제품/기술 분야를 선택하여 주시기 바랍니다.(최대 3개선택 가능)',
        '수출실적':     '(기업현황) 최근 3년간 수출 실적 여부를 선택하여 주시기 바랍니다.',
        '수출국가':     '① 주요 수출 국가를 입력하여 주시기 바랍니다.',
    }
    df_c=pd.DataFrame({k:df_companies[v] for k,v in col_map.items() if v in df_companies.columns})
    df_c=df_c.fillna("")

    df_notices=pd.read_csv("notices_db.csv",dtype=str).fillna("")

    # 발송 이력 로드
    history_file="send_history.csv"
    if os.path.exists(history_file):
        df_hist=pd.read_csv(history_file,dtype=str).fillna("")
        already_sent=set(zip(df_hist['기업명'],df_hist['pblancId']))
        print(f"발송 이력: {len(df_hist)}건 → 중복 제외 적용")
    else:
        already_sent=set()
        print("발송 이력 없음 (첫 실행)")

    print(f"\n{'='*55}")
    print(f"매칭 시작 — 기업 {len(df_c)}개사 × 공고 {len(df_notices)}건")
    print(f"{'='*55}\n")

    all_results=[]
    for _,row in df_c.iterrows():
        company=row['기업명']
        interest=row['관심사업분야']
        codes=[v for k,v in REALM_CODE.items() if k in interest] or ["02","04"]
        notices_filtered=df_notices[df_notices['분야코드_대응'].isin(codes)] if '분야코드_대응' in df_notices.columns else df_notices

        scored=[r for _,n in notices_filtered.iterrows()
                if (r:=score_notice(n.to_dict(),row,already_sent))]
        scored.sort(key=lambda x:-x['점수'])
        top=scored[:5]
        print(f"▶ {company}: {len(scored)}건 매칭 → 상위 {len(top)}건")
        all_results.extend(top)

    path=save_excel(all_results)
    print(f"\n{'='*55}")
    print(f"저장 완료: {path} (총 {len(all_results)}건)")
    print("담당자 검토(○/✕) 입력 후 send.py 실행하세요.")


if __name__=="__main__":
    main()
