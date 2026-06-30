"""
원스톱 스케일업 — 공고 매칭 시스템 _UJH
"""
import streamlit as st
import pandas as pd
import requests
import re, os, json, io, subprocess
from datetime import datetime, timedelta

# Playwright 브라우저 자동 설치 (Streamlit Cloud 대응)
@st.cache_resource
def install_playwright():
    try:
        subprocess.run(["playwright", "install", "chromium"], 
                      capture_output=True, timeout=120)
    except Exception:
        pass

install_playwright()

st.set_page_config(
    page_title="원스톱 스케일업_UJH",
    page_icon="📢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── 비밀번호 인증 ─────────────────────────────────────
def check_password():
    if st.session_state.get("authenticated"):
        return True
    st.markdown("""
    <div style="max-width:380px;margin:100px auto;text-align:center;">
      <div style="width:52px;height:52px;background:#ECFDF5;
                  border:1px solid #A7F3D0;border-radius:14px;
                  display:flex;align-items:center;justify-content:center;
                  margin:0 auto 20px;font-size:24px;box-shadow:0 2px 8px rgba(16,185,129,0.15);">📢</div>
      <h2 style="color:#0F172A;font-size:22px;font-weight:700;margin:0 0 6px;
                 font-family:'Inter','Apple SD Gothic Neo',sans-serif;">원스톱 스케일업</h2>
      <p style="color:#64748B;font-size:13px;margin:0 0 32px;
                font-family:'Inter','Apple SD Gothic Neo',sans-serif;">혁신제품지원센터 공고 매칭 시스템</p>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        pw = st.text_input("비밀번호", type="password",
                           placeholder="비밀번호 입력", label_visibility="collapsed")
        if st.button("로그인", use_container_width=True, type="primary"):
            correct = st.secrets.get("password", "scaleup2026")
            if pw == correct:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다")
    return False

if not check_password():
    st.stop()

# ── 상수 ──────────────────────────────────────────────
DRIVE_FOLDER_ID  = "1iWGYjaoslqST45ggDlg-IPMLaUHCYmV_"
LOGO_URL = "https://raw.githubusercontent.com/onestopkipcc-web/scaleup-matching/main/logo.png"
API_KEY          = "Nt604D"
BASE_URL         = "https://www.bizinfo.go.kr/uss/rss/bizinfoApi.do"
SELECTED_FILE    = "선정기업_명단.xlsx"   # ← 핵심 변경
NOTICES_FILE     = "notices_db.xlsx"
HISTORY_FILE     = "send_history.xlsx"
KEYWORDS_FILE    = "keywords.json"
DETAIL_FILE      = "notices_detail.xlsx"  # 크롤링 전문 DB
AI_ANALYSIS_FILE = "ai_analysis.json"     # AI 분석 결과 영구 저장
CALID_FILE       = "calendar_id.txt"
CATCAL_FILE      = "category_calendars.json"
INDCAL_FILE      = "individual_calendars.json"

# ── 키워드 두 축 구조 ────────────────────────────────
# 대상: 해외조달시장 진출 희망 조달 기업 (혁신제품 지정 여부 무관)

# 축1: 지원대상 (누구를 위한 공고인가)
TARGET_KW = {
    "조달기업특화": [
        # 조달 자격·인증
        "조달청","G-PASS","혁신제품","MAS","우선구매",
        "공공조달","해외조달","나라장터","조달우수","혁신조달",
        "혁신기업","혁신바우처","혁신성장","공공기관",
        "벤처나라","우수제품","성능인증","우수조달",
        "다수공급자계약","조달등록","조달기업","공공구매",
        "혁신제품지정","공공판로","정부조달","조달시장",
        # 해외조달
        "해외조달시장","UN조달","해외공공조달","글로벌조달",
        "조달한류","공공혁신",
    ],
    "수출기업특화": [
        # 수출 단계별
        "수출기업","수출유망","수출강소","글로벌강소","수출초보",
        "수출성장","수출실적기업","수출유망중소기업","수출유망기업",
        "내수기업","수출전환","첫수출","수출입문",
        # 기관·제도
        "FTA","무역","KOTRA","해외개척","무역보험","한국무역",
        "무역협회","수출보험","수출금융","수출보증",
        # 수출 목적지
        "글로벌","해외","수출","신흥시장","新남방","신북방",
    ],
    "중소벤처일반": [
        # 기업 유형
        "중소기업","중소·중견","벤처기업","이노비즈","메인비즈",
        "강소기업","스타기업","유망기업","월드클래스","히든챔피언",
        "중견기업","소기업","예비창업","창업기업","스타트업",
        "소상공인 제외","초기기업","시드","시리즈",
        # 성장 단계
        "성장기업","스케일업","Scale-up","유니콘","예비유니콘",
        "아기유니콘","팁스","tips","도약","도전",
    ],
}
# 축2: 사업성격 (어떤 종류의 지원인가)
TYPE_KW = {
    "공공조달": [
        "조달청","혁신조달","시범구매","우선구매","해외조달",
        "공공구매","MAS등록","나라장터","조달시장","공공기관납품",
        "조달등록","혁신구매","공공혁신","정부구매","공공기관구매",
        "조달지원","혁신제품구매","혁신시장","조달우수제품",
        "공공판로","공공기관","납품실적","조달실적",
    ],
    "해외진출": [
        "해외판로","수출바우처","해외마케팅","판로개척","해외진출",
        "수출지원","해외시장","수출컨소시엄","해외무역",
        "수출인큐베이터","해외거점","글로벌진출","해외진출지원",
        "해외시장개척","수출역량","해외수출","수출활성화",
        "해외판로개척","수출플랫폼","해외기업","바이어발굴",
        "해외네트워크","글로벌파트너","수출상담","무역사절단",
    ],
    "마케팅홍보": [
        "전시회","박람회","해외전시","전시참가","무역박람회",
        "홍보","브랜드","온라인마케팅","해외홍보","쇼룸",
        "국제전시","CES","MWC","전시출품","참가기업모집",
        "홍보영상","광고","SNS마케팅","디지털마케팅",
        "브랜드개발","CI","BI","패키지","디자인개발",
        "카탈로그","홍보물","콘텐츠","제품홍보",
    ],
    "인증특허": [
        "인증","특허","지식재산","KC인증","CE인증","ISO",
        "해외인증","IP","상표등록","국제인증","품질인증",
        "FDA","UL","해외규격","국제규격","인증취득",
        "특허출원","실용신안","디자인등록","상표권",
        "지식재산권","IP보호","기술보호","특허분석",
        "성능인증","NEP","NET","GR인증","녹색인증",
    ],
    "기술개발": [
        "기술개발","R&D","기술사업화","연구개발","기술혁신",
        "기술이전","실증","사업화","기술혁신개발","과제",
        "연구과제","개발과제","정부R&D","국가R&D",
        "기술혁신","제품개발","시제품","프로토타입",
        "실증사업","규제샌드박스","실증특례","혁신실증",
        "공동연구","산학연","기술협력","연구기관",
    ],
    "금융융자": [
        "융자","보증","정책자금","이차보전","경영안정",
        "투자","펀드","자금지원","신용보증",
        "저리융자","장기융자","운전자금","시설자금",
        "R&D자금","수출금융","무역금융","수출보험",
        "벤처투자","엔젤투자","크라우드펀딩","매칭펀드",
        "기술보증","신용대출","대출지원","금리우대",
    ],
    "내수판로": [
        "온라인몰","플랫폼","유통","알리바바","바이코리아",
        "내수판로","홈쇼핑","라이브커머스","B2B",
        "쿠팡","네이버쇼핑","카카오","오픈마켓","G마켓",
        "판로확대","국내판로","내수시장","유통채널",
        "대형마트","백화점","편의점","납품","구매상담",
        "온라인판매","이커머스","전자상거래",
    ],
    "인력채용": [
        "청년채용","고용","인재","채용","일자리","취업",
        "청년","인턴","현장실습","고용장려","채용지원",
        "인력양성","전문인력","R&D인력","핵심인력",
        "산업인력","기술인력","직무교육","재직자",
    ],
}

# 축3: 제품분야 역방향 매칭 (WALLA 15개 카테고리 → 공고 키워드)
# 기업이 선택한 제품분야 → 공고에 이 키워드가 있으면 업종 적합성 가산
INDUSTRY_KW = {
    "바이오·의료": [
        "바이오","의료","의료기기","헬스케어","제약","헬스","생명공학","의약",
        "진단","치료","의료용","임상","체외진단","의약품","바이오텍",
        "의료AI","디지털헬스","원격의료","웨어러블의료",
        # 동의어 확장
        "메디컬","헬스테크","의생명","생물의약","디지털치료제",
        "의료정보","병원정보","의료데이터","EMR","PHR",
        "재활","보조기기","의료소프트웨어","수술로봇","진단기기",
    ],
    "미래차·모빌리티": [
        "미래차","모빌리티","전기차","자율주행","수소차","친환경차","UAM",
        "퍼스널모빌리티","전동화","충전인프라","배터리","이차전지",
        "자동차부품","차량","운송","교통수단","커넥티드카",
        # 동의어 확장
        "EV","전동킥보드","PM","도심항공","전기버스","수소버스",
        "차량공유","카셰어링","라이드헤일링","MaaS","통합모빌리티",
        "ADAS","자동차SW","차량용반도체","OBC","BMS",
    ],
    "환경·에너지 기술": [
        "환경","에너지","탄소중립","친환경","온실가스","신재생","태양광",
        "수소","폐기물","재활용","탄소","녹색","저탄소","청정",
        "에너지효율","ESG","탄소저감","기후","풍력","연료전지",
        "환경오염","수처리","대기","토양정화","친환경소재",
        # 동의어 확장
        "탄소발자국","넷제로","RE100","그린뉴딜","순환경제",
        "폐기물처리","하수처리","대기오염","미세먼지","소음진동",
        "환경모니터링","환경측정","에너지저장","ESS","BESS",
        "태양전지","풍력발전","바이오에너지","지열","조력",
    ],
    "스마트 제조·산업·기계": [
        "스마트제조","스마트공장","제조혁신","산업기계","자동화","로봇",
        "공정혁신","MES","제조","기계","장비","설비","부품","소재",
        "정밀기계","산업용","제조업","공정","생산","품질관리",
        "CNC","공작기계","산업로봇","협동로봇","용접","도금",
        # 동의어 확장
        "DX전환","디지털공장","산업IoT","IIoT","예지보전",
        "품질검사","비전검사","비파괴검사","측정","계측",
        "금형","주조","단조","압출","표면처리","열처리",
        "PCB제조","반도체장비","디스플레이장비","이차전지장비",
    ],
    "스마트 건설": [
        "스마트건설","건설","건축","BIM","모듈러","건설기술","건설자재","플랜트",
        "건설장비","건설안전","건축자재","인프라건설","토목","시공",
        "건물","구조물","시설물","유지보수","안전진단","건설IT",
        # 동의어 확장
        "OSC","탈현장건설","3D프린팅건설","드론측량","건설드론",
        "시설관리","FM","AMO","건물에너지","BEMS","FEMS",
        "안전모니터링","작업자안전","건설현장안전","붕괴예측",
    ],
    "도시·교통 인프라": [
        "스마트시티","도시","교통","인프라","철도","도로","물류","스마트교통",
        "도시개발","대중교통","지하철","버스","항만","공항",
        "물류센터","SCM","공급망","유통인프라","스마트물류",
        # 동의어 확장
        "C-ITS","도로인프라","교통신호","주차관제","주차공유",
        "라스트마일","배송","풀필먼트","크로스도킹","콜드체인",
        "스마트항만","항만물류","선박물류","물류자동화","WMS",
    ],
    "방산·국방 기술": [
        "방산","국방","방위","군수","보안","방위산업","군","방어",
        "국방기술","방위기술","무기","군용","군사","사이버보안",
        "정보보안","보안솔루션","네트워크보안","물리보안",
        # 동의어 확장
        "ITAR","방산수출","국방R&D","ADD","국방조달",
        "보안인증","CC인증","ISMS","ISMS-P","보안감사",
        "침해대응","취약점","위협인텔리전스","SOC","SIEM",
    ],
    "정보통신 기술": [
        "ICT","정보통신","소프트웨어","앱","플랫폼","클라우드","SaaS","통신",
        "SW","IT","시스템","솔루션","네트워크","서버","데이터센터",
        "사물인터넷","IoT","5G","통신망","디지털","스마트",
        "보안","ERP","CRM","그룹웨어","SI",
        # 동의어 확장
        "API","마이크로서비스","컨테이너","쿠버네티스","DevOps",
        "엣지컴퓨팅","엣지","멀티클라우드","하이브리드클라우드",
        "RPA","업무자동화","디지털워크플레이스","협업툴",
        "6G","위성통신","저궤도위성","O-RAN","오픈랜",
    ],
    "AI·데이터 기술": [
        "AI","인공지능","빅데이터","데이터","머신러닝","딥러닝","디지털전환","DX",
        "자연어처리","컴퓨터비전","생성형AI","ChatGPT","LLM",
        "데이터분석","데이터플랫폼","MLOps","AI솔루션","지능형",
        "예측","추천","자동화AI","AI서비스",
        # 동의어 확장
        "파운데이션모델","멀티모달","RAG","AI에이전트","AI플랫폼",
        "데이터레이크","데이터웨어하우스","데이터마트","ETL",
        "시각AI","음성AI","영상분석","이상탐지","예측유지보수",
        "AI반도체","NPU","온디바이스AI","경량화모델",
    ],
    "농축산·수산·식품": [
        "농업","농식품","식품","수산","축산","스마트팜","농산물","푸드테크",
        "농기계","농촌","원예","작물","가공식품","식품안전",
        "수산물","수산양식","해양수산","축산물","낙농",
        # 동의어 확장
        "정밀농업","농업IoT","농업드론","수직농장","식물공장",
        "대체육","배양육","세포농업","푸드업사이클","기능성식품",
        "HACCP","식품인증","GMP식품","유기농","GAP",
        "어업","어선","어망","수산가공","냉동수산",
    ],
    "우주·항공·해양": [
        "우주","항공","드론","해양","선박","해운","위성",
        "UAM","도심항공","항공기","항공부품","발사체",
        "해양플랜트","해양기술","조선","선박부품","해양수산",
        # 동의어 확장
        "소형위성","큐브샛","우주부품","발사서비스","뉴스페이스",
        "MRO","항공정비","항공소재","항공SW","항공안전",
        "수중드론","AUV","ROV","수중로봇","해양로봇",
        "선박자율운항","자율선박","스마트쉽","친환경선박",
    ],
    "핀테크·금융 IT": [
        "핀테크","금융","결제","블록체인","암호화폐","보험테크",
        "디지털금융","오픈뱅킹","간편결제","송금","대출플랫폼",
        "자산관리","투자플랫폼","금융데이터","RegTech",
        # 동의어 확장
        "마이데이터","신용평가","대안신용","P2P금융","크라우드펀딩",
        "CBDC","STO","NFT","디지털자산","웹3",
        "보험금청구","자동심사","손해사정","보험데이터",
        "PG","VAN","전자지갑","선불카드","BNPL",
    ],
    "전기·전자": [
        "전기","전자","반도체","디스플레이","배터리","전력","회로","센서",
        "전장","전기부품","전자부품","LED","OLED","PCB",
        "전력반도체","시스템반도체","전력변환","전기설비",
        "계측기","검사장비","시험장비",
        # 동의어 확장
        "SiC","GaN","MLCC","MEMS","RF부품","안테나",
        "마이크로LED","MiniLED","플렉서블디스플레이","투명디스플레이",
        "사이니지","전광판","스마트사이니지","디지털사이니지",
        "스마트미터","AMI","전력IT","에너지관리","HEMS",
    ],
    "교육·HR테크": [
        "에듀테크","교육","HR","인재","학습","이러닝","채용플랫폼",
        "온라인교육","직무교육","기업교육","학습관리","LMS",
        "인적자원","인사관리","채용","HRD","HRM",
        # 동의어 확장
        "적응형학습","메타버스교육","XR교육","AI튜터","코딩교육",
        "평생학습","직업훈련","직업교육","자격증","역량평가",
        "피플애널리틱스","인사AI","채용AI","온보딩","퇴직관리",
    ],
    "콘텐츠·미디어·문화": [
        "콘텐츠","미디어","문화","게임","엔터","영상","음악","OTT",
        "웹툰","웹소설","애니메이션","VFX","CG","디지털콘텐츠",
        "메타버스","XR","VR","AR","MR","실감콘텐츠",
        "MCN","유튜브","SNS","인플루언서","소셜미디어",
        "광고","마케팅","브랜드","PR","디지털광고",
        "관광","여행","스마트관광","여가","레저","스포츠테크",
    ],
}

# 별점 판정 함수용 flat 리스트 (설정 탭 표시용)
DEFAULT_HIGH = (
    TARGET_KW["조달기업특화"] +
    TYPE_KW["공공조달"] +
    # 원스톱 스케일업 핵심 대상 추가 키워드
    [
        "혁신형","혁신성장","혁신창업","혁신기업",
        "수출기업","수출유망","수출강소","수출지원",
        "해외조달","해외공공조달","UN조달","글로벌조달",
        "혁신제품","혁신제품지정","G-PASS","MAS","우수조달",
        "공공조달","조달시장","조달기업","조달지원",
        "중소벤처","벤처기업","이노비즈","메인비즈",
    ]
)
DEFAULT_MID = (
    TARGET_KW["수출기업특화"] +
    TYPE_KW["해외진출"] +
    TYPE_KW["마케팅홍보"] +
    TYPE_KW["인증특허"] +
    # 원스톱 대상 기업 MID 추가
    [
        "수출바우처","해외판로","해외마케팅","바이어발굴",
        "해외전시","무역사절단","수출인큐베이터",
        "인증","특허","지식재산","해외인증","국제인증",
        "스케일업","Scale-up","성장기업","글로벌","판로개척",
    ]
)
REALM_CODE   = {
    "금융":"01","기술개발":"02","인력":"03","수출":"04",
    "내수":"05","창업":"06","경영":"07","기타":"09",
}
# 테스트 수신자 — 설정 탭에서 변경 가능 (session_state 우선)
_DEFAULT_TEST_RECIPIENTS = ["fbwlgns819@naver.com","fbwlgns819@kip.re.kr"]


def save_ai_analysis(drive):
    """ai_analysis session_state를 드라이브에 영구 저장"""
    data = st.session_state.get('ai_analysis', {})
    if not data:
        return True
    import json
    content = json.dumps(data, ensure_ascii=False, indent=None).encode('utf-8')
    return drive_upload(drive, AI_ANALYSIS_FILE, content, "application/json")

def load_ai_analysis(drive):
    """드라이브에서 ai_analysis를 복원 → session_state에 저장"""
    if 'ai_analysis' not in st.session_state:
        st.session_state['ai_analysis'] = {}
    saved = load_json(drive, AI_ANALYSIS_FILE)
    if saved and isinstance(saved, dict):
        # 기존 session_state와 병합 (새 분석이 우선)
        merged = {**saved, **st.session_state['ai_analysis']}
        st.session_state['ai_analysis'] = merged
        return len(saved)
    return 0


def reason_to_sentence(reason_str):
    """매칭근거 문자열을 메일용 자연스러운 한 줄 문장으로 변환"""
    if not reason_str: return ""
    parts = [p.strip() for p in reason_str.split("+")]
    targets, types, tags, industry, location = [], [], [], [], ""
    for p in parts:
        if "조달기업 대상" in p:    targets.append("조달")
        elif "수출기업 대상" in p:  targets.append("수출")
        elif "중소기업" in p:       targets.append("중소기업")
        if "공공조달" in p:         types.append("공공조달 진출")
        elif "해외진출" in p:       types.append("해외 판로 개척")
        elif "기술개발" in p:       types.append("기술개발")
        elif "마케팅" in p:         types.append("홍보·마케팅")
        elif "인증특허" in p or "인증·특허" in p: types.append("인증·특허")
        elif "금융" in p and "융자" in p:         types.append("자금 지원")
        elif "내수" in p:           types.append("내수 판로")
        if "수요태그(" in p:
            tags.append(p.split("(")[1].rstrip(")"))
        if "업종일치(" in p:
            industry = p.split("(")[1].rstrip(")")
        if "동일지역" in p: location = "지역 가점"
    parts_out = []
    if targets:
        parts_out.append("·".join(list(dict.fromkeys(targets))) + " 기업 대상")
    if types:
        mt = types[0] + (" 외 " + str(len(types)-1) + "개 지원" if len(types) > 1 else "")
        parts_out.append(mt)
    if tags:
        parts_out.append(", ".join(tags[:2]) + " 역량 활용 가능")
    if industry:
        parts_out.append(industry[:12] + " 분야 연관")
    if location:
        parts_out.append(location)
    return " · ".join(parts_out) if parts_out else reason_str[:40]


def get_test_recipients():
    saved = st.session_state.get('test_recipients_str','')
    if saved:
        return [e.strip() for e in saved.split(',') if e.strip()]
    return _DEFAULT_TEST_RECIPIENTS

# ── 구글 인증 ─────────────────────────────────────────
# ── 구글 인증 (requests 기반, httplib2 미사용) ──────────
def get_creds():
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GRequest
    SCOPES = [
        'https://www.googleapis.com/auth/gmail.send',
        'https://www.googleapis.com/auth/calendar',
        'https://www.googleapis.com/auth/drive',
    ]
    if 'g_creds' in st.session_state:
        creds = st.session_state['g_creds']
        if creds and not creds.expired:
            return creds
    if 'google' in st.secrets:
        creds = Credentials.from_authorized_user_info(
            json.loads(st.secrets['google']['token']), SCOPES)
    elif os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    else:
        return None
    if creds and creds.expired and creds.refresh_token:
        try: creds.refresh(GRequest())
        except: return None
    st.session_state['g_creds'] = creds
    return creds

def gapi(method, url, **kwargs):
    """인증된 구글 API 요청 (requests 직접 호출)"""
    creds = get_creds()
    if not creds:
        raise Exception("인증 실패")
    headers = kwargs.pop('headers', {})
    headers['Authorization'] = f'Bearer {creds.token}'
    return requests.request(method, url, headers=headers, **kwargs)

# ── gmail API (requests 기반) ──────────────────────────
def gmail_send(raw_b64):
    resp = gapi('POST',
        'https://gmail.googleapis.com/gmail/v1/users/me/messages/send',
        json={'raw': raw_b64})
    resp.raise_for_status()
    return resp.json()

# ── calendar API (requests 기반) ───────────────────────
def cal_list_events(cal_id, private_prop=None):
    params = {'maxResults': 10}
    if private_prop:
        params['privateExtendedProperty'] = private_prop
    resp = gapi('GET',
        f'https://www.googleapis.com/calendar/v3/calendars/{cal_id}/events',
        params=params)
    return resp.json() if resp.ok else {'items': []}

def cal_insert_event(cal_id, body):
    resp = gapi('POST',
        f'https://www.googleapis.com/calendar/v3/calendars/{cal_id}/events',
        json=body)
    return resp.json() if resp.ok else {}

def cal_create(summary, description):
    """캘린더 생성 → cal_id 반환"""
    resp = gapi('POST',
        'https://www.googleapis.com/calendar/v3/calendars',
        json={'summary': summary, 'description': description, 'timeZone': 'Asia/Seoul'})
    return resp.json().get('id', '') if resp.ok else ''

def cal_share(cal_id, email):
    """캘린더를 특정 이메일(reader)로 공유"""
    resp = gapi('POST',
        f'https://www.googleapis.com/calendar/v3/calendars/{cal_id}/acl',
        json={'scope': {'type': 'user', 'value': email}, 'role': 'reader'})
    return resp.ok

# ── drive API (requests 기반) ──────────────────────────
def drive_list_files(name, folder_id):
    params = {
        'q': f"name='{name}' and '{folder_id}' in parents and trashed=false",
        'fields': 'files(id,name)',
        'orderBy': 'modifiedTime desc',
    }
    resp = gapi('GET', 'https://www.googleapis.com/drive/v3/files', params=params)
    return resp.json().get('files', []) if resp.ok else []

def drive_download_file(file_id):
    resp = gapi('GET',
        f'https://www.googleapis.com/drive/v3/files/{file_id}',
        params={'alt': 'media'})
    return resp.content if resp.ok else None

def drive_upload_file(name, folder_id, content_bytes, mime, file_id=None):
    bnd  = b"----MIMEBoundary"
    crlf = b"\r\n"
    ct   = "multipart/related; boundary=----MIMEBoundary"

    if file_id:
        # 기존 파일 업데이트: parents 제외, 내용만 교체
        meta = json.dumps({"name": name}).encode()
        body = b"--" + bnd + crlf
        body += b"Content-Type: application/json; charset=UTF-8" + crlf + crlf
        body += meta + crlf
        body += b"--" + bnd + crlf
        body += b"Content-Type: " + mime.encode() + crlf + crlf
        body += content_bytes + crlf
        body += b"--" + bnd + b"--"
        resp = gapi("PATCH",
            f"https://www.googleapis.com/upload/drive/v3/files/{file_id}",
            params={"uploadType": "multipart"},
            data=body,
            headers={"Content-Type": ct})
    else:
        # 신규 파일 생성: parents 포함
        meta = json.dumps({"name": name, "parents": [folder_id]}).encode()
        body = b"--" + bnd + crlf
        body += b"Content-Type: application/json; charset=UTF-8" + crlf + crlf
        body += meta + crlf
        body += b"--" + bnd + crlf
        body += b"Content-Type: " + mime.encode() + crlf + crlf
        body += content_bytes + crlf
        body += b"--" + bnd + b"--"
        resp = gapi("POST",
            "https://www.googleapis.com/upload/drive/v3/files",
            params={"uploadType": "multipart"},
            data=body,
            headers={"Content-Type": ct})

    if not resp.ok:
        raise Exception(f"HTTP {resp.status_code}: {resp.text[:200]}")
    return True

def get_services(): return None, None, None
def _get_drive():   return 'drive'
def _get_gmail():   return 'gmail'
def _get_cal():     return 'cal' 

# ── 드라이브 유틸 ─────────────────────────────────────
def drive_file_id(drive, filename):
    files = drive_list_files(filename, DRIVE_FOLDER_ID)
    return files[0]['id'] if files else None

def drive_download(drive, filename):
    fid = drive_file_id(drive, filename)
    return drive_download_file(fid) if fid else None

def drive_upload(drive, filename, content_bytes, mime):
    try:
        fid = drive_file_id(drive, filename)
        result = drive_upload_file(filename, DRIVE_FOLDER_ID, content_bytes, mime, fid)
        return result
    except Exception as e:
        st.error(f"드라이브 저장 실패 ({filename}): {e}")
        return False

def load_excel(drive, filename):
    content = drive_download(drive, filename)
    if content:
        try:
            df = pd.read_excel(io.BytesIO(content), dtype=str).fillna("")
            # 선정기업 명단인데 '기업명' 컬럼이 없으면
            # 1행이 필수/운영관리 구분행 → header=1로 재시도
            if filename == SELECTED_FILE and '기업명' not in df.columns:
                df = pd.read_excel(io.BytesIO(content), header=1, dtype=str).fillna("")
            return df
        except: return pd.DataFrame()
    return pd.DataFrame()

def save_excel(drive, df, filename, sheet="데이터", hcolor="1F4E79", star_col=None):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active
    ws.title = sheet; ws.sheet_view.showGridLines = False
    s   = Side(style="thin", color="BFBFBF")
    bdr = Border(left=s, right=s, top=s, bottom=s)
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(str(col))*2, 14)
        c = ws.cell(row=1, column=ci, value=col)
        c.fill=PatternFill("solid",start_color=hcolor,end_color=hcolor)
        c.font=Font(name="맑은 고딕",bold=True,color="FFFFFF",size=10)
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border=bdr
    ws.row_dimensions[1].height = 24
    star_colors = {"★★★":"FFF2CC","★★":"E2EFDA"}
    rev_colors  = {"○":"D5E8D4","✕":"FFE6E6"}
    for ri, row in enumerate(df.itertuples(index=False), 2):
        vals = list(row)
        bg   = "FFFFFF"
        if star_col and star_col in df.columns:
            bg = star_colors.get(vals[df.columns.tolist().index(star_col)], "FFFFFF")
        ws.row_dimensions[ri].height = 18
        for ci, val in enumerate(vals, 1):
            cname    = df.columns[ci-1]
            cell_bg  = rev_colors.get(str(val), bg) if cname=="담당자검토" else bg
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill=PatternFill("solid",start_color=cell_bg,end_color=cell_bg)
            c.font=Font(name="맑은 고딕",size=9)
            c.alignment=Alignment(horizontal="left",vertical="center")
            c.border=bdr
    buf = io.BytesIO(); wb.save(buf)
    return drive_upload(drive, filename, buf.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def load_json(drive, filename):
    content = drive_download(drive, filename)
    if content:
        try: return json.loads(content.decode('utf-8'))
        except: return {}
    return {}

def save_json(drive, data, filename):
    return drive_upload(drive, filename,
        json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'),
        "application/json")

def load_text(drive, filename):
    content = drive_download(drive, filename)
    return content.decode('utf-8').strip() if content else ""

def load_keywords(drive):
    kw = load_json(drive, KEYWORDS_FILE)
    global TARGET_KW, TYPE_KW
    if "TARGET_KW" in kw: TARGET_KW = kw["TARGET_KW"]
    if "TYPE_KW"   in kw: TYPE_KW   = kw["TYPE_KW"]
    return kw.get("HIGH", DEFAULT_HIGH), kw.get("MID", DEFAULT_MID)

def load_kw_config(drive):
    """가중치, 제외키워드, 피드백 등 전체 키워드 설정 로드"""
    kw = load_json(drive, KEYWORDS_FILE) or {}
    weights = kw.get("weights", {
        "지원대상": 3, "사업성격": 2, "기업키워드": 2,
        "핵심수요": 3, "업종역방향": 1, "소재지가산": 3, "세그먼트": 2
    })
    return {
        "weights":          weights,
        "exclude_keywords": kw.get("exclude_keywords", []),
        "feedback":         kw.get("feedback", {}),
        "star3_threshold":  kw.get("star3_threshold", 20),
        "star2_threshold":  kw.get("star2_threshold", 10),
    }

# ── 유틸 ─────────────────────────────────────────────
def strip_html(html):
    return re.sub(r'<[^>]+>', ' ', html or '').strip()

def parse_deadline(s):
    try:
        end = s.split('~')[-1].strip()
        return datetime.strptime(re.sub(r'\.', '-', end), "%Y-%m-%d")
    except: return None

def score_notice(notice, row, already_sent, HIGH, MID, feedback=None, kw_config=None):
    pid = notice.get('pblancId', '')
    if (row['기업명'], pid) in already_sent: return None
    dl = notice.get('마감일', '')
    if dl and dl < datetime.today().strftime("%Y-%m-%d"): return None
    if str(row.get('수출실적',''))=='아니오' and '수출' in str(notice.get('분야','')): return None

    text = " ".join([str(notice.get(k,'')) for k in ['공고명','사업개요','전문내용','해시태그','주관기관','지원대상']])

    # ── 제외 키워드 필터 ──────────────────────────────
    if kw_config:
        for excl_kw in kw_config.get("exclude_keywords", []):
            if excl_kw and excl_kw in text:
                return None  # 제외 키워드 포함 공고 즉시 제거

    # 가중치 설정 (없으면 기본값)
    W = kw_config.get("weights", {}) if kw_config else {}
    W_TARGET = W.get("지원대상", 3)
    W_TYPE   = W.get("사업성격", 2)
    W_KW     = W.get("기업키워드", 2)
    W_DEMAND = W.get("핵심수요", 3)
    W_IND    = W.get("업종역방향", 1)
    W_LOC    = W.get("소재지가산", 3)
    W_SEG    = W.get("세그먼트", 4)

    # ── 강화된 부정 필터 — 자격 미달이 명확한 공고 사전 제거 ──
    # 업력 제한: 창업 초기 공고에 오래된 기업 매칭 방지
    biz_start = str(row.get('설립연도', row.get('창업연도', '')))
    if biz_start and biz_start.isdigit():
        biz_age = datetime.today().year - int(biz_start)
        # 공고에 "창업 7년 이내" 등 업력 제한이 명시된 경우
        age_patterns = re.findall(r'창업\s*(\d+)년\s*이내', text)
        for ap in age_patterns:
            if biz_age > int(ap):
                return None  # 업력 초과 → 제외
        startup_only = any(p in text for p in ['예비창업자', '창업팀', '예비 창업'])
        if startup_only:
            return None  # 예비창업자 전용 공고

    # 매출 규모 제한 (공고에 "매출 N억 이하" 명시 시 대형 기업 제외)
    revenue_str = str(row.get('매출규모', row.get('매출액', '')))
    if revenue_str and '억' in revenue_str:
        try:
            rev_amt = float(re.search(r'([\d.]+)', revenue_str).group(1))
            revenue_limits = re.findall(r'매출액?\s*(\d+)억\s*원?\s*이하', text)
            for rl in revenue_limits:
                if rev_amt > float(rl) * 1.5:  # 1.5배 여유 허용
                    return None
        except Exception:
            pass

    # ── 기업 세그먼트 자동 분류 ───────────────────────────
    # 기업 프로파일에 따라 3개 세그먼트로 분류 → 세그먼트별 점수 보정
    export_yn  = str(row.get('수출실적','')) not in ['아니오','','nan']
    trl_str    = str(row.get('TRL단계',''))
    trl_high   = any(t in trl_str for t in ['8','9'])
    trl_low    = any(t in trl_str for t in ['4','5','6'])
    demand     = str(row.get('핵심수요태그',''))
    has_procurement_tag = any(k in demand for k in ['혁신제품','G-PASS','우수조달','시범구매','MAS'])
    has_export_tag      = any(k in demand for k in ['해외조달','수출바우처','해외전시','해외진출'])
    has_rd_tag          = any(k in demand for k in ['기술사업화','특허','R&D','기술개발'])

    # 세그먼트 판정 (우선순위 순)
    if has_procurement_tag or (trl_high and not export_yn):
        segment = 'procurement'   # 조달강화형: 혁신제품·G-PASS 보유, 국내 조달 집중
    elif export_yn or has_export_tag:
        segment = 'export'        # 해외진출형: 수출 실적 있거나 해외 수요 명확
    elif trl_low or has_rd_tag:
        segment = 'rd'            # 기술개발형: TRL 낮거나 기술사업화 수요
    else:
        segment = 'general'       # 일반형

    # 세그먼트별 사업성격 점수 보정값 (매칭 시 type 카테고리에 가산)
    SEGMENT_BOOST = {
        'procurement': {'공공조달': 4, '해외진출': 2, '기술개발': 0, '금융융자': 0},
        'export':      {'해외진출': 4, '마케팅홍보': 2, '공공조달': 1, '인증특허': 1},
        'rd':          {'기술개발': 4, '인증특허': 2, '금융융자': 2, '공공조달': 0},
        'general':     {'공공조달': 1, '해외진출': 1, '기술개발': 1, '금융융자': 1},
    }
    seg_boost = SEGMENT_BOOST.get(segment, {})

    # ── TRL 필터: 8-9단계 기업은 R&D 공고 제외 ──────
    trl = str(row.get('TRL단계',''))
    if any(t in trl for t in ['8','9']):
        if any(kw in text for kw in ['R&D','연구개발','기술개발과제','기초연구','원천기술']):
            return None

    # ── 소재지 필터: 지역 공고 판단 ─────────────────
    location = str(row.get('소재지',''))
    location_score = 0
    notice_region_tag = ''   # 공고에서 감지된 지역명 (검토 화면 표시용)

    REGIONS = ['서울','부산','대구','인천','광주','대전','울산','세종',
               '경기','강원','충북','충남','전북','전남','경북','경남','제주']
    REGION_ALIAS = {
        '서울': ['서울','수도권'],
        '경기': ['경기','수도권'],
        '인천': ['인천','수도권'],
        '부산': ['부산'],
        '대구': ['대구'],
        '광주': ['광주'],
        '대전': ['대전','충청'],
        '울산': ['울산'],
        '세종': ['세종'],
        '강원': ['강원'],
        '충북': ['충북','충청'],
        '충남': ['충남','충청'],
        '전북': ['전북','전라'],
        '전남': ['전남','전라'],
        '경북': ['경북'],
        '경남': ['경남'],
        '제주': ['제주'],
    }

    if location and location != 'nan':
        notice_name = str(notice.get('공고명',''))
        notice_full  = " ".join([
            notice_name,
            str(notice.get('사업개요','')),
            str(notice.get('지원대상','')),
            str(notice.get('해시태그','')),
            str(notice.get('전문내용','')),  # 전문내용도 소재지 판단에 반영
        ])
        co_region = next((r for r in REGIONS if r in location), None)

        import re as _re

        # ① 공고명 [] 패턴 최우선
        bracket_match = _re.search(r'\[([^\]]+)\]', notice_name)
        if bracket_match:
            bracket_text = bracket_match.group(1)
            bracket_regions = [r for r in REGIONS if r in bracket_text]
            if bracket_regions:
                notice_region_tag = f"[{bracket_text}]"
                if co_region:
                    aliases = REGION_ALIAS.get(co_region, [co_region])
                    if any(a in bracket_text for a in aliases):
                        location_score = 3
                    else:
                        location_score = -5

        # ② 주관기관이 지방자치단체
        elif co_region:
            organizer = str(notice.get('주관기관',''))
            org_regions = [r for r in REGIONS if r in organizer]
            if org_regions:
                notice_region_tag = f"주관기관({organizer[:10]})"
                aliases = REGION_ALIAS.get(co_region, [co_region])
                if any(a in organizer for a in aliases):
                    location_score = 2
                else:
                    location_score = -5  # -3 → -5 강화

        # ③ 전문내용 포함 전체 텍스트에서 "○○ 소재 기업" 패턴
        elif co_region:
            if '전국' not in notice_full and '전 지역' not in notice_full:
                sojaepat = _re.findall(r'([가-힣]{2,4}(?:도|시|군|구))\s*소재', notice_full)
                if sojaepat:
                    notice_region_tag = ", ".join(sojaepat[:2])
                    aliases = REGION_ALIAS.get(co_region, [co_region])
                    if any(any(a in s for a in aliases) for s in sojaepat):
                        location_score = 2
                    else:
                        location_score = -5  # -3 → -5 강화

                # ④ 신규: 전문에서 지역 직접 언급 패턴
                # "광주 소재", "경남 지역", "○○시에 위치한" 등
                if location_score == 0:
                    region_mentions = [r for r in REGIONS
                                       if _re.search(rf'{r}.{{0,6}}(?:소재|지역|소속|위치|관내|내)', notice_full)]
                    if region_mentions:
                        notice_region_tag = ", ".join(region_mentions[:2])
                        aliases = REGION_ALIAS.get(co_region, [co_region])
                        if any(any(a in rm for a in aliases) for rm in region_mentions):
                            location_score = 2
                        else:
                            location_score = -3

    # ── 축1: 지원대상 매칭 ────────────────────────────
    matched_target = {}
    for cat, kws in TARGET_KW.items():
        hits = [kw for kw in kws if kw in text]
        if hits: matched_target[cat] = hits

    # ── 축2: 사업성격 매칭 ────────────────────────────
    matched_type = {}
    for cat, kws in TYPE_KW.items():
        hits = [kw for kw in kws if kw in text]
        if hits: matched_type[cat] = hits

    # ── 기업 키워드 + 핵심수요태그 매칭 ────────────────
    raw = ",".join([str(row.get(k,'')) for k in ['기술키워드','제품분야','키워드보완','핵심수요태그']])
    co_kws = [k.strip() for k in raw.split(',') if k.strip() and k.strip()!='nan']
    matched_co = [kw for kw in co_kws if kw in text]

    # ── 기술키워드 역매핑 확장: 기업 키워드 → INDUSTRY_KW 카테고리 → 공고 키워드 검색 ──
    # 너무 일반적인 단어는 역매핑에서 제외 (거의 모든 공고에 등장해 변별력 없음)
    GENERIC_WORDS = {
        '시스템','솔루션','서비스','플랫폼','기술','개발','소프트웨어','SW',
        'IT','ICT','디지털','스마트','데이터','네트워크','클라우드','AI',
        '관리','지원','운영','구축','통합','연계','자동화','혁신',
    }
    tech_kws = [k.strip() for k in str(row.get('기술키워드','')).split(',') if k.strip() and k.strip()!='nan']
    expanded_industry_hits = []
    for tech_kw in tech_kws:
        kw_lower = tech_kw.lower().replace(' ', '')
        if len(kw_lower) < 3:
            continue  # 너무 짧은 키워드는 오매칭 위험
        for ind_cat, ind_kws in INDUSTRY_KW.items():
            matched_cat = False
            for ind_kw in ind_kws:
                ind_lower = ind_kw.lower().replace(' ', '')
                if len(ind_lower) >= 3 and (ind_lower in kw_lower or kw_lower in ind_lower):
                    matched_cat = True
                    break
            if matched_cat:
                # 일반 단어 제외 후 카테고리 키워드만 추출
                cat_hits = [
                    k for k in ind_kws
                    if k in text and len(k) >= 2 and k not in GENERIC_WORDS
                ]
                for h in cat_hits[:2]:
                    if h not in expanded_industry_hits:
                        expanded_industry_hits.append(h)

    for h in expanded_industry_hits:
        if h not in matched_co:
            matched_co.append(h)

    # 핵심수요태그는 가중치 높게
    demand_tags = [t.strip() for t in str(row.get('핵심수요태그','')).split(',') if t.strip() and t.strip()!='nan']
    matched_demand = [kw for kw in demand_tags if kw in text]

    # ── 축3: 제품분야 역방향 매칭 ────────────────────
    # 기업 제품분야 → 해당 업종 키워드가 공고에 있으면 업종 적합성 확인
    co_industry    = str(row.get('제품분야',''))
    matched_industry = []
    for ind_cat, ind_kws in INDUSTRY_KW.items():
        # 기업이 이 카테고리를 선택했는지 확인
        if any(cat_part in co_industry for cat_part in ind_cat.split('·')):
            # 공고에 해당 업종 키워드가 있으면 매칭
            hits = [kw for kw in ind_kws if kw in text]
            matched_industry.extend(hits)
    # 역방향 매칭 점수 (업종 적합성 가산)
    ind_score = min(len(matched_industry) * 2, 6)  # 최대 6점 상한

    # ── 수출국가 가산 ─────────────────────────────────
    cn = str(row.get('수출국가',''))
    xs = 2 if (cn and cn!='nan' and cn in text) else 0

    # ── 역매칭: 특정 업종 명시 공고에서 기업 키워드 없으면 제외 ──
    INDUSTRY_SPECIFIC = [
        '농식품','농업','식품','농산물','수산','임업',
        '의료','바이오','제약','화장품','뷰티',
        '건설','건축','토목','부동산',
        '관광','숙박','외식','요식업',
        '섬유','패션','의류',
    ]
    for ind_kw in INDUSTRY_SPECIFIC:
        if ind_kw in text and not any(
            ind_kw in kw or kw in ind_kw for kw in co_kws + demand_tags
        ):
            return None

    # ── 점수 계산 ────────────────────────────────────
    seg_score = sum(min(seg_boost.get(cat, 0), W_SEG) for cat in matched_type.keys())

    feedback_penalty = 0
    fb = feedback or (kw_config.get("feedback", {}) if kw_config else {})
    if fb:
        co_feedback = fb.get(row.get("기업명",""), {})
        for cat in matched_type.keys():
            reject_count = co_feedback.get(cat, 0)
            if reject_count >= 3:   feedback_penalty -= 6
            elif reject_count >= 2: feedback_penalty -= 3

    score = (
        len(sum(matched_target.values(), [])) * W_TARGET +
        len(sum(matched_type.values(),   [])) * W_TYPE   +
        len(matched_co)     * W_KW    +
        len(matched_demand) * W_DEMAND +
        ind_score * W_IND   +
        xs + (location_score * W_LOC // 3) + seg_score + feedback_penalty
    )
    if matched_demand: score += 8   # 핵심수요 직접매칭 강력 보너스
    if score <= 0: return None

    # ── ④ 별점 판정 — 점수 기반 통합 ──────────────────
    star3_threshold = kw_config.get("star3_threshold", 18) if kw_config else 18
    star2_threshold = kw_config.get("star2_threshold",  8) if kw_config else 8

    hard_star3 = (
        bool(matched_demand)
        or ("조달기업특화" in matched_target and "공공조달" in matched_type)
        or ("조달기업특화" in matched_target and "해외진출" in matched_type)
        or ("수출기업특화" in matched_target and "해외진출" in matched_type)
        or (bool(matched_co) and "공공조달" in matched_type)
    )

    if hard_star3 or score >= star3_threshold:
        stars = "★★★"
    elif score >= star2_threshold:
        stars = "★★"
    else:
        return None

    # 소재지 불일치 시 별점 강등
    if location_score < 0:
        if stars == "★★★":   stars = "★★"
        elif stars == "★★":  stars = "★"
    if stars == "★": return None


    # ── 매칭 근거 텍스트 ──────────────────────────────
    target_str = " / ".join([f"{k}({','.join(v)})" for k,v in matched_target.items()])
    type_str   = " / ".join([f"{k}({','.join(v)})" for k,v in matched_type.items()])

    return {
        "기업명":       row['기업명'],
        "관련도":       stars,
        "점수":         score,
        "세그먼트":     segment,
        "공고ID":       pid,
        "공고명":       notice.get('공고명',''),
        "주관기관":     notice.get('주관기관',''),
        "접수기간":     notice.get('접수기간',''),
        "지원대상":     notice.get('지원대상',''),
        "마감일":       dl,
        "사업개요":     str(notice.get('사업개요',''))[:300]+"...",
        "지원대상매칭": target_str,
        "사업성격매칭": type_str,
        "기업키워드매칭": ", ".join(matched_co),
        "핵심수요매칭":   ", ".join(matched_demand),
        "업종역방향매칭": ", ".join(matched_industry[:5]),
        "소재지점수":     location_score,
        "공고지역":     notice_region_tag,
        "매칭근거":     _build_reason(stars, matched_target, matched_type, matched_co,
                                matched_demand, location_score, matched_industry),
        "공고링크":     notice.get('공고링크',''),
        "공고유형":     "맞춤",   # 매칭 실행 시 공통/맞춤 재분류됨
        "담당자검토":   "",
        "검토의견":     "",
    }

# Claude API 비용 상수 (claude-sonnet-4-5 기준)
CLAUDE_INPUT_COST  = 3.0 / 1_000_000   # $3 per 1M input tokens
CLAUDE_OUTPUT_COST = 15.0 / 1_000_000  # $15 per 1M output tokens
USD_TO_KRW         = 1380

def estimate_cost(n_notices):
    """n건 분석 시 예상 비용 계산"""
    avg_input  = 800   # 토큰 (기업정보 300 + 공고내용 500)
    avg_output = 300   # 토큰 (JSON 응답)
    total_usd  = n_notices * (avg_input * CLAUDE_INPUT_COST + avg_output * CLAUDE_OUTPUT_COST)
    total_krw  = total_usd * USD_TO_KRW
    return total_usd, total_krw

def claude_call_raw(prompt, max_tokens=1000):
    """단순 텍스트 프롬프트 → Claude 응답 문자열 반환"""
    api_key = ""
    try:
        api_key = st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            pass
    if not api_key:
        st.warning("⚠️ ANTHROPIC_API_KEY 없음")
        return ''
    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01"
    }
    payload = {
        "model": "claude-sonnet-4-6",
        "max_tokens": max_tokens,
        "messages": [{"role": "user", "content": prompt}]
    }
    try:
        resp = requests.post("https://api.anthropic.com/v1/messages",
                             headers=headers, json=payload, timeout=30)
        if resp.ok:
            content = resp.json().get('content', [])
            return content[0].get('text', '') if content else ''
        else:
            st.warning(f"⚠️ Claude API 오류 {resp.status_code}: {resp.text[:100]}")
    except Exception as e:
        st.warning(f"⚠️ Claude API 예외: {e}")
    return ''


def claude_analyze(company_info, notice_info):
    """Claude API로 공고-기업 적합성 분석"""
    api_key = ""
    try:
        api_key = st.secrets["ANTHROPIC_API_KEY"]
    except:
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except:
            pass

    if not api_key:
        try:
            available = list(st.secrets.keys())
        except:
            available = []
        return {"error": f"API 키 없음 — Secrets 키 목록: {available}"}

    # 전문 내용 우선 활용
    notice_content = (
        notice_info.get('전문내용','') or
        notice_info.get('사업개요','')
    )[:1500]

    prompt = f"""당신은 정부 지원사업 매칭 전문가입니다.
아래 기업 정보와 공고를 보고 이 기업이 이 공고에 지원하는 게 적합한지 판단하세요.

## 기업 정보
- 기업명: {company_info.get('기업명','')}
- 소재지: {company_info.get('소재지','')}
- 기업유형: {company_info.get('기업유형','')}
- 관심분야: {company_info.get('관심사업분야','')}
- 제품/기술 분야: {company_info.get('제품분야','')}
- 기술키워드: {company_info.get('기술키워드','')}
- 핵심수요: {company_info.get('핵심수요태그','')}
- 희망서비스: {company_info.get('희망서비스요약','')}
- 수출실적: {company_info.get('수출실적','')} | 수출국가: {company_info.get('수출국가','')}
- TRL단계: {company_info.get('TRL단계','')}
- 매출규모: {company_info.get('매출규모','')}

## 공고 정보
- 공고명: {notice_info.get('공고명','')}
- 주관기관: {notice_info.get('주관기관','')}
- 지원대상: {notice_info.get('지원대상','')}
- 지원금액: {notice_info.get('지원금액','')}
- 선정규모: {notice_info.get('선정규모','')}
- 지역제한: {notice_info.get('공고지역','') or '전국'}
- 공고내용: {notice_content}

## 판단 항목 (각각 구체적으로)
1. 업종일치: 기업 제품/기술이 이 공고가 원하는 업종/분야와 맞는가
2. 자격충족: 지원 자격(기업유형·매출·수출실적 등) 충족 가능성
3. 지역적합: 소재지와 공고 지역이 맞는가
4. 수요일치: 기업이 원하는 지원(핵심수요태그)과 공고 내용이 맞는가
5. 경쟁력: 이 기업이 선정될 가능성이 있는가

JSON 형식으로만 답하세요:
{{
  "추천여부": "추천 또는 검토 또는 비추천",
  "적합도": "높음 또는 보통 또는 낮음",
  "한줄요약": "15자 이내 핵심",
  "업종일치": "O 또는 X 또는 △",
  "자격충족": "O 또는 X 또는 △",
  "지역적합": "O 또는 X 또는 △",
  "수요일치": "O 또는 X 또는 △",
  "판단근거": "3~4문장으로 구체적 근거 (기업의 어떤 특성이 공고의 어떤 조건과 맞거나 안 맞는지)",
  "주의사항": "신청 전 반드시 확인할 사항 (없으면 없음)"
}}"""

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-5",
                "max_tokens": 600,
                "messages": [{"role": "user", "content": prompt}]
            },
            timeout=30
        )
        if resp.ok:
            text = resp.json()['content'][0]['text']
            import re as _re
            json_match = _re.search(r'[{].*[}]', text, _re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
            return {"error": "응답 파싱 실패", "raw": text[:200]}
        else:
            try:
                err_detail = resp.json().get('error', {}).get('message', resp.text[:300])
            except:
                err_detail = resp.text[:300]
            return {"error": f"API 오류 {resp.status_code}: {err_detail}"}
    except Exception as e:
        return {"error": str(e)}

def _build_reason(stars, matched_target, matched_type, matched_co,
                  matched_demand=[], location_score=0, matched_industry=[]):
    parts = []
    if '조달기업특화' in matched_target:
        parts.append("조달기업 대상")
    if '수출기업특화' in matched_target:
        parts.append("수출기업 대상")
    if '공공조달' in matched_type:
        parts.append("공공조달 연계")
    if '해외진출' in matched_type:
        parts.append("해외진출 지원")
    if '마케팅홍보' in matched_type:
        parts.append("마케팅·홍보")
    if '인증특허' in matched_type:
        parts.append("인증·특허 지원")
    if '기술개발' in matched_type:
        parts.append("기술개발 지원")
    if '금융융자' in matched_type:
        parts.append("금융·융자 지원")
    if matched_demand:
        parts.append(f"수요태그({', '.join(matched_demand[:2])})")
    if matched_industry:
        parts.append(f"업종일치({', '.join(matched_industry[:2])})")
    if matched_co and not matched_demand:
        parts.append(f"키워드({', '.join(matched_co[:2])})")
    if location_score > 0:
        parts.append("✓ 동일지역")
    elif location_score < 0:
        parts.append("△ 타지역")
    return " + ".join(parts) if parts else "키워드 매칭"

# ── 안내 박스 ─────────────────────────────────────────
def info_box(title, desc, how_to=None):
    with st.expander(f"ℹ️ {title}", expanded=False):
        st.markdown(desc)
        if how_to:
            st.divider()
            st.markdown("**✏️ 수정 방법**")
            st.markdown(how_to)

# ── 앱 초기화 — AI 분석 결과 드라이브에서 복원 ──────
if 'ai_analysis_loaded' not in st.session_state:
    try:
        _drive_init = _get_drive()
        _loaded = load_ai_analysis(_drive_init)
        st.session_state['ai_analysis_loaded'] = True
        if _loaded > 0:
            pass  # 조용히 복원 (배너 표시 안 함)
    except Exception:
        st.session_state['ai_analysis_loaded'] = True  # 실패해도 반복 시도 안 함

# ── CSS ──────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

:root {
  --accent:       #10B981;
  --accent-light: #ECFDF5;
  --accent-dark:  #059669;
  --border:       #E2E8F0;
  --border-2:     #CBD5E1;
  --surface:      #FFFFFF;
  --bg:           #F8FAFC;
  --text-1:       #0F172A;
  --text-2:       #475569;
  --text-3:       #94A3B8;
  --blue:         #3B82F6;
  --yellow:       #F59E0B;
  --red:          #EF4444;
  --radius:       8px;
  --shadow:       0 1px 3px rgba(0,0,0,0.08);
  --shadow-md:    0 4px 12px rgba(0,0,0,0.08);
}

/* ── 전체 폰트 ── */
.stApp {
  font-family: 'Inter','Apple SD Gothic Neo','Malgun Gothic',sans-serif !important;
  background: var(--bg) !important;
}

/* ── 타이포 ── */
h1 { font-size:22px !important; font-weight:700 !important; color:var(--text-1) !important; letter-spacing:-0.3px !important; }
h2 { font-size:17px !important; font-weight:600 !important; color:var(--text-1) !important; }
h3 { font-size:15px !important; font-weight:600 !important; color:var(--text-1) !important; }
strong { color:var(--accent-dark) !important; font-weight:600 !important; }
code {
  color:var(--accent-dark) !important;
  background:var(--accent-light) !important;
  border-radius:4px !important; padding:1px 6px !important; font-size:12px !important;
}

/* ── 메트릭 카드 ── */
[data-testid="metric-container"] {
  background: var(--surface) !important;
  border: 1px solid var(--border) !important;
  border-radius: var(--radius) !important;
  padding: 14px 18px !important;
  box-shadow: var(--shadow) !important;
}
[data-testid="stMetricLabel"] {
  font-size: 11px !important; font-weight:500 !important;
  text-transform: uppercase !important; letter-spacing:0.5px !important;
  color: var(--text-2) !important;
}
[data-testid="stMetricValue"] { font-size:24px !important; font-weight:700 !important; }

/* ── primary 버튼 ── */
button[kind="primary"] {
  background: var(--accent) !important;
  color: #fff !important; border:none !important; font-weight:600 !important;
  box-shadow: 0 1px 3px rgba(16,185,129,0.3) !important;
}
button[kind="primary"]:hover { background: var(--accent-dark) !important; }

/* ── 일반 버튼 ── */
.stButton button {
  border-radius: var(--radius) !important;
  font-size:13px !important; font-weight:500 !important;
  border: 1px solid var(--border-2) !important;
  transition: all 0.15s !important;
}
.stButton button:hover { border-color: var(--accent) !important; color: var(--accent-dark) !important; }

/* ── 입력창 포커스 ── */
.stTextInput input:focus, .stTextArea textarea:focus {
  border-color: var(--accent) !important;
  box-shadow: 0 0 0 3px rgba(16,185,129,0.15) !important;
}

/* ── 탭 ── */
.stTabs [data-baseweb="tab-list"] {
  border-bottom: 2px solid var(--border) !important;
}
.stTabs [aria-selected="true"] {
  color: var(--accent-dark) !important;
  font-weight:600 !important;
  border-bottom: 2px solid var(--accent) !important;
  background: var(--accent-light) !important;
}

/* ── 알림 박스 ── */
.stSuccess, .stSuccess * { color: #065F46 !important; }
.stSuccess { background: #ECFDF5 !important; border-left: 3px solid var(--accent) !important; border-radius:var(--radius) !important; }
.stWarning, .stWarning * { color: #78350F !important; }
.stWarning { background: #FFFBEB !important; border-left: 3px solid var(--yellow) !important; border-radius:var(--radius) !important; }
.stError, .stError * { color: #7F1D1D !important; }
.stError { background: #FEF2F2 !important; border-left: 3px solid var(--red) !important; border-radius:var(--radius) !important; }
.stInfo, .stInfo * { color: #1E3A8A !important; }
.stInfo { background: #EFF6FF !important; border-left: 3px solid var(--blue) !important; border-radius:var(--radius) !important; }

/* ── 구분선 ── */
hr { border: none !important; border-top: 1px solid var(--border) !important; }

/* ── 캡션 ── */
.stCaption, [data-testid="stCaptionContainer"] { color: var(--text-3) !important; font-size:12px !important; }

/* ── progress ── */
.stProgress > div > div { background: var(--accent) !important; }
[data-testid="stProgressBar"] > div { background: var(--accent) !important; }

/* ── link button ── */
.stLinkButton a {
  border: 1px solid var(--border-2) !important;
  border-radius: var(--radius) !important;
  font-size:13px !important; text-decoration:none !important;
  transition: all 0.15s !important;
}
.stLinkButton a:hover { border-color: var(--accent) !important; color: var(--accent-dark) !important; }

/* ── 사이드바 항상 표시 ── */
[data-testid="stSidebarCollapseButton"] { display:none !important; }
[data-testid="collapsedControl"]        { display:none !important; }
section[data-testid="stSidebar"] {
  display:block !important; transform:translateX(0) !important;
  min-width:244px !important; visibility:visible !important;
}
section[data-testid="stSidebar"][aria-expanded="false"] {
  display:block !important; transform:translateX(0) !important; margin-left:0 !important;
}

/* ── 시스템 UI 숨김 ── */
.stTooltipIcon                    { display:none !important; }
div[data-testid="stStatusWidget"] { display:none !important; }
#MainMenu { visibility:hidden; }
footer    { visibility:hidden; }
header    { visibility:hidden; }

/* ── 슬라이더 진행바만 에메랄드 ── */
[data-testid="stSlider"] [role="progressbar"] { background: var(--accent) !important; }
[data-testid="stSlider"] [role="slider"] {
  background: var(--accent) !important; border-color: var(--accent) !important;
}
</style>
""", unsafe_allow_html=True)


# ── 사이드바 ──────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📢 원스톱 스케일업")
    st.caption("혁신제품지원센터")
    st.divider()

    # 메뉴 그룹핑
    st.caption("▸ 운영")
    page = st.radio("메뉴", [
        "대시보드",
        "기업 관리",
        "공고·매칭",
        "발송",
        "안내 메일",
        "발송 이력",
        "캘린더",
        "설정",
        "시스템 명세",
    ], label_visibility="collapsed")
    st.divider()
    test_mode = st.toggle("테스트 모드", value=True)
    if test_mode: st.warning("테스트 메일 발송")
    else:         st.success("실제 기업 발송")

# 구글 서비스 — 필요할 때 get_creds()로 직접 인증


# ══════════════════════════════════════════════════════
# 대시보드
# ══════════════════════════════════════════════════════
if page == "대시보드":
    drive = _get_drive()
    st.title("📊 운영 대시보드")

    with st.spinner("데이터 로딩 중..."):
        df_c  = load_excel(drive, SELECTED_FILE)
        df_n  = load_excel(drive, NOTICES_FILE)
        df_h  = load_excel(drive, HISTORY_FILE)
        df_det = load_excel(drive, DETAIL_FILE)

    # ── 상단 핵심 지표 4개 ─────────────────────────────
    m1, m2, m3, m4 = st.columns(4)

    # 선정 기업
    total_co = len(df_c)
    sel_co   = (df_c['선정구분']=='선정').sum() if '선정구분' in df_c.columns else total_co
    res_co   = (df_c['선정구분']=='예비').sum() if '선정구분' in df_c.columns else 0
    m1.metric("선정 기업", f"{sel_co}개사", f"예비 {res_co}개사")

    # 공고 DB
    active_n = 0
    if not df_n.empty and '마감일' in df_n.columns:
        today_str = datetime.today().strftime('%Y-%m-%d')
        active_n = (df_n['마감일'] >= today_str).sum()
    m2.metric("활성 공고", f"{active_n:,}건", f"전체 {len(df_n):,}건")

    # 전문 크롤링
    crawl_ok = 0
    crawl_fail = 0
    if not df_det.empty and '크롤링성공' in df_det.columns:
        crawl_ok   = (df_det['크롤링성공']=='Y').sum()
        crawl_fail = (df_det['크롤링성공']!='Y').sum()
    m3.metric("전문 수집", f"{crawl_ok}건", f"실패 {crawl_fail}건")

    # 발송 이력
    this_month_h = 0
    if not df_h.empty and '발송일' in df_h.columns:
        this_month = datetime.today().strftime('%Y-%m')
        this_month_h = df_h['발송일'].astype(str).str.startswith(this_month).sum()
    m4.metric("이번 달 발송", f"{this_month_h}건", f"누적 {len(df_h)}건")

    st.divider()

    # ── 이번 주 할 일 체크리스트 ───────────────────────
    st.subheader("🗂 이번 주 운영 체크리스트")

    # 각 단계 상태 자동 판단
    last_collect = "—"
    if not df_n.empty and '수정일' in df_n.columns:
        last_collect = df_n['수정일'].max()[:10] if df_n['수정일'].max() else "—"

    last_crawl = "—"
    need_crawl_cnt = 0
    if not df_det.empty and '크롤링일' in df_det.columns:
        last_crawl = df_det['크롤링일'].max()[:10] if df_det['크롤링일'].max() else "—"
    if not df_n.empty and not df_det.empty and 'pblancId' in df_n.columns and 'pblancId' in df_det.columns:
        crawled_pids = set(df_det[df_det.get('크롤링성공',pd.Series(''))=='Y']['pblancId']) if '크롤링성공' in df_det.columns else set()
        need_crawl_cnt = len(set(df_n['pblancId']) - crawled_pids)

    results      = st.session_state.get('match_results', [])
    review_state = st.session_state.get('review_state', {})
    approved     = sum(1 for v in review_state.values() if v=="○")
    pending_rev  = len(results) - sum(1 for v in review_state.values() if v in ["○","✕"])

    steps_data = [
        {
            "step": "① 공고 수집",
            "status": "완료" if not df_n.empty else "필요",
            "ok": not df_n.empty,
            "detail": f"마지막 수집: {last_collect} / 활성 공고 {active_n}건" if not df_n.empty else "공고 수집 탭에서 실행 필요",
            "action": "공고 수집"
        },
        {
            "step": "② 전문 크롤링",
            "status": "완료" if need_crawl_cnt == 0 else f"{need_crawl_cnt}건 대기",
            "ok": need_crawl_cnt == 0,
            "detail": f"수집 완료 {crawl_ok}건 / 미수집 {need_crawl_cnt}건" if not df_det.empty else "크롤링 미실행",
            "action": "공고 수집"
        },
        {
            "step": "③ 매칭 실행",
            "status": "완료" if results else "미실행",
            "ok": bool(results),
            "detail": f"매칭 결과 {len(results)}건 로드됨" if results else "매칭 결과 탭에서 실행 필요",
            "action": "매칭 결과"
        },
        {
            "step": "④ 검토 & 승인",
            "status": "완료" if (results and pending_rev==0) else (f"{pending_rev}건 미검토" if results else "매칭 후 진행"),
            "ok": results and pending_rev == 0,
            "detail": f"승인 {approved}건 / 미검토 {pending_rev}건" if results else "매칭 실행 후 가능",
            "action": "매칭 결과"
        },
        {
            "step": "⑤ 발송",
            "status": "발송 가능" if approved > 0 else "승인 후 진행",
            "ok": False,
            "detail": f"승인된 공고 {approved}건 발송 대기" if approved > 0 else "검토 승인 후 발송 가능",
            "action": "발송 관리"
        },
    ]

    for s in steps_data:
        icon = "✅" if s['ok'] else ("🟡" if "대기" in s['status'] or "가능" in s['status'] else "⬜")
        c1, c2, c3 = st.columns([2, 4, 2])
        with c1:
            st.write(f"{icon} **{s['step']}**")
        with c2:
            st.caption(s['detail'])
        with c3:
            badge_color = "#10B981" if s['ok'] else ("#F59E0B" if icon=="🟡" else "#E2E8F0")
            st.markdown(
                f"<span style='background:{badge_color};color:#000;padding:2px 10px;"
                f"border-radius:10px;font-size:11px;font-weight:700;'>{s['status']}</span>",
                unsafe_allow_html=True
            )

    st.divider()

    # ── 월간 공고 캘린더 ─────────────────────────────────
    st.subheader("📅 공고 마감 캘린더")

    import calendar as cal_mod

    # 월 선택
    cal_col1, cal_col2, _ = st.columns([1, 1, 4])
    with cal_col1:
        cal_year  = st.selectbox("연도", [2025, 2026, 2027],
                                  index=1, key="cal_year", label_visibility="collapsed")
    with cal_col2:
        cal_month = st.selectbox("월", list(range(1, 13)),
                                  index=datetime.today().month - 1,
                                  key="cal_month", label_visibility="collapsed",
                                  format_func=lambda x: f"{x}월")

    # 해당 월 공고 필터링
    if not df_n.empty and '마감일' in df_n.columns:
        month_str = f"{cal_year}-{cal_month:02d}"
        df_cal = df_n[df_n['마감일'].astype(str).str.startswith(month_str)].copy()

        # 날짜별 공고 그룹핑
        notices_by_day = {}
        for _, row in df_cal.iterrows():
            try:
                day = int(str(row['마감일'])[8:10])
                n_name = str(row.get('공고명', ''))[:18]
                if day not in notices_by_day:
                    notices_by_day[day] = []
                notices_by_day[day].append(n_name)
            except Exception:
                pass
    else:
        notices_by_day = {}
        df_cal = pd.DataFrame()

    # 달력 HTML 생성
    first_weekday, days_in_month = cal_mod.monthrange(cal_year, cal_month)
    # 0=월요일 기준으로 변환 (일요일=6 → 0으로)
    first_weekday = (first_weekday + 1) % 7  # 일요일 시작

    today = datetime.today()
    is_current_month = (today.year == cal_year and today.month == cal_month)

    # 요일 헤더
    day_headers = ["일", "월", "화", "수", "목", "금", "토"]
    header_html = "".join([
        f"<th style='padding:8px;text-align:center;font-size:12px;font-weight:600;"
        f"color:{'#EF4444' if i==0 else '#3B82F6' if i==6 else '#475569'};'>{d}</th>"
        for i, d in enumerate(day_headers)
    ])

    # 날짜 칸 생성
    cells = ["<td></td>"] * first_weekday
    for day in range(1, days_in_month + 1):
        ns = notices_by_day.get(day, [])
        is_today = is_current_month and day == today.day
        is_sun   = (first_weekday + day - 1) % 7 == 0
        is_sat   = (first_weekday + day - 1) % 7 == 6

        day_color = "#EF4444" if is_sun else "#3B82F6" if is_sat else "#0F172A"
        bg_color  = "#ECFDF5" if is_today else "#FFFFFF"
        border    = "2px solid #10B981" if is_today else "1px solid #E2E8F0"

        notices_html = ""
        for n in ns[:2]:  # 최대 2개 표시
            notices_html += (
                f"<div style='background:#EFF6FF;border-left:2px solid #3B82F6;"
                f"border-radius:3px;padding:1px 4px;margin-top:2px;"
                f"font-size:10px;color:#1E40AF;line-height:1.4;"
                f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>"
                f"📌 {n}</div>"
            )
        if len(ns) > 2:
            notices_html += (
                f"<div style='font-size:10px;color:#94A3B8;margin-top:2px;'>"
                f"+{len(ns)-2}건 더</div>"
            )

        cells.append(
            f"<td style='padding:6px;vertical-align:top;min-width:80px;min-height:80px;"
            f"background:{bg_color};border:{border};border-radius:4px;'>"
            f"<div style='font-size:13px;font-weight:700;color:{day_color};"
            f"margin-bottom:3px;'>{day}</div>"
            f"{notices_html}</td>"
        )

    # 7칸씩 행으로 나누기
    while len(cells) % 7 != 0:
        cells.append("<td style='background:#F8FAFC;border:1px solid #E2E8F0;'></td>")

    rows_html = ""
    for i in range(0, len(cells), 7):
        row_cells = "".join(cells[i:i+7])
        rows_html += f"<tr>{row_cells}</tr>"

    cal_html = f"""
    <div style="overflow-x:auto;">
    <table style="width:100%;border-collapse:separate;border-spacing:3px;
                  font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;">
      <thead>
        <tr style="background:#F8FAFC;">{header_html}</tr>
      </thead>
      <tbody>{rows_html}</tbody>
    </table>
    </div>
    """

    if notices_by_day:
        notice_cnt = len(df_cal)
        st.caption(f"{cal_year}년 {cal_month}월 마감 공고 총 **{notice_cnt}건** — 파란 테두리: 오늘")
    else:
        st.caption(f"{cal_year}년 {cal_month}월 마감 공고 없음 (공고 수집 후 표시)")

    st.markdown(cal_html, unsafe_allow_html=True)

    # 해당 월 공고 목록 (클릭 시 상세)
    if notices_by_day:
        with st.expander(f"📋 {cal_month}월 공고 전체 목록 ({len(df_cal)}건)"):
            for _, row in df_cal.sort_values('마감일').iterrows():
                c1, c2, c3 = st.columns([4, 2, 1])
                with c1:
                    st.write(f"**{row.get('공고명','')[:35]}**")
                with c2:
                    dl = str(row.get('마감일',''))
                    try:
                        days_left = (datetime.strptime(dl, '%Y-%m-%d') - datetime.today()).days
                        d_label = f"D-{days_left}" if days_left >= 0 else f"마감 {abs(days_left)}일 전"
                        color = "#EF4444" if days_left <= 3 else "#F59E0B" if days_left <= 7 else "#10B981"
                        st.markdown(f"<span style='color:{color};font-weight:600;font-size:12px;'>{d_label}</span>",
                                    unsafe_allow_html=True)
                    except Exception:
                        st.caption(dl)
                with c3:
                    if row.get('공고링크'):
                        st.markdown(f"[🔗]({row.get('공고링크','')})")

    st.divider()

    # ── 현황 분석 탭 ────────────────────────────────────
    st.subheader("📈 현황 분석")
    tab_c1, tab_c2, tab_c3, tab_c4 = st.tabs(["분야별 공고", "기업 관심분야", "매칭 현황", "발송 추이"])

    with tab_c1:
        if df_n.empty:
            st.info("공고 수집 후 확인 가능")
        elif '분야' in df_n.columns:
            import plotly.express as px
            realm_count = df_n['분야'].value_counts().reset_index()
            realm_count.columns = ['분야', '공고 수']
            realm_count = realm_count[realm_count['분야'] != ''].head(10)
            fig = px.bar(realm_count, x='분야', y='공고 수',
                         color_discrete_sequence=['#10B981'],
                         template='simple_white', height=300)
            fig.update_layout(margin=dict(l=0,r=0,t=10,b=0),
                              font_family="'Apple SD Gothic Neo','Malgun Gothic',sans-serif",
                              plot_bgcolor='white', paper_bgcolor='white')
            st.plotly_chart(fig, use_container_width=True)
            st.caption(f"총 {len(df_n):,}건 / {len(realm_count)}개 분야")
        else:
            st.info("분야 컬럼 없음")

    with tab_c2:
        if df_c.empty:
            st.info("선정기업 명단 업로드 후 확인 가능")
        else:
            import plotly.express as px
            col_a, col_b = st.columns(2)
            with col_a:
                if '관심사업분야' in df_c.columns:
                    all_interests = []
                    for val in df_c['관심사업분야']:
                        for item in str(val).split(','):
                            item = item.strip()
                            if item and item != 'nan':
                                all_interests.append(item)
                    interest_count = pd.Series(all_interests).value_counts().reset_index()
                    interest_count.columns = ['관심분야', '기업 수']
                    st.caption("관심사업분야 분포")
                    fig = px.bar(interest_count, x='관심분야', y='기업 수',
                                 color_discrete_sequence=['#10B981'],
                                 template='simple_white', height=250)
                    fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), plot_bgcolor='white', paper_bgcolor='white')
                    st.plotly_chart(fig, use_container_width=True)
            with col_b:
                if '소재지' in df_c.columns:
                    region_count = df_c['소재지'].value_counts().reset_index()
                    region_count.columns = ['소재지', '기업 수']
                    st.caption("소재지 분포")
                    fig = px.bar(region_count, x='소재지', y='기업 수',
                                 color_discrete_sequence=['#3B82F6'],
                                 template='simple_white', height=250)
                    fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), plot_bgcolor='white', paper_bgcolor='white')
                    st.plotly_chart(fig, use_container_width=True)

    with tab_c3:
        if not results:
            st.info("매칭 실행 후 확인 가능")
        else:
            import plotly.express as px
            df_r = pd.DataFrame(results)
            col_a, col_b = st.columns(2)
            with col_a:
                if '관련도' in df_r.columns:
                    star_count = df_r['관련도'].value_counts().reset_index()
                    star_count.columns = ['관련도', '건수']
                    st.caption("별점 분포")
                    fig = px.bar(star_count, x='관련도', y='건수',
                                 color_discrete_sequence=['#F59E0B'],
                                 template='simple_white', height=200)
                    fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), plot_bgcolor='white', paper_bgcolor='white')
                    st.plotly_chart(fig, use_container_width=True)
            with col_b:
                if '공고명' in df_r.columns:
                    top_notices = df_r.groupby('공고명').size().sort_values(ascending=False).head(10).reset_index()
                    top_notices.columns = ['공고명', '추천 기업 수']
                    top_notices['공고명'] = top_notices['공고명'].str[:18] + '...'
                    st.caption("공고별 추천 기업 수 (상위 10)")
                    fig = px.bar(top_notices, x='추천 기업 수', y='공고명',
                                 orientation='h',
                                 color_discrete_sequence=['#EF4444'],
                                 template='simple_white', height=260)
                    fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), yaxis={'categoryorder':'total ascending'},
                                      plot_bgcolor='white', paper_bgcolor='white')
                    st.plotly_chart(fig, use_container_width=True)
            if '공고유형' in df_r.columns:
                st.divider()
                type_count = df_r['공고유형'].value_counts()
                c1, c2, c3 = st.columns(3)
                c1.metric("맞춤 공고", f"{type_count.get('맞춤', 0)}건")
                c2.metric("공통 공고", f"{type_count.get('공통', 0)}건")
                c3.metric("기업당 평균", f"{len(df_r)/df_r['기업명'].nunique():.1f}건")

    with tab_c4:
        if df_h.empty:
            st.info("발송 이력이 쌓이면 확인 가능")
        else:
            if '발송일' in df_h.columns:
                import plotly.express as px
                df_h2 = df_h.copy()
                df_h2['발송일'] = pd.to_datetime(df_h2['발송일'], errors='coerce')
                df_h2['월'] = df_h2['발송일'].dt.to_period('M').astype(str)
                monthly = df_h2.groupby('월').size().reset_index(name='발송 건수')
                fig = px.line(monthly, x='월', y='발송 건수', markers=True,
                              color_discrete_sequence=['#10B981'],
                              template='simple_white', height=220)
                fig.update_layout(margin=dict(l=0,r=0,t=10,b=0), plot_bgcolor='white', paper_bgcolor='white')
                st.plotly_chart(fig, use_container_width=True)
            col_a, col_b, col_c = st.columns(3)
            col_a.metric("총 발송", f"{len(df_h)}건")
            if '신청여부' in df_h.columns:
                applied = (df_h['신청여부']=='Y').sum()
                col_b.metric("신청 건", f"{applied}건",
                             f"{applied/len(df_h)*100:.1f}%" if len(df_h) else "")
            if '선정결과' in df_h.columns:
                selected = (df_h['선정결과']=='선정').sum()
                col_c.metric("선정 건", f"{selected}건")


# ══════════════════════════════════════════════════════
# 기업 관리
# ══════════════════════════════════════════════════════
elif page == "기업 관리":
    drive = _get_drive()
    st.title("기업 관리")
    info_box("기업 관리",
        """
`선정기업_명단.xlsx` 기반 기업 정보 관리

**주요 기능**
- **파일 업로드** — 선정기업 명단 xlsx 업로드 → 드라이브 자동 저장
- **키워드 보완** — 기술키워드 부족 시 담당자 직접 추가 → 매칭 정확도 향상
- **수신거부** — 체크 시 이후 매칭·발송에서 자동 제외
- **구글계정** — 개별 캘린더 공유 시 활용

**매칭 반영 필드** → `기술키워드` + `제품분야` + `키워드보완` 합산
        """,
        "키워드 추가 → 기업 항목 열기 → '키워드 보완' 입력 → 저장")

    with st.spinner("드라이브에서 선정기업 명단 로딩 중..."):
        df_c = load_excel(drive, SELECTED_FILE)

    if df_c.empty:
        st.warning("드라이브에 선정기업 명단이 없음")
        st.info("선정기업_명단.xlsx 파일 업로드 → 드라이브 자동 저장")
        uploaded = st.file_uploader("선정기업_명단.xlsx 업로드", type=["xlsx"])
        if uploaded:
            df_new = pd.read_excel(uploaded, dtype=str).fillna("")
            for col in ['키워드보완','수신거부','메모','구글계정']:
                if col not in df_new.columns: df_new[col] = ''
            with st.spinner("드라이브 저장 중..."):
                if save_excel(drive, df_new, SELECTED_FILE, "선정기업명단", "1F4E79"):
                    st.success(f"{len(df_new)}개사 저장 완료!"); st.rerun()
    else:
        for col in ['키워드보완','수신거부','메모','구글계정']:
            if col not in df_c.columns: df_c[col] = ''

        # ── 선정/예비 구분 ─────────────────────────────────
        has_status = '선정구분' in df_c.columns
        df_선정 = df_c[df_c['선정구분']=='선정'] if has_status else df_c
        df_예비 = df_c[df_c['선정구분']=='예비'] if has_status else pd.DataFrame()

        # ── 상단 메트릭 ────────────────────────────────────
        c1,c2,c3,c4 = st.columns(4)
        c1.metric("전체", f"{len(df_c)}개사")
        c2.metric("선정", f"{len(df_선정)}개사")
        c3.metric("예비", f"{len(df_예비)}개사")
        c4.metric("수신거부", f"{(df_c['수신거부']=='Y').sum()}개사")
        st.divider()

        # ── 파일 교체 ──────────────────────────────────────
        with st.expander("📁 선정기업 명단 파일 교체"):
            uploaded = st.file_uploader("새 파일 업로드", type=["xlsx"], key="replace")
            if uploaded:
                df_new = pd.read_excel(uploaded, dtype=str).fillna("")
                for col in ['키워드보완','수신거부','메모','구글계정']:
                    if col not in df_new.columns: df_new[col] = ''
                if save_excel(drive, df_new, SELECTED_FILE, "선정기업명단", "1F4E79"):
                    st.success("교체 완료!"); st.rerun()

        if '기업명' not in df_c.columns:
            st.error(f"'기업명' 컬럼을 찾을 수 없음 — 현재 컬럼: {', '.join(df_c.columns.tolist())}")
            st.stop()

        # ── 탭 분리 ────────────────────────────────────────
        if has_status:
            tab_labels = [f"🟢 선정 {len(df_선정)}개사", f"🟡 예비 {len(df_예비)}개사", f"📋 전체 {len(df_c)}개사"]
        else:
            tab_labels = [f"📋 전체 {len(df_c)}개사"]
        tabs = st.tabs(tab_labels)

        def render_company_list(df_target, tab_key):
            search = st.text_input("🔍 기업명 검색", key=f"search_{tab_key}")
            df_show = df_target[df_target['기업명'].str.contains(search, na=False)] if search else df_target

            for idx, row in df_show.iterrows():
                unsub = str(row.get('수신거부','')) == 'Y'
                icon  = "🚫" if unsub else "🏢"
                status_badge = f"[{row.get('선정구분','')}]" if has_status and row.get('선정구분','') else ""
                score_badge  = f"총점 {row.get('평가_총점','')}점" if row.get('평가_총점','') else ""

                with st.expander(f"{icon} **{row.get('기업명','')}**  {status_badge}  |  {row.get('소재지','')}  |  {score_badge}"):
                    c1,c2 = st.columns(2)
                    with c1:
                        st.markdown(f"**소재지:** {row.get('소재지','')}")
                        st.markdown(f"**이메일:** {row.get('이메일','')}")
                        st.markdown(f"**관심분야:** {row.get('관심사업분야','')}")
                        st.markdown(f"**수출:** {row.get('수출실적','')} / {row.get('수출국가','')}")
                        st.markdown(f"**TRL단계:** {row.get('TRL단계','')}")
                        if row.get('평가_총점',''):
                            st.markdown(f"**평가점수:** 정량 {row.get('평가_정량점수','')}점 / 총점 {row.get('평가_총점','')}점")
                    with c2:
                        st.markdown(f"**제품분야:** {row.get('제품분야','')}")
                        st.markdown(f"**기술키워드:** {row.get('기술키워드','')}")
                        st.markdown(f"**핵심수요태그:** {row.get('핵심수요태그','')}")
                        st.markdown(f"**사업자번호:** {row.get('사업자등록번호','')}")
                        if row.get('평가_내부논의',''):
                            st.markdown(f"**평가의견:** {row.get('평가_내부논의','')}")

                    extra_kw   = st.text_input("키워드 보완", value=row.get('키워드보완',''),
                                    key=f"kw_{tab_key}_{idx}", placeholder="예: 스마트팜, IoT")
                    google_acc = st.text_input("구글계정", value=row.get('구글계정',''),
                                    key=f"ga_{tab_key}_{idx}", placeholder="example@gmail.com")
                    unsub_cb   = st.checkbox("수신거부", value=unsub, key=f"unsub_{tab_key}_{idx}")
                    memo       = st.text_input("메모", value=row.get('메모',''), key=f"memo_{tab_key}_{idx}")

                    if st.button("💾 저장", key=f"save_{tab_key}_{idx}"):
                        df_c.at[idx,'키워드보완'] = extra_kw
                        df_c.at[idx,'구글계정']   = google_acc
                        df_c.at[idx,'수신거부']   = 'Y' if unsub_cb else ''
                        df_c.at[idx,'메모']       = memo
                        with st.spinner("드라이브 저장 중..."):
                            if save_excel(drive, df_c, SELECTED_FILE, "선정기업명단", "1F4E79"):
                                st.success(f"{row['기업명']} 저장 완료!")

        if has_status:
            with tabs[0]: render_company_list(df_선정, "sel")
            with tabs[1]: render_company_list(df_예비, "res")
            with tabs[2]: render_company_list(df_c,   "all")
        else:
            with tabs[0]: render_company_list(df_c, "all")


# ══════════════════════════════════════════════════════
# 공고 수집
# ══════════════════════════════════════════════════════
elif page == "공고·매칭":
    drive = _get_drive()
    st.title("공고·매칭")

    tab_cm1, tab_cm2 = st.tabs(["📥 공고 수집", "🎯 매칭 실행·검토"])

    with tab_cm1:
        info_box("공고 수집",
            """
기업마당 API → 전체 공고 수집 → `notices_db.xlsx` 누적 저장

**수집 방식**
- 금융·기술개발·인력·수출·내수·창업·경영·기타 **8개 분야 전체** 수집
- `pblancId` 기준 **중복 저장 방지** (기존 공고 스킵)
- 내용 수정된 공고 **자동 업데이트**
- 마감일 자동 파싱 (비정형 마감일은 공란 처리)

**권장 수집 주기** — 주 1회 (매주 월요일)
            """,
            "수집 분야 변경 → 화면의 '수집 분야 선택' 옵션에서 체크박스로 선택")
    with st.spinner("드라이브에서 공고 DB 로딩 중..."):
        df_n = load_excel(drive, NOTICES_FILE)

    if not df_n.empty:
        c1,c2,c3 = st.columns(3)
        # 방금 수집한 경우 세션 값 우선 표시
        db_count = st.session_state.get('notices_count', len(df_n))
        c1.metric("현재 DB", f"{db_count:,}건")
        c2.metric("마지막 수집일", df_n['수집일'].max() if '수집일' in df_n.columns else "—")
        c3.metric("마감일 파싱 성공",
                  f"{(df_n['마감일']!='').sum()}건" if '마감일' in df_n.columns else "—")

    st.divider()

    # 분야 선택 옵션
    REALM_OPTIONS = {
        "금융":     "01",
        "기술개발": "02",
        "인력":     "03",
        "수출":     "04",
        "내수":     "05",
        "창업":     "06",
        "경영":     "07",
        "기타":     "09",
    }

    with st.expander("⚙️ 수집 분야 선택 (기본: 전체)", expanded=False):
        st.caption("원칙은 전체 수집 — 특정 분야만 빠르게 확인할 때 선택")
        col1, col2, col3, col4 = st.columns(4)
        selected_realms = {}
        for i, (name, code) in enumerate(REALM_OPTIONS.items()):
            col = [col1, col2, col3, col4][i % 4]
            selected_realms[code] = col.checkbox(name, value=True, key=f"realm_{code}")

        selected_codes = [code for code, checked in selected_realms.items() if checked]
        if not selected_codes:
            st.warning("최소 1개 이상 선택 필요")
            selected_codes = list(REALM_OPTIONS.values())

        st.caption(f"선택된 분야: **{len(selected_codes)}개** / {', '.join([k for k,v in REALM_OPTIONS.items() if v in selected_codes])}")

    if st.button("🔄 지금 수집 실행", type="primary"):
        REALM_CODES = selected_codes
        all_items, seen = [], set()
        prog = st.progress(0); log_area = st.empty(); logs = []
        realm_name_map = {v:k for k,v in REALM_OPTIONS.items()}

        import time as _time
        for idx, code in enumerate(REALM_CODES):
            params = {"crtfcKey":API_KEY,"dataType":"json","searchCnt":"0","searchLclasId":code}
            realm_name = realm_name_map.get(code, code)
            success = False
            for retry in range(3):  # 최대 3회 재시도
                try:
                    _time.sleep(0.8 if retry == 0 else 2)  # 첫 요청 0.8초, 재시도 2초 대기
                    resp  = requests.get(BASE_URL, params=params, timeout=40)
                    items = resp.json().get('jsonArray', [])
                    for item in items:
                        pid = item.get('pblancId','')
                        if pid and pid not in seen:
                            seen.add(pid); all_items.append(item)
                    logs.append(f"✅ {realm_name}: {len(items)}건")
                    success = True
                    break
                except Exception as e:
                    if retry < 2:
                        logs.append(f"⚠️ {realm_name}: 재시도 중... ({retry+1}/3)")
                    else:
                        logs.append(f"❌ {realm_name}: 수집 실패 ({e})")
                log_area.code("\n".join(logs))
            prog.progress((idx+1)/len(REALM_CODES)); log_area.code("\n".join(logs))

        def to_row(item):
            def pdl(s):
                try: return datetime.strptime(re.sub(r'\.', '-', s.split('~')[-1].strip()), "%Y-%m-%d").strftime("%Y-%m-%d")
                except: return ""
            return {"pblancId":item.get('pblancId',''),"공고명":item.get('pblancNm',''),
                    "주관기관":item.get('jrsdInsttNm',''),"분야":item.get('pldirSportRealmLclasCodeNm',''),
                    "세부분야":item.get('pldirSportRealmMlsfcCodeNm',''),
                    "접수기간":item.get('reqstBeginEndDe',''),"마감일":pdl(item.get('reqstBeginEndDe','')),
                    "지원대상":item.get('trgetNm',''),
                    "사업개요":strip_html(item.get('bsnsSumryCn','')),  # API 전체 반환값 저장
                    "해시태그":item.get('hashtags',''),"공고링크":item.get('pblancUrl',''),
                    "전문내용":"",  # 크롤링 후 채워짐
                    "수정일":item.get('updtPnttm',''),"수집일":datetime.today().strftime("%Y-%m-%d")}

        today = datetime.today().strftime("%Y-%m-%d")
        ex_map = {r['pblancId']:r.get('수정일','') for _,r in df_n.iterrows()} if not df_n.empty else {}
        new_rows, upd_rows = [], []
        dup_count = 0
        for item in all_items:
            pid = item.get('pblancId','')
            if not pid: continue
            row = to_row(item)
            if pid not in ex_map:
                new_rows.append(row)
            elif ex_map[pid] != item.get('updtPnttm',''):
                upd_rows.append(row)
            else:
                dup_count += 1

        if not df_n.empty:
            upd_ids  = {r['pblancId'] for r in upd_rows}
            df_base  = df_n[~df_n['pblancId'].isin(upd_ids)].copy()
            # 기존 공고도 수집일 오늘로 갱신 (마지막 수집일 반영)
            df_base['수집일'] = today
            df_final = pd.concat([df_base, pd.DataFrame(new_rows+upd_rows)], ignore_index=True)
        else:
            df_final = pd.DataFrame(new_rows)

        summary = "\n".join(logs) + f"\n\n📊 결과 요약\n  신규: {len(new_rows)}건 / 업데이트: {len(upd_rows)}건 / 중복유지: {dup_count}건 / 총 DB: {len(df_final):,}건"
        log_area.code(summary)

        with st.spinner("드라이브 저장 중..."):
            save_ok = save_excel(drive, df_final, NOTICES_FILE, "공고DB", "00897B")

        prog.progress(1.0)
        if save_ok:
            st.session_state['notices_count'] = len(df_final)
            st.session_state['notices_new']   = len(new_rows)
            st.session_state['notices_upd']   = len(upd_rows)
            st.session_state['pending_crawl'] = True  # 크롤링 필요 플래그
            st.success(
                f"✅ 수집 완료 — 총 {len(df_final):,}건 "
                f"(신규 {len(new_rows)} / 업데이트 {len(upd_rows)} / 중복유지 {dup_count})"
            )
            st.rerun()
        else:
            st.error("❌ 드라이브 저장 실패 — 설정 메뉴에서 드라이브 연동 확인 필요")
            st.info(f"수집은 완료: 총 {len(df_final):,}건 (신규 {len(new_rows)} / 업데이트 {len(upd_rows)} / 중복 {dup_count})")

    if not df_n.empty:
        st.divider()
        st.subheader("공고 DB 미리보기 (최근 20건)")
        cols = [c for c in ["공고명","주관기관","분야","접수기간","마감일"] if c in df_n.columns]
        st.dataframe(df_n[cols].head(20), use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("📄 공고 전문 크롤링")

    with st.spinner("전문 DB 현황 확인 중..."):
        df_detail = load_excel(drive, DETAIL_FILE)

    today_str = datetime.today().strftime('%Y-%m-%d')
    FRESHNESS_DAYS = 7  # 7일 이내 크롤링 = 최신

    if not df_detail.empty and '크롤링일' in df_detail.columns:
        df_detail['크롤링일'] = df_detail['크롤링일'].astype(str)
        fresh   = df_detail[df_detail['크롤링일'] >= (datetime.today() - timedelta(days=FRESHNESS_DAYS)).strftime('%Y-%m-%d')]
        stale   = df_detail[df_detail['크롤링일'] <  (datetime.today() - timedelta(days=FRESHNESS_DAYS)).strftime('%Y-%m-%d')]
        ok_pids = set(df_detail[df_detail['크롤링성공']=='Y']['pblancId'].tolist())
    else:
        fresh = stale = pd.DataFrame(); ok_pids = set()

    # 최신화 필요 공고 계산
    if not df_n.empty:
        active_pids = set(df_n[
            (df_n['마감일'] == '') | (df_n['마감일'] >= today_str)
        ]['pblancId'].tolist())
        need_crawl_pids = active_pids - ok_pids
    else:
        active_pids = need_crawl_pids = set()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("전문 수집", f"{len(ok_pids):,}건")
    c2.metric("7일 이내 최신", f"{len(fresh):,}건")
    c3.metric("갱신 필요", f"{len(stale):,}건")
    c4.metric("미수집 (활성공고)", f"{len(need_crawl_pids):,}건")

    if st.session_state.get('pending_crawl'):
        st.info(f"💡 공고 수집 완료 — 전문 크롤링을 실행하면 매칭 정확도가 높아져요.")

    col_a, col_b = st.columns([2,1])
    with col_a:
        st.caption(f"7일 이내 수집된 전문은 재수집 안 함. 미수집 {len(need_crawl_pids)}건 + 갱신필요 {len(stale)}건 대상.")
    with col_b:
        crawl_toggle = st.toggle("크롤링 실행", value=bool(st.session_state.get('pending_crawl')), key="crawl_toggle")

    if crawl_toggle:
        st.warning("⚠️ 수동 크롤링 — 공고 수에 따라 수십 분 소요될 수 있습니다. 설정한 건수마다 자동 중간 저장되어, 중간에 끊겨도 이미 처리한 건은 보존됩니다. 끊기면 그냥 다시 실행하면 미수집분부터 이어집니다.")

        col1, col2, col3 = st.columns(3)
        with col1:
            crawl_limit = st.number_input(
                "최대 크롤링 건수 (0 = 전체)",
                min_value=0, max_value=2000, value=50,
                help="테스트 시 50건 권장, 전체는 0 입력"
            )
        with col2:
            crawl_delay = st.selectbox(
                "요청 간격",
                ["빠름 (0.5초)", "보통 (1.2초)", "느림 (2초)"],
                index=1
            )
            delay_map = {"빠름 (0.5초)": 0.5, "보통 (1.2초)": 1.2, "느림 (2초)": 2.0}
            delay_sec = delay_map[crawl_delay]
        with col3:
            batch_size = st.number_input(
                "중간 저장 단위 (건)",
                min_value=10, max_value=200, value=30,
                help="이 건수마다 드라이브에 저장 → 중간에 끊겨도 이전 작업 보존"
            )

        if st.button("🕷️ 지금 크롤링 실행", type="primary", key="crawl_btn"):
            import time as _time
            from bs4 import BeautifulSoup

            today = datetime.today().strftime('%Y-%m-%d')

            def clean_notice_text(raw_text):
                """bizinfo 공고 본문에서 사이트 네비게이션/추천목록/평점위젯/첨부파일 영역을
                제거하고 실제 사업개요+신청자격+지원내용 영역만 남긴다.
                (실측 검증: 평균 1723자 -> 805자로 잡음 약 50% 제거, 핵심 정보는 보존)"""
                text = re.sub(r'본문\s*바로가기.*?화면크기', ' ', raw_text, flags=re.DOTALL)
                end_markers = [
                    r'이\s*공고를\s*열람한\s*사용자',  # 관련 공고 추천 영역 시작
                    r'NO\.\s*1\b',                      # 추천 목록 항목 시작
                    r'정보에\s*만족하셨나요',            # 평점 위젯
                    r'본문출력파일',                    # 첨부파일 목록 시작
                ]
                cut_idx = len(text)
                for pat in end_markers:
                    m = re.search(pat, text)
                    if m:
                        cut_idx = min(cut_idx, m.start())
                text = text[:cut_idx]
                text = re.sub(r'\s+', ' ', text).strip()
                return text

            import re

            # 크롤링 대상 필터
            # 날짜 기반 최신화: 미수집 + 7일 초과 갱신필요 + 마감 안 지난 것
            stale_pids = set(stale['pblancId'].tolist()) if not stale.empty else set()
            fresh_pids = set(fresh[fresh['크롤링성공']=='Y']['pblancId'].tolist()) if not fresh.empty else set()
            df_target = df_n[
                (~df_n['pblancId'].isin(fresh_pids)) &  # 최신 성공 건 제외
                ((df_n['마감일'] == '') | (df_n['마감일'] >= today))
            ].copy()

            if crawl_limit > 0:
                df_target = df_target.head(crawl_limit)

            st.info(f"크롤링 대상: {len(df_target)}건 (총 {-(-len(df_target)//batch_size)}개 배치, {batch_size}건씩 저장)")
            prog = st.progress(0); log_area = st.empty(); batch_status = st.empty()
            logs = []; new_records = []; success = fail = 0
            total_saved = 0

            HEADERS = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept-Language": "ko-KR,ko;q=0.9",
                "Referer": "https://www.bizinfo.go.kr/",
            }

            def save_batch(records_to_save):
                """new_records를 notices_detail.xlsx에 누적 저장. 매번 드라이브에서 최신본을 다시 읽어 합친다."""
                if not records_to_save:
                    return True
                with st.spinner(f"드라이브 저장 중... (누적 {total_saved + len(records_to_save)}건)"):
                    df_latest = load_excel(drive, DETAIL_FILE)  # 다른 배치/실행과 충돌 방지 위해 매번 최신본 로드
                    df_new_batch = pd.DataFrame(records_to_save)
                    df_merged = pd.concat([df_latest, df_new_batch], ignore_index=True) if not df_latest.empty else df_new_batch
                    # pblancId 중복 시 가장 마지막(최신) 레코드만 유지
                    if 'pblancId' in df_merged.columns:
                        df_merged = df_merged.drop_duplicates('pblancId', keep='last')
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine='openpyxl') as w:
                        df_merged.to_excel(w, index=False)
                    return drive_upload(drive, DETAIL_FILE, buf.getvalue(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            # bizinfo 상세페이지는 SPA(JS 렌더링) 구조라 requests로는 빈 body만 받아온다
            # (실측: html 8~9만자, body 자식태그 0개 — 진단 확정됨) → Playwright로 브라우저를 띄워 렌더링 후 텍스트 추출
            from playwright.sync_api import sync_playwright

            with sync_playwright() as pw:
                browser = pw.chromium.launch(headless=True)
                page = browser.new_page(user_agent=HEADERS["User-Agent"])

                for i, (_, row) in enumerate(df_target.iterrows()):
                    url = row.get('공고링크','')
                    pid = row.get('pblancId','')
                    name = row.get('공고명','')[:25]
                    if not url or not pid: continue

                    try:
                        _time.sleep(delay_sec)
                        page.goto(url, timeout=20000, wait_until="networkidle")
                        full_text = ""
                        for sel in ['.view-content','.detail-content','#bizSumryCn',
                                    '.bbs-view-content','#content','.board-view']:
                            try:
                                el = page.query_selector(sel)
                                if el:
                                    t = el.inner_text()
                                    if len(t) > 200:
                                        full_text = t
                                        break
                            except Exception:
                                continue
                        if len(full_text) < 100:
                            full_text = page.inner_text('body')

                        full_text = clean_notice_text(full_text)

                        amount = ""
                        for pat in [
                            r'(?:지원|보조|사업비?)\s{0,3}(?:금액|한도)[^0-9]{0,10}([0-9][0-9,]*\s{0,2}(?:백만원|천만원|만원|억원))',
                            r'(?:총|최대)\s{0,3}([0-9][0-9,]*\s{0,2}(?:백만원|천만원|만원|억원))\s{0,3}(?:이내|한도|지원)',
                            r'([0-9][0-9,]*\s{0,2}(?:백만원|천만원|만원|억원))\s{0,3}이내\s{0,3}지원',
                        ]:
                            m = re.search(pat, full_text)
                            if m: amount = m.group(0)[:30]; break
                        scale = ""
                        for pat in [
                            r'([0-9]+)\s{0,2}개\s{0,3}(?:사|업체|기업)\s{0,4}(?:내외|이내|선정|모집)',
                            r'선정\s{0,4}(?:규모|예정)[^0-9]{0,10}([0-9]+)\s{0,2}(?:개|개사|개업체)',
                            r'모집\s{0,4}(?:규모|인원)[^0-9]{0,10}([0-9]+)\s{0,2}(?:개|개사|개업체|명)',
                        ]:
                            m = re.search(pat, full_text)
                            if m: scale = m.group(0)[:30]; break

                        MIN_TEXT_LEN = 200
                        ok = len(full_text) >= MIN_TEXT_LEN
                        if ok:
                            # ── ③ 키워드 태그 자동 추출 ─────────────
                            # 전문에서 TARGET_KW/TYPE_KW 매칭 결과를 사전 계산
                            tag_target = [cat for cat, kws in TARGET_KW.items()
                                          if any(kw in full_text for kw in kws)]
                            tag_type   = [cat for cat, kws in TYPE_KW.items()
                                          if any(kw in full_text for kw in kws)]
                            # 지역 태그
                            bracket_m = re.search(r'\[([^\]]+)\]', name or '')
                            tag_region = bracket_m.group(1) if bracket_m else ''
                            # 업력 제한
                            age_m = re.findall(r'창업\s*(\d+)년\s*이내', full_text)
                            tag_age_limit = age_m[0] if age_m else ''
                            # 선착순 여부
                            tag_urgent = '선착순' if '선착순' in full_text else ''

                            new_records.append({
                                'pblancId':    pid,
                                '전문내용':    full_text[:4000],
                                '지원금액':    amount,
                                '선정규모':    scale,
                                '크롤링방법':  'Playwright',
                                '크롤링일':    today,
                                '크롤링성공':  'Y',
                                '태그_지원대상': ', '.join(tag_target),
                                '태그_사업성격': ', '.join(tag_type),
                                '태그_지역':    tag_region,
                                '태그_업력제한': tag_age_limit,
                                '태그_긴급':    tag_urgent,
                            })
                            success += 1
                            logs.append(f"✅ {name} ({len(full_text)}자)")
                        else:
                            new_records.append({'pblancId':pid,'전문내용':f"[진단] Playwright 렌더링 후에도 본문 {len(full_text)}자",
                                               '지원금액':'','선정규모':'','크롤링방법':'Playwright-diag',
                                               '크롤링일':today,'크롤링성공':'N'})
                            fail += 1
                            logs.append(f"❌ {name} (본문 {len(full_text)}자 — 렌더링 후에도 부족)")
                    except Exception as e:
                        new_records.append({'pblancId':pid,'전문내용':f"[진단] 예외: {str(e)[:100]}",'지원금액':'',
                                           '선정규모':'','크롤링방법':'Playwright','크롤링일':today,'크롤링성공':'N'})
                        fail += 1
                        logs.append(f"❌ {name} ({str(e)[:30]})")

                    prog.progress((i+1)/len(df_target))
                    log_area.code("\n".join(logs[-10:]))

                    # ── 배치 단위 중간 저장 ──────────────────
                    if len(new_records) >= batch_size:
                        if save_batch(new_records):
                            total_saved += len(new_records)
                            batch_status.success(f"💾 중간 저장 완료 — 누적 {total_saved}건 ({i+1}/{len(df_target)} 진행)")
                            new_records = []  # 저장된 건 비우고 다음 배치 시작
                        else:
                            batch_status.error("⚠️ 중간 저장 실패 — 다음 배치에서 재시도")

                browser.close()

            # 잔여분(배치 크기 미만으로 남은 마지막 묶음) 저장
            if new_records:
                if save_batch(new_records):
                    total_saved += len(new_records)
                    st.success(f"✅ 크롤링 완료 — 성공 {success}건 / 실패 {fail}건 / 총 저장 {total_saved}건 → notices_detail.xlsx")
                    st.rerun()
                else:
                    st.error(f"드라이브 저장 실패 — 단, 이전 배치 {total_saved}건은 이미 저장되어 있습니다.")
            elif total_saved > 0:
                st.success(f"✅ 크롤링 완료 — 성공 {success}건 / 실패 {fail}건 / 총 저장 {total_saved}건 → notices_detail.xlsx")
                st.rerun()
            else:
                st.info("새로 크롤링할 공고 없음")

    with tab_cm2:
        info_box("매칭 실행·검토",
            """
선정기업 × 공고 DB 교차 매칭 → 담당자 검토 → 발송 승인

**매칭 로직**
- 공고 전문(크롤링 완료 건) + API 사업개요를 통합 반영해 한 번에 매칭
- 키워드 7개 축 × 별점 판정 → 점수 높은 순 정렬 → 기업당 최대 N건 추출
- 이미 발송한 공고 / 마감 지난 공고 / 수신거부 기업 자동 제외

**검토 방법** — ○ 승인 / ✕ 제외 클릭 → 공고 원문 링크로 내용 확인 후 판단
            """,
            "★★★ 위주 먼저 검토 권장")

        tab1, tab2 = st.tabs(["▶ 매칭 실행", "🔍 검토 & 승인"])

        with tab1:
            col1, col2 = st.columns(2)
            with col1:
                max_per = st.slider("기업당 최대 추천 건수", 5, 20, 12)
            with col2:
                st.markdown("**발송 대상 그룹**")
                target_group = st.radio(
                    "발송 대상 그룹", ["선정 50개사", "예비 20개사", "전체 70개사"],
                    horizontal=True, label_visibility="collapsed", key="match_target_group"
                )

            # ── 스마트 매칭 권고 알림 ────────────────────────
            df_n_check = load_excel(drive, NOTICES_FILE)

            # 드라이브에서 마지막 매칭 정보 복원 (앱 재시작 후에도 유지)
            if 'last_match_info' not in st.session_state:
                kw_saved = load_json(drive, KEYWORDS_FILE) or {}
                if 'last_match' in kw_saved:
                    st.session_state['last_match_info'] = kw_saved['last_match']

            last_match_info = st.session_state.get('last_match_info', {})
            need_rematch    = False
            rematch_reason  = ""

            if not df_n_check.empty:
                # 공고 DB 최신 날짜
                latest_notice_date = ""
                if '수정일' in df_n_check.columns:
                    latest_notice_date = df_n_check['수정일'].max()[:10] if df_n_check['수정일'].max() else ""

                last_match_date    = last_match_info.get('date','')
                last_match_notices = last_match_info.get('notice_count', 0)
                cur_notice_count   = len(df_n_check)

                if not last_match_date:
                    need_rematch  = True
                    rematch_reason = "아직 매칭을 실행한 적이 없습니다."
                elif latest_notice_date and latest_notice_date > last_match_date:
                    new_cnt = cur_notice_count - last_match_notices
                    need_rematch  = True
                    rematch_reason = f"마지막 매칭({last_match_date}) 이후 공고 {new_cnt:+d}건 변동이 있습니다."
                else:
                    days_since = (datetime.today() - datetime.strptime(last_match_date, '%Y-%m-%d')).days
                    if days_since >= 3:
                        need_rematch  = True
                        rematch_reason = f"마지막 매칭({last_match_date})으로부터 {days_since}일 경과했습니다."

            if need_rematch:
                st.warning(f"🔄 **매칭 재실행 권장** — {rematch_reason}")
            elif last_match_info:
                st.success(f"✅ 매칭 최신 상태 — 마지막 실행: {last_match_info.get('date','')} / {last_match_info.get('result_count',0)}건")

            # 전문 수집 현황
            df_det_check = load_excel(drive, DETAIL_FILE)
            detail_ok_count = 0
            if not df_det_check.empty and '크롤링성공' in df_det_check.columns:
                detail_ok_count = (df_det_check['크롤링성공']=='Y').sum()
            if detail_ok_count > 0:
                st.success(f"📄 전문 수집 완료: {detail_ok_count}건 — 매칭에 자동 반영됩니다.")
            else:
                st.info("📄 전문 미수집 — API 사업개요만으로 매칭합니다.")

            if st.button("🔍 매칭 실행", type="primary"):
                with st.spinner("드라이브 데이터 로딩 중..."):
                    df_c    = load_excel(drive, SELECTED_FILE)
                    df_n    = load_excel(drive, NOTICES_FILE)
                    df_h    = load_excel(drive, HISTORY_FILE)
                    df_det  = load_excel(drive, DETAIL_FILE)
                    HIGH, MID = load_keywords(drive)

                if df_n.empty: st.error("notices_db 없음 → 공고 수집 먼저"); st.stop()
                if df_c.empty: st.error("선정기업 명단 없음 → 기업 관리에서 업로드"); st.stop()

                # 발송 대상 필터
                if '선정구분' in df_c.columns:
                    if target_group == "선정 50개사":
                        df_c = df_c[df_c['선정구분'] == '선정']
                    elif target_group == "예비 20개사":
                        df_c = df_c[df_c['선정구분'] == '예비']
                else:
                    st.warning("선정기업 명단에 '선정구분' 컬럼이 없어 전체 기업으로 매칭합니다.")
                if df_c.empty:
                    st.error(f"'{target_group}' 대상 기업이 없습니다."); st.stop()
                if '수신거부' in df_c.columns: df_c = df_c[df_c['수신거부']!='Y']

                # 전문내용 enrich 맵 구성
                detail_map = {}
                if not df_det.empty and 'pblancId' in df_det.columns:
                    # 버그 수정: df.get()은 컬럼 Series를 반환 → == 'Y' 비교 시 항상 빈 결과
                    # 올바른 방식: df['컬럼'] == 'Y'
                    if '크롤링성공' in df_det.columns:
                        df_det_ok = df_det[df_det['크롤링성공'] == 'Y']
                    else:
                        df_det_ok = df_det
                    df_det_ok = df_det_ok.drop_duplicates(subset='pblancId', keep='last')
                    detail_map = df_det_ok.set_index('pblancId').to_dict('index')

                # 진단: notices_db × notices_detail 교집합 확인
                notice_pids_set = set(df_n['pblancId'].str.strip()) if 'pblancId' in df_n.columns else set()
                overlap_count = len(notice_pids_set & set(detail_map.keys()))
                st.caption(f"📄 전문 DB {len(detail_map)}건 로드 | notices_db {len(df_n)}건 | 교집합(전문 반영 가능): {overlap_count}건")

                def enrich(n_dict):
                    """공고 dict에 전문내용을 병합해 매칭 정확도를 높인다.
                    score_notice()는 '사업개요'와 '전문내용' 두 필드를 모두 읽으므로 양쪽에 반영."""
                    pid = n_dict.get('pblancId','')
                    if pid in detail_map:
                        d = detail_map[pid]
                        full = str(d.get('전문내용',''))
                        if len(full) >= 200:
                            n_dict['전문내용'] = full   # score_notice 502줄에서 직접 읽힘
                            n_dict['사업개요'] = full   # 소재지 판정(578줄)에도 전문 내용 반영
                        n_dict['지원금액'] = d.get('지원금액','')
                        n_dict['선정규모'] = d.get('선정규모','')
                    return n_dict

                already_sent = set(zip(df_h['기업명'], df_h['pblancId'])) if not df_h.empty else set()
                all_results  = []
                prog         = st.progress(0, text="매칭 준비 중...")
                status_area  = st.empty()
                notice_recommend_count = {}
                match_errors = []

                # 피드백 + 가중치 + 제외키워드 통합 로드
                kw_config = load_kw_config(drive)
                feedback_map = kw_config.get('feedback', {})
                if feedback_map:
                    st.caption(f"🔄 피드백 반영 중 — {len(feedback_map)}개사 패턴 로드")
                if kw_config.get('exclude_keywords'):
                    st.caption(f"🚫 제외 키워드 {len(kw_config['exclude_keywords'])}개 적용")

                total_cos = len(df_c)
                for idx, (_, row) in enumerate(df_c.iterrows()):
                    co_name = row.get('기업명', f'기업{idx+1}')
                    prog.progress((idx+1)/total_cos,
                                  text=f"매칭 중... {idx+1}/{total_cos} ({co_name[:10]})")

                    try:
                        scored = []
                        for _, n in df_n.iterrows():
                            try:
                                r = score_notice(enrich(n.to_dict()), row, already_sent,
                                                 HIGH, MID, feedback_map, kw_config)
                                if r:
                                    scored.append(r)
                            except Exception as e_inner:
                                # 개별 공고 오류는 skip
                                pass

                        for r in scored:
                            pid = r.get('공고ID','')
                            cnt = notice_recommend_count.get(pid, 0)
                            if cnt >= 15:   r['점수'] -= 20
                            elif cnt >= 10: r['점수'] -= 15
                            elif cnt >= 7:  r['점수'] -= 10
                            elif cnt >= 4:  r['점수'] -= 6
                            elif cnt >= 2:  r['점수'] -= 3

                        scored.sort(key=lambda x: -x['점수'])
                        top = scored[:max_per]

                        for r in top:
                            pid = r.get('공고ID','')
                            notice_recommend_count[pid] = notice_recommend_count.get(pid, 0) + 1
                            r['_recommend_count'] = notice_recommend_count[pid]

                        all_results.extend(top)

                    except Exception as e_outer:
                        match_errors.append(f"{co_name}: {str(e_outer)[:60]}")
                        continue

                prog.progress(1.0, text="매칭 완료!")

                if match_errors:
                    with st.expander(f"⚠️ 처리 중 오류 {len(match_errors)}건"):
                        for err in match_errors:
                            st.caption(err)

                for r in all_results:
                    r['공고유형'] = '공통' if notice_recommend_count.get(r.get('공고ID',''), 1) >= 4 else '맞춤'

                enriched_count = len(detail_map)
                st.session_state['match_results'] = all_results
                st.session_state['df_companies_cache'] = df_c
                # 매칭 실행 정보 저장 — session_state + 드라이브 JSON (앱 재시작 후에도 유지)
                match_info = {
                    'date':         datetime.today().strftime('%Y-%m-%d'),
                    'time':         datetime.today().strftime('%H:%M'),
                    'notice_count': len(df_n),
                    'result_count': len(all_results),
                    'target_group': target_group,
                }
                st.session_state['last_match_info'] = match_info
                # 드라이브에도 저장
                kw_data_save = load_json(drive, KEYWORDS_FILE) or {}
                kw_data_save['last_match'] = match_info
                save_json(drive, kw_data_save, KEYWORDS_FILE)
                st.success(
                    f"✅ 매칭 완료 — 총 {len(all_results)}건 "
                    f"(전문 DB {enriched_count}건 반영 / 공고 총 {df_n.shape[0]}건 중 매칭) "
                    f"→ '검토 & 승인' 탭으로 이동"
                )

                # ── ① 버려진 공고 분석 ───────────────────────
                matched_pids = {r.get('공고ID','') for r in all_results}
                all_pids     = set(df_n['pblancId'].astype(str)) if 'pblancId' in df_n.columns else set()
                dropped_pids = all_pids - matched_pids
                dropped_df   = df_n[df_n['pblancId'].astype(str).isin(dropped_pids)]

                with st.expander(f"🗑️ 버려진 공고 분석 ({len(dropped_df)}건 / 전체 {len(df_n)}건)", expanded=False):
                    if not dropped_df.empty:
                        # 분야별 분포
                        c1, c2 = st.columns(2)
                        with c1:
                            st.caption("**버려진 공고 분야 분포**")
                            if '분야' in dropped_df.columns:
                                realm_drop = dropped_df['분야'].value_counts().head(8)
                                for realm, cnt in realm_drop.items():
                                    pct = cnt / len(dropped_df) * 100
                                    st.write(f"• {realm}: {cnt}건 ({pct:.0f}%)")
                        with c2:
                            st.caption("**버려진 공고 키워드 빈도**")
                            # 버려진 공고들에서 TARGET_KW/TYPE_KW 빈도 계산
                            kw_counter = {}
                            for _, row_d in dropped_df.iterrows():
                                text_d = str(row_d.get('사업개요','')) + str(row_d.get('공고명',''))
                                for cat, kws in {**TARGET_KW, **TYPE_KW}.items():
                                    if any(kw in text_d for kw in kws):
                                        kw_counter[cat] = kw_counter.get(cat, 0) + 1
                            for cat, cnt in sorted(kw_counter.items(), key=lambda x:-x[1])[:8]:
                                st.write(f"• {cat}: {cnt}건")

                        st.divider()
                        # 역제안: 버려진 공고 중 키워드 추가 시 매칭 가능한 것
                        st.caption("**💡 키워드 추가 시 매칭 가능할 수 있는 공고 (상위 10건)**")
                        st.caption("아래 공고들은 기업 키워드가 보완되면 매칭될 수 있습니다.")
                        dropped_sample = dropped_df[dropped_df['마감일'] >= datetime.today().strftime('%Y-%m-%d')].head(10)
                        for _, row_d in dropped_sample.iterrows():
                            name_d = str(row_d.get('공고명',''))[:40]
                            dl_d   = str(row_d.get('마감일',''))
                            realm_d = str(row_d.get('분야',''))
                            st.write(f"📌 **{name_d}** | {realm_d} | 마감 {dl_d}")

                if enriched_count == 0 and detail_ok_count > 0:
                    st.error("⚠️ 전문 DB 로드 실패 — '공고 수집' 탭에서 크롤링 상태를 확인하세요.")

        with tab2:
            results = st.session_state.get('match_results', [])
            if not results:
                st.info("매칭 실행 탭에서 먼저 실행 필요")
            else:
                df_show = pd.DataFrame(results)
                if 'review_state'  not in st.session_state: st.session_state['review_state']  = {}
                if 'ai_analysis'   not in st.session_state: st.session_state['ai_analysis']   = {}
                if 'custom_deadline' not in st.session_state: st.session_state['custom_deadline'] = {}

                # ── 검토 상태 저장/불러오기 ─────────────────────
                sv1, sv2, sv3 = st.columns([2, 2, 4])
                with sv1:
                    if st.button("💾 검토 상태 저장", help="드라이브에 현재 검토 상태를 저장합니다"):
                        import json
                        review_data = {
                            'review_state': st.session_state['review_state'],
                            'saved_at': datetime.today().strftime('%Y-%m-%d %H:%M')
                        }
                        drive_upload(drive, "review_state.json",
                                     json.dumps(review_data, ensure_ascii=False).encode('utf-8'),
                                     "application/json")
                        st.success("검토 상태 저장 완료")
                with sv2:
                    if st.button("📂 검토 상태 불러오기", help="드라이브에서 이전에 저장한 검토 상태를 불러옵니다"):
                        import json
                        saved = load_json(drive, "review_state.json")
                        if saved and 'review_state' in saved:
                            st.session_state['review_state'] = saved['review_state']
                            st.success(f"불러오기 완료 ({saved.get('saved_at','날짜 미상')} 저장본)")
                            st.rerun()
                        else:
                            st.warning("저장된 검토 상태가 없습니다")
                with sv3:
                    saved_json = load_json(drive, "review_state.json")
                    if saved_json:
                        st.caption(f"마지막 저장: {saved_json.get('saved_at','—')}")

                st.divider()

                # ── 필터 ────────────────────────────────────────
                c1, c2 = st.columns(2)
                with c1: filter_stars = st.multiselect("관련도", ["★★★","★★"], default=["★★★","★★"])
                with c2: filter_co    = st.selectbox("기업", ["전체"]+sorted(df_show['기업명'].unique().tolist()))
                filtered = df_show[df_show['관련도'].isin(filter_stars)]
                if filter_co != "전체": filtered = filtered[filtered['기업명']==filter_co]

                ap      = sum(1 for v in st.session_state['review_state'].values() if v=="○")
                rj      = sum(1 for v in st.session_state['review_state'].values() if v=="✕")
                pending = len(filtered) - ap - rj

                # ── 상단 진행 현황 ──────────────────────────────
                c1,c2,c3,c4 = st.columns(4)
                c1.metric("전체", f"{len(filtered)}건")
                c2.metric("✅ 승인", f"{ap}건")
                c3.metric("❌ 제외", f"{rj}건")
                c4.metric("⏳ 미검토", f"{pending}건")

                # ── 뷰 모드 선택 ───────────────────────────────
                view_mode = st.radio(
                    "보기 방식",
                    ["🏢 기업별 검토 (테이블)", "🤖 일괄 검토 (AI 활용)"],
                    horizontal=True, key="review_view_mode"
                )

                st.divider()

                # ══════════════════════════════════════════════
                # 기업별 검토 모드
                # ══════════════════════════════════════════════
                if view_mode == "🏢 기업별 검토 (테이블)":

                    companies_in_result = filtered['기업명'].unique().tolist()

                    def co_pending_count(co):
                        return sum(1 for _, r in filtered[filtered['기업명']==co].iterrows()
                                   if st.session_state['review_state'].get(
                                       f"{r['기업명']}_{r.get('공고ID','')}", "") == "")

                    co_labels = [
                        f"{'✅' if co_pending_count(co)==0 else '⏳'} {co}  ({co_pending_count(co)}건 미검토)"
                        for co in companies_in_result
                    ]
                    co_map = dict(zip(co_labels, companies_in_result))

                    # 기업명 자체를 session_state에 저장 (레이블 재생성 시 인덱스 틀어짐 방지)
                    if 'review_co_name' not in st.session_state or \
                       st.session_state['review_co_name'] not in companies_in_result:
                        st.session_state['review_co_name'] = companies_in_result[0]
                    cur_name    = st.session_state['review_co_name']
                    current_idx = companies_in_result.index(cur_name)

                    nav1, nav2, nav3, nav4 = st.columns([1, 5, 1, 2])
                    with nav1:
                        prev_clicked = st.button("◀ 이전", disabled=current_idx==0, key="btn_prev_co")
                    with nav2:
                        # 드롭다운 선택 (표시용) — index를 current_idx로 고정
                        # on_change 없이 매 rerun마다 index로 위치를 강제 지정
                        chosen = st.selectbox(
                            "기업 선택", co_labels,
                            index=current_idx,
                            label_visibility="collapsed",
                            key=f"co_select_{cur_name}"   # 기업명이 바뀌면 key도 바뀌어 위젯 초기화
                        )
                        # 드롭다운으로 직접 선택한 경우
                        chosen_name = co_map.get(chosen, cur_name)
                        if chosen_name != cur_name:
                            st.session_state['review_co_name'] = chosen_name
                            st.rerun()
                    with nav3:
                        next_clicked = st.button("다음 ▶", disabled=current_idx==len(co_labels)-1, key="btn_next_co")
                    with nav4:
                        st.caption(f"{current_idx+1} / {len(co_labels)}개사")

                    if prev_clicked:
                        st.session_state['review_co_name'] = companies_in_result[current_idx - 1]
                        st.rerun()
                    if next_clicked:
                        st.session_state['review_co_name'] = companies_in_result[current_idx + 1]
                        st.rerun()

                    selected_co = st.session_state['review_co_name']
                    co_rows = filtered[filtered['기업명'] == selected_co].copy()
                    co_rows['점수_num'] = pd.to_numeric(co_rows['점수'], errors='coerce').fillna(0)
                    co_rows = co_rows.sort_values('점수_num', ascending=False)

                    co_ap  = sum(1 for _, r in co_rows.iterrows() if st.session_state['review_state'].get(f"{r['기업명']}_{r.get('공고ID','')}", "")=="○")
                    co_rj  = sum(1 for _, r in co_rows.iterrows() if st.session_state['review_state'].get(f"{r['기업명']}_{r.get('공고ID','')}", "")=="✕")
                    co_pen = len(co_rows) - co_ap - co_rj

                    # ── 기업 정보 패널 ──────────────────────────
                    st.divider()
                    with st.expander(f"🏢 {selected_co} 기업 정보", expanded=True):
                        co_info_panel = {}
                        if 'df_companies_cache' in st.session_state:
                            df_co2 = st.session_state['df_companies_cache']
                            m = df_co2[df_co2['기업명']==selected_co]
                            if not m.empty: co_info_panel = m.iloc[0].to_dict()

                        # 상단 핵심 지표
                        pi1, pi2, pi3, pi4 = st.columns(4)
                        pi1.metric("공고 현황", f"전체 {len(co_rows)}건",
                                   f"✅{co_ap} / ❌{co_rj} / ⏳{co_pen}")
                        pi2.metric("소재지", co_info_panel.get('소재지','—'))
                        biz_type = co_info_panel.get('기업유형','—')
                        pi3.metric("기업유형", biz_type[:16] + ('...' if len(str(biz_type))>16 else '') if biz_type else '—')
                        pi4.metric("TRL / 매출", f"TRL {co_info_panel.get('TRL단계','—')} / {co_info_panel.get('매출규모','—')[:6]}" if co_info_panel.get('매출규모') else f"TRL {co_info_panel.get('TRL단계','—')}")

                        st.divider()

                        # 매칭 관련 정보
                        ic1, ic2 = st.columns(2)
                        with ic1:
                            st.markdown("**관심사업분야**")
                            st.write(co_info_panel.get('관심사업분야','—'))
                            st.markdown("**기술키워드**")
                            st.write(co_info_panel.get('기술키워드','—'))
                            st.markdown("**핵심수요태그**")
                            st.write(co_info_panel.get('핵심수요태그','—'))
                            if co_info_panel.get('키워드보완'):
                                st.markdown("**보완키워드** ✏️")
                                st.write(co_info_panel.get('키워드보완',''))
                        with ic2:
                            st.markdown("**제품분야**")
                            st.write(co_info_panel.get('제품분야','—'))
                            st.markdown("**수출실적 / 수출국가**")
                            st.write(f"{co_info_panel.get('수출실적','—')} / {co_info_panel.get('수출국가','—')}")
                            st.markdown("**이메일**")
                            st.write(co_info_panel.get('이메일','—'))

                        # 주관식 답변 (있는 경우만 표시)
                        has_subjective = any([
                            co_info_panel.get('희망서비스요약',''),
                            co_info_panel.get('평가_서비스요청',''),
                            co_info_panel.get('평가_내부논의',''),
                            co_info_panel.get('메모',''),
                        ])
                        if has_subjective:
                            st.divider()
                            if co_info_panel.get('희망서비스요약',''):
                                st.markdown("**📝 희망 서비스** (신청 시 작성)")
                                st.info(co_info_panel.get('희망서비스요약',''))
                            if co_info_panel.get('평가_서비스요청',''):
                                st.markdown("**📝 서비스 요청** (평가 시 작성)")
                                st.info(co_info_panel.get('평가_서비스요청',''))
                            if co_info_panel.get('평가_내부논의',''):
                                st.markdown("**📋 평가 내부 의견**")
                                st.warning(co_info_panel.get('평가_내부논의',''))
                            if co_info_panel.get('메모',''):
                                st.markdown("**🗒 운영 메모**")
                                st.write(co_info_panel.get('메모',''))

                    # ── 일괄 처리 버튼 ──────────────────────────
                    ba1, ba2, ba3, _ = st.columns([1, 1, 2, 2])
                    with ba1:
                        if st.button("✅ 전체 승인", key="bulk_approve"):
                            for _, r in co_rows.iterrows():
                                st.session_state['review_state'][f"{r['기업명']}_{r.get('공고ID','')}"] = "○"
                            st.rerun()
                    with ba2:
                        if st.button("❌ 전체 제외", key="bulk_reject"):
                            for _, r in co_rows.iterrows():
                                st.session_state['review_state'][f"{r['기업명']}_{r.get('공고ID','')}"] = "✕"
                            st.rerun()
                    with ba3:
                        co_ai_done  = sum(1 for _, r in co_rows.iterrows()
                                          if f"{r['기업명']}_{r.get('공고ID','')}" in st.session_state.get('ai_analysis', {}))
                        co_ai_total = len(co_rows)
                        if st.button(f"🤖 이 기업 AI 전체 분석 ({co_ai_done}/{co_ai_total})", key="co_bulk_ai"):
                            prog_co_ai = st.progress(0, text="AI 분석 중...")
                            for ai_i, (_, ai_row) in enumerate(co_rows.iterrows()):
                                ai_key = f"{ai_row['기업명']}_{ai_row.get('공고ID','')}"
                                if ai_key not in st.session_state.get('ai_analysis', {}):
                                    ci = {}
                                    if 'df_companies_cache' in st.session_state:
                                        df_co3 = st.session_state['df_companies_cache']
                                        mx = df_co3[df_co3['기업명']==ai_row['기업명']]
                                        if not mx.empty: ci = mx.iloc[0].to_dict()
                                    ci['기업명'] = ai_row['기업명']
                                    if 'ai_analysis' not in st.session_state: st.session_state['ai_analysis'] = {}
                                    st.session_state['ai_analysis'][ai_key] = claude_analyze(ci, ai_row.to_dict())
                                prog_co_ai.progress((ai_i+1)/co_ai_total, text=f"AI 분석 중... {ai_i+1}/{co_ai_total}")
                            save_ai_analysis(_get_drive())
                            st.success(f"✅ {selected_co} AI 분석 완료 — 드라이브 저장됨"); st.rerun()

                    st.divider()

                    # ── 공고 목록 (카드 형태) ───────────────────
                    for i, (idx, row) in enumerate(co_rows.iterrows()):
                        key     = f"{row['기업명']}_{row.get('공고ID','')}"
                        current = st.session_state['review_state'].get(key, "")
                        ai_res  = st.session_state.get('ai_analysis', {}).get(key)

                        try: loc_v = int(str(row.get('소재지점수','0')))
                        except: loc_v = 0
                        loc_txt  = "🟢 지역일치" if loc_v > 0 else ("🔴 타지역" if loc_v < 0 else "🔵 전국")
                        dl       = row.get('마감일','')
                        dl_txt   = "⚠️ 비정형" if not dl or not dl.strip() else dl
                        star     = row.get('관련도','')
                        nm       = row.get('공고명','')
                        org      = row.get('주관기관','')
                        reason   = row.get('매칭근거','')

                        # D-day 계산
                        d_day_txt = ""
                        if dl and dl.strip():
                            try:
                                days_left = (datetime.strptime(dl, '%Y-%m-%d') - datetime.today()).days
                                d_day_txt = f"D-{days_left}" if days_left >= 0 else f"마감"
                                d_color   = "#EF4444" if days_left <= 3 else "#F59E0B" if days_left <= 7 else "#10B981"
                            except: d_day_txt = dl; d_color = "#94A3B8"
                        else:
                            d_day_txt = "상시"; d_color = "#94A3B8"

                        # 상태 색상
                        status_color = {"○": "#ECFDF5", "✕": "#FEF2F2"}.get(current, "#F8FAFC")
                        status_border = {"○": "#10B981", "✕": "#EF4444"}.get(current, "#E2E8F0")
                        status_icon  = {"○": "✅ 승인됨", "✕": "❌ 제외됨"}.get(current, "⏳ 미검토")

                        # AI 배지
                        ai_badge = ""
                        if ai_res and not ai_res.get('error'):
                            rec = ai_res.get('추천여부','')
                            ai_badge = {"추천":"🟢 AI추천","검토":"🟡 AI검토","비추천":"🔴 AI비추천"}.get(rec,"")

                        # ── 카드 상단: 별점 + 공고명 + 상태 ──
                        top_c1, top_c2 = st.columns([8, 2])
                        with top_c1:
                            star_color = {"★★★":"#F59E0B","★★":"#10B981"}.get(star,"#94A3B8")
                            badges = f"<span style='font-size:13px;font-weight:700;color:{star_color};'>{star}</span>"
                            if ai_badge:
                                ai_col = {"🟢 AI추천":"#10B981","🟡 AI검토":"#F59E0B","🔴 AI비추천":"#EF4444"}.get(ai_badge,"#94A3B8")
                                badges += f" &nbsp;<span style='font-size:11px;background:{ai_col}22;color:{ai_col};padding:2px 7px;border-radius:10px;font-weight:600;'>{ai_badge}</span>"
                            st.markdown(
                                f"{badges}<br>"
                                f"<span style='font-size:15px;font-weight:700;color:#0F172A;'>{nm}</span><br>"
                                f"<span style='font-size:12px;color:#64748B;'>{org}</span>",
                                unsafe_allow_html=True
                            )
                        with top_c2:
                            st.markdown(
                                f"<div style='text-align:right;'>"
                                f"<span style='font-size:14px;font-weight:700;color:{d_color};'>{d_day_txt}</span><br>"
                                f"<span style='font-size:11px;color:#94A3B8;'>{dl_txt}</span><br>"
                                f"<span style='font-size:11px;color:#94A3B8;'>{loc_txt}</span>"
                                f"</div>",
                                unsafe_allow_html=True
                            )

                        # ── 매칭근거 + AI 요약 ──
                        if reason:
                            st.markdown(
                                f"<div style='background:#F0FDF4;border-left:3px solid #10B981;"
                                f"padding:6px 10px;border-radius:4px;margin:4px 0;font-size:12px;color:#065F46;'>"
                                f"↳ {reason[:80]}{'...' if len(reason)>80 else ''}</div>",
                                unsafe_allow_html=True
                            )
                        if ai_res and not ai_res.get('error'):
                            summary = ai_res.get('한줄요약','')
                            if summary:
                                st.markdown(
                                    f"<div style='font-size:12px;color:#374151;padding:3px 0;'>"
                                    f"🤖 {summary[:80]}</div>",
                                    unsafe_allow_html=True
                                )

                        # ── 액션 버튼 행 ──
                        b1, b2, b3, b4, _ = st.columns([1.5, 1.5, 1.5, 1.5, 3])
                        with b1:
                            lbl = "✅ 승인" if current != "○" else "↩ 취소"
                            btn_type = "primary" if current != "○" else "secondary"
                            if st.button(lbl, key=f"o_{key}_{i}", use_container_width=True, type=btn_type):
                                st.session_state['review_state'][key] = "" if current=="○" else "○"
                                st.rerun()
                        with b2:
                            lbl = "❌ 제외" if current != "✕" else "↩ 취소"
                            if st.button(lbl, key=f"x_{key}_{i}", use_container_width=True):
                                st.session_state['review_state'][key] = "" if current=="✕" else "✕"
                                st.rerun()
                        with b3:
                            if st.button("📋 상세", key=f"det_{key}_{i}", use_container_width=True):
                                toggle_key = f"show_detail_{key}"
                                st.session_state[toggle_key] = not st.session_state.get(toggle_key, False)
                                st.rerun()
                        with b4:
                            if not ai_res:
                                if st.button("🤖 AI분석", key=f"ai_{key}_{i}", use_container_width=True):
                                    with st.spinner("분석 중..."):
                                        ci = {}
                                        if 'df_companies_cache' in st.session_state:
                                            df_co3 = st.session_state['df_companies_cache']
                                            mx = df_co3[df_co3['기업명']==row['기업명']]
                                            if not mx.empty: ci = mx.iloc[0].to_dict()
                                        ci['기업명'] = row['기업명']
                                        if 'ai_analysis' not in st.session_state: st.session_state['ai_analysis'] = {}
                                        st.session_state['ai_analysis'][key] = claude_analyze(ci, row.to_dict())
                                    st.rerun()
                            else:
                                if st.button("🤖 AI상세", key=f"ai_{key}_{i}", use_container_width=True):
                                    toggle_ai = f"show_ai_{key}"
                                    st.session_state[toggle_ai] = not st.session_state.get(toggle_ai, False)
                                    st.rerun()

                        # ── 공고 상세 펼침 (토글) ──
                        if st.session_state.get(f"show_detail_{key}", False):
                            with st.container():
                                st.markdown("<div style='background:#F8FAFC;border:1px solid #E2E8F0;border-radius:8px;padding:14px;margin:6px 0;'>", unsafe_allow_html=True)
                                d1, d2 = st.columns(2)
                                with d1:
                                    st.caption("**지원대상**"); st.write(row.get('지원대상','—'))
                                    if row.get('지원금액'): st.caption("**지원금액**"); st.write(row.get('지원금액',''))
                                    if row.get('선정규모'): st.caption("**선정규모**"); st.write(row.get('선정규모',''))
                                with d2:
                                    st.caption("**접수기간**"); st.write(row.get('접수기간','—'))
                                    if row.get('공고링크'):
                                        st.link_button("🔗 공고 원문 보러가기", row.get('공고링크',''))
                                overview = row.get('사업개요','')
                                if 'HOME' in overview[:30]:
                                    cut = overview.find('소관부처')
                                    if cut > 0: overview = overview[cut:]
                                st.caption("**사업개요**"); st.write(overview[:500])
                                st.markdown("</div>", unsafe_allow_html=True)

                        # ── AI 분석 상세 펼침 (토글) ──
                        if ai_res and not ai_res.get('error') and st.session_state.get(f"show_ai_{key}", False):
                            rec       = ai_res.get('추천여부','')
                            fit       = ai_res.get('적합도','')
                            summary   = ai_res.get('한줄요약','')
                            reason_ai = ai_res.get('판단근거', ai_res.get('적합이유',''))
                            caution   = ai_res.get('주의사항','')
                            checks    = {"업종일치": ai_res.get('업종일치','—'), "자격충족": ai_res.get('자격충족','—'),
                                         "지역적합": ai_res.get('지역적합','—'), "수요일치": ai_res.get('수요일치','—')}
                            icon_map  = {"O":"✅","X":"❌","△":"⚠️"}
                            rec_icon  = {"추천":"🟢","검토":"🟡","비추천":"🔴"}.get(rec,"⚪")

                            st.markdown(
                                f"<div style='background:#EFF6FF;border:1px solid #BFDBFE;border-radius:8px;padding:14px;margin:6px 0;'>",
                                unsafe_allow_html=True
                            )
                            ai_top, ai_body = st.columns([1, 5])
                            with ai_top:
                                st.markdown(f"**{rec_icon} {rec}**")
                                st.caption(f"적합도 {fit}")
                            with ai_body:
                                st.markdown(f"**{summary}**")
                                st.write(reason_ai)
                                if caution and caution not in ['없음','','nan']:
                                    st.warning(f"⚠️ {caution}")
                                check_str = "  ".join([f"{icon_map.get(v,'—')} {k}" for k,v in checks.items()])
                                st.caption(check_str)
                            st.markdown("</div>", unsafe_allow_html=True)

                        st.divider()

                # ══════════════════════════════════════════════
                # 일괄 검토 (AI 활용) 모드
                # ══════════════════════════════════════════════
                else:
                    # 전체 AI 분석 — 상세 검토 탭에서만
                    with st.expander("🤖 전체 AI 분석 (상세 검토용)"):
                        usd, krw = estimate_cost(len(filtered))
                        st.caption(f"현재 필터 기준 {len(filtered)}건 분석 / 예상 비용 ${usd:.3f} (약 {krw:.0f}원)")
                        confirm = st.text_input("확인코드 입력 후 실행", key="bulk_ai_confirm", placeholder="분석실행")
                        if st.button("전체 분석 시작", key="bulk_ai_btn", type="primary"):
                            if confirm == "분석실행":
                                prog_ai = st.progress(0, text="AI 분석 중...")
                                for ai_i, (_, ai_row) in enumerate(filtered.iterrows()):
                                    ai_key = f"{ai_row['기업명']}_{ai_row.get('공고ID','')}"
                                    if ai_key not in st.session_state['ai_analysis']:
                                        co_info = {}
                                        if 'df_companies_cache' in st.session_state:
                                            df_co = st.session_state['df_companies_cache']
                                            co_rows2 = df_co[df_co['기업명']==ai_row['기업명']]
                                            if not co_rows2.empty:
                                                co_info = co_rows2.iloc[0].to_dict()
                                        co_info['기업명'] = ai_row['기업명']
                                        st.session_state['ai_analysis'][ai_key] = claude_analyze(co_info, ai_row.to_dict())
                                    prog_ai.progress((ai_i+1)/len(filtered), text=f"AI 분석 중... {ai_i+1}/{len(filtered)}")
                                save_ai_analysis(_get_drive())
                                st.success(f"완료 — {len(filtered)}건 (드라이브 저장됨)")
                                st.rerun()
                            else:
                                st.error("확인코드가 틀렸습니다 ('분석실행' 입력)")

                    st.divider()
                    st.caption("공고 상세 내용을 확인하며 검토합니다.")
                    for i,(idx,row) in enumerate(filtered.iterrows()):
                        key      = f"{row['기업명']}_{row.get('공고ID','')}"
                        current  = st.session_state['review_state'].get(key,"")
                        icon     = "🟡" if not current else ("✅" if current=="○" else "❌")
                        deadline = row.get('마감일','')
                        is_irregular = not deadline or deadline.strip() == ''
                        deadline_display = f"⚠️ 비정형" if is_irregular else deadline
                        reason = row.get('매칭근거','')
                        loc_icon = ""
                        if row.get('공고지역',''):
                            _ls = str(row.get('소재지점수','0'))
                            _lv = int(_ls) if _ls.lstrip('-').isdigit() else 0
                            loc_icon = " ✅" if _lv > 0 else (" ⚠️" if _lv < 0 else "")
                        with st.expander(
                            f"{icon} **{row['기업명']}**  |  {row.get('관련도','')}  |  "
                            f"{row.get('공고명','')[:28]}{loc_icon}  |  {reason[:22]}"
                        ):
                            left, right = st.columns(2)
                            with left:
                                st.markdown("**🏢 기업 정보**")
                                co_info = {}
                                if 'df_companies_cache' in st.session_state:
                                    df_co = st.session_state['df_companies_cache']
                                    co_rows = df_co[df_co['기업명']==row['기업명']]
                                    if not co_rows.empty:
                                        co_info = co_rows.iloc[0].to_dict()
                                co_loc      = co_info.get('소재지','—')
                                notice_loc  = row.get('공고지역','')
                                loc_score_v = str(row.get('소재지점수','0'))
                                loc_score   = int(loc_score_v) if loc_score_v.lstrip('-').isdigit() else 0
                                if notice_loc:
                                    loc_tag = "🟢 일치" if loc_score>0 else ("🔴 불일치" if loc_score<0 else "")
                                    st.markdown(f"- **소재지:** {co_loc} &nbsp; {loc_tag}")
                                else:
                                    st.markdown(f"- **소재지:** {co_loc}")
                                st.markdown(f"- **관심분야:** {co_info.get('관심사업분야','—')}")
                                st.markdown(f"- **기술키워드:** {co_info.get('기술키워드','—')}")
                                st.markdown(f"- **제품분야:** {co_info.get('제품분야','—')}")
                                st.markdown(f"- **수출실적:** {co_info.get('수출실적','—')} / {co_info.get('수출국가','—')}")
                                if co_info.get('TRL단계'): st.markdown(f"- **TRL:** {co_info.get('TRL단계')}")
                                if co_info.get('핵심수요태그'): st.markdown(f"- **핵심수요:** {co_info.get('핵심수요태그')}")
                                if co_info.get('키워드보완'): st.markdown(f"- **보완키워드:** {co_info.get('키워드보완')}")
                            with right:
                                st.markdown("**📋 공고 정보**")
                                if notice_loc:
                                    region_tag = f"🟢 `{notice_loc}` (귀사 소재지 포함)" if loc_score>0 else (f"🔴 `{notice_loc}` (미포함)" if loc_score<0 else f"`{notice_loc}`")
                                    st.markdown(f"- **지역제한:** {region_tag}")
                                else:
                                    st.markdown("- **지역제한:** 전국 공고")
                                st.markdown(f"- **주관기관:** {row.get('주관기관','—')}")
                                if row.get('지원금액',''): st.markdown(f"- **지원금액:** {row.get('지원금액','')}")
                                if row.get('선정규모',''): st.markdown(f"- **선정규모:** {row.get('선정규모','')}")
                                st.markdown(f"- **지원대상:** {row.get('지원대상','—')}")
                                st.markdown(f"- **접수기간:** {row.get('접수기간','—')}")
                                st.markdown(f"- **마감일:** {deadline_display}")
                                if row.get('공고링크',''): st.markdown(f"[🔗 공고 원문 보기]({row.get('공고링크','')})")
                            st.divider()
                            st.markdown("**🔍 매칭 근거**")
                            rc1,rc2,rc3,rc4 = st.columns(4)
                            with rc1:
                                st.caption("지원대상")
                                v = row.get('지원대상매칭','—')
                                st.markdown(f"`{v}`" if v and v!='—' else "—")
                            with rc2:
                                st.caption("사업성격")
                                v = row.get('사업성격매칭','—')
                                st.markdown(f"`{v}`" if v and v!='—' else "—")
                            with rc3:
                                st.caption("업종 역방향")
                                v = row.get('업종역방향매칭','—')
                                st.markdown(f"`{v}`" if v and v!='—' else "—")
                            with rc4:
                                st.caption("기업키워드·수요")
                                v = (row.get('핵심수요매칭','') or row.get('기업키워드매칭','')) or '—'
                                st.markdown(f"`{v}`" if v and v!='—' else "—")
                            st.divider()
                            st.caption("사업개요")
                            st.markdown(row.get('사업개요',''))
                            if is_irregular:
                                st.divider()
                                st.caption("📅 비정형 마감일 — 공고 원문 확인 후 직접 입력")
                                custom_dl = st.text_input("마감일 (YYYY-MM-DD)",
                                    value=st.session_state['custom_deadline'].get(key,''),
                                    key=f"dl_{key}_{i}", placeholder="예: 2026-06-30")
                                if custom_dl:
                                    st.session_state['custom_deadline'][key] = custom_dl
                                    for r in results:
                                        if f"{r['기업명']}_{r.get('공고ID','')}" == key:
                                            r['마감일'] = custom_dl; break
                            st.divider()
                            bc1,bc2,bc3,bc4 = st.columns([1,1,1.2,2])
                            with bc1:
                                if st.button("○ 승인", key=f"o_{key}_{i}", type="primary"):
                                    st.session_state['review_state'][key]="○"; st.rerun()
                            with bc2:
                                if st.button("✕ 제외", key=f"x_{key}_{i}"):
                                    st.session_state['review_state'][key]="✕"; st.rerun()
                            with bc3:
                                _usd,_krw = estimate_cost(1)
                                if st.button(f"🤖 AI 분석 (~{_krw:.0f}원)", key=f"ai_{key}_{i}"):
                                    if 'ai_analysis' not in st.session_state:
                                        st.session_state['ai_analysis'] = {}
                                    with st.spinner("Claude 분석 중..."):
                                        co_info = {}
                                        if 'df_companies_cache' in st.session_state:
                                            df_co = st.session_state['df_companies_cache']
                                            co_rows = df_co[df_co['기업명']==row['기업명']]
                                            if not co_rows.empty: co_info = co_rows.iloc[0].to_dict()
                                        co_info['기업명'] = row['기업명']
                                        result = claude_analyze(co_info, row.to_dict())
                                        st.session_state['ai_analysis'][key] = result
                                    st.rerun()
                            ai_result = st.session_state.get('ai_analysis', {}).get(key)
                            if ai_result:
                                if 'error' in ai_result:
                                    st.error(f"분석 오류: {ai_result['error']}")
                                else:
                                    rec       = ai_result.get('추천여부','')
                                    fit       = ai_result.get('적합도','')
                                    rec_color = {"추천":"🟢","검토":"🟡","비추천":"🔴"}.get(rec,"⚪")
                                    fit_color = {"높음":"#10B981","보통":"#F59E0B","낮음":"#EF4444"}.get(fit,"#0F172A")
                                    icon_map  = {"O":"✅","X":"❌","△":"⚠️"}
                                    checks = {
                                        "업종일치": ai_result.get('업종일치','—'),
                                        "자격충족": ai_result.get('자격충족','—'),
                                        "지역적합": ai_result.get('지역적합','—'),
                                        "수요일치": ai_result.get('수요일치','—'),
                                    }
                                    check_html = "".join([
                                        f"<span style='margin-right:12px;font-size:12px;'>"
                                        f"{icon_map.get(v,'—')} {k}</span>"
                                        for k,v in checks.items()
                                    ])
                                    caution_html = ""
                                    if ai_result.get('주의사항','') not in ['없음','','nan']:
                                        caution_txt = ai_result.get('주의사항','')
                                        caution_html = f"<p style='margin:8px 0 0;font-size:11px;color:#F59E0B;'>⚠️ {caution_txt}</p>"
                                    judgment = ai_result.get('판단근거', ai_result.get('적합이유',''))
                                    summary_txt = ai_result.get('한줄요약','')
                                    st.markdown(f"""
    <div style="background:rgba(74,158,255,0.08);border:1px solid rgba(74,158,255,0.2);
                border-radius:8px;padding:14px 16px;margin-top:8px;">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
        <span style="font-size:11px;font-weight:700;color:#10B981;letter-spacing:1px;">🤖 CLAUDE 분석</span>
        <span style="font-size:13px;font-weight:700;color:{fit_color};">{rec_color} {rec}</span>
        <span style="font-size:12px;color:rgba(255,255,255,0.5);">적합도: {fit}</span>
      </div>
      <p style="margin:0 0 8px;font-size:13px;font-weight:600;color:#0F172A;">{summary_txt}</p>
      <div style="margin-bottom:10px;">{check_html}</div>
      <p style="margin:0 0 6px;font-size:12px;color:rgba(255,255,255,0.65);line-height:1.7;">{judgment}</p>
      {caution_html}
    </div>
                                    """, unsafe_allow_html=True)
                st.divider()
                c1, c2, c3 = st.columns(3)
                with c1:
                    if st.button("✅ 검토 완료 저장", type="primary"):
                        for r in results:
                            r['담당자검토'] = st.session_state['review_state'].get(
                                f"{r['기업명']}_{r.get('공고ID','')}", "")
                        st.session_state['match_results'] = results
                        st.success(f"저장 완료 — 승인 {ap}건 → '발송 관리' 메뉴로 이동")
                with c2:
                    if st.button("📥 매칭결과 엑셀 저장"):
                        fname = f"매칭결과_{datetime.today().strftime('%Y%m%d')}.xlsx"
                        ai_data = st.session_state.get('ai_analysis', {})
                        export_rows = []
                        for r in results:
                            r2  = dict(r)
                            key = f"{r2.get('기업명','')}_{r2.get('공고ID','')}"
                            ai  = ai_data.get(key, {})
                            if ai and not ai.get('error'):
                                r2['AI_추천여부'] = ai.get('추천여부','')
                                r2['AI_적합도']   = ai.get('적합도','')
                                r2['AI_한줄요약'] = ai.get('한줄요약','')
                                r2['AI_판단근거'] = ai.get('판단근거', ai.get('적합이유',''))
                                r2['AI_주의사항'] = ai.get('주의사항','')
                                r2['AI_업종일치'] = ai.get('업종일치','')
                                r2['AI_자격충족'] = ai.get('자격충족','')
                                r2['AI_지역적합'] = ai.get('지역적합','')
                                r2['AI_수요일치'] = ai.get('수요일치','')
                            export_rows.append(r2)
                        with st.spinner("드라이브 저장 중..."):
                            save_excel(drive, pd.DataFrame(export_rows), fname,
                                       "매칭결과", "C55A11", star_col="관련도")
                        ai_cnt = sum(1 for r in export_rows if r.get('AI_추천여부',''))
                        st.success(f"✅ {fname} 저장 완료 — AI 분석 {ai_cnt}건 포함")
                with c3:
                    with st.expander("🔄 피드백 반영 (매칭 개선)"):
                        st.caption("""제외한 공고 패턴을 분석해 다음 매칭에 자동 반영합니다.
    기업별로 자주 제외된 사업성격을 낮은 우선순위로 설정합니다.""")
                        if st.button("피드백 적용", key="apply_feedback"):
                            # 제외된 공고의 사업성격 패턴 집계
                            rejected = [r for r in results
                                        if st.session_state['review_state'].get(
                                            f"{r['기업명']}_{r.get('공고ID','')}", "") == "✕"]

                            # 기업별 제외 패턴
                            co_reject_types = {}
                            for r in rejected:
                                co  = r.get('기업명','')
                                typ = r.get('사업성격매칭','')
                                if co and typ:
                                    if co not in co_reject_types:
                                        co_reject_types[co] = []
                                    for t in typ.split('/'):
                                        t = t.strip().split('(')[0].strip()
                                        if t: co_reject_types[co].append(t)

                            # keywords.json에 제외 패턴 저장
                            kw_data = load_json(drive, KEYWORDS_FILE) or {}
                            feedback = kw_data.get('feedback', {})
                            for co, types in co_reject_types.items():
                                if co not in feedback:
                                    feedback[co] = {}
                                for t in types:
                                    feedback[co][t] = feedback[co].get(t, 0) + 1
                            kw_data['feedback'] = feedback

                            if save_json(drive, kw_data, KEYWORDS_FILE):
                                summary_lines = [f"{co}: {dict(v)}" for co, v in list(co_reject_types.items())[:5]]
                                st.success(f"피드백 반영 완료 — {len(rejected)}건 제외 패턴 저장")
                                for line in summary_lines:
                                    st.caption(line)
                            else:
                                st.error("저장 실패")


# ══════════════════════════════════════════════════════
# 발송 관리
# ══════════════════════════════════════════════════════
elif page == "발송":
    drive = _get_drive()
    st.title("발송 관리")
    info_box("발송 관리",
        """
담당자 승인(○) 건 → 기업별 HTML 메일 발송 + 캘린더 D-day 등록

**발송 방식**
- HTML 메일 — 공고명 클릭 가능한 링크 포함, 네이버·회사메일 호환
- 캘린더 등록 — 마감 당일·D-7·D-3 이벤트 자동 등록
  - 공통 캘린더: 전체 선정기업 공유
  - 개별 캘린더: 해당 기업 전용 (생성된 경우)
- 발송 이력 — send_history.xlsx 자동 기록 → 다음 매칭 시 중복 제외

**테스트 모드** — 사이드바 토글 ON 시 본인 메일로만 발송
실제 발송 전 테스트 모드로 먼저 확인 권장
        """,
        "발신자 이름·서명 변경 → `app.py` HTML 템플릿 수정")

    results  = st.session_state.get('match_results', [])
    approved = [r for r in results if r.get('담당자검토')=='○']
    matched_group = st.session_state.get('match_target_group', '미확인')

    if not approved:
        st.warning("승인된 공고 없음 — '매칭 결과'에서 검토 완료 후 진행")
    else:
        st.info(f"📌 현재 매칭 대상 그룹: **{matched_group}** (다른 그룹 발송 시 '매칭 결과'에서 그룹 변경 후 재매칭 필요)")
        companies = list(set(r['기업명'] for r in approved))
        c1,c2,c3  = st.columns(3)
        c1.metric("승인 건수", f"{len(approved)}건")
        c2.metric("대상 기업", f"{len(companies)}개사")
        c3.metric("발송 모드", "테스트" if test_mode else "실제")
        if test_mode: st.warning("⚠️ 테스트 모드 — 본인 메일로만 발송")
        else:         st.success("✅ 실제 모드 — 기업 담당자 이메일로 발송")

        st.divider()
        st.subheader("발송 미리보기")
        preview_co      = st.selectbox("기업 선택", companies)
        preview_notices = [r for r in approved if r['기업명']==preview_co]
        with st.expander(f"📧 {preview_co} 메일 미리보기"):
            custom_n = [n for n in preview_notices if n.get('공고유형','맞춤')=='맞춤']
            common_n = [n for n in preview_notices if n.get('공고유형','맞춤')=='공통']
            if custom_n:
                st.caption("**🎯 맞춤 공고**")
                for i,n in enumerate(custom_n,1):
                    st.markdown(f"**{i}. {n.get('관련도','')} [{n.get('공고명','')}]({n.get('공고링크','#')})**")
                    st.caption(f"주관: {n.get('주관기관','')}  |  기간: {n.get('접수기간','')}")
            if common_n:
                st.divider()
                st.caption("**📌 공통 공고 (이런 공고도 있어요)**")
                for i,n in enumerate(common_n,1):
                    st.markdown(f"{i}. [{n.get('공고명','')}]({n.get('공고링크','#')})")
                    st.caption(f"주관: {n.get('주관기관','')}  |  기간: {n.get('접수기간','')}")

        st.divider()
        if st.button("📤 발송 실행", type="primary"):
            from email.mime.multipart import MIMEMultipart
            from email.mime.text import MIMEText
            import base64

            cal_id        = load_text(drive, CALID_FILE)
            ind_cals      = load_json(drive, INDCAL_FILE)
            CALENDAR_LINK = f"https://calendar.google.com/calendar?cid={cal_id}" if cal_id else ""
            df_c_cur      = load_excel(drive, SELECTED_FILE)

            history_records = []; prog = st.progress(0); log = st.empty(); logs = []
            grouped = {}
            for r in approved: grouped.setdefault(r['기업명'],[]).append(r)

            for idx,(company,notices) in enumerate(grouped.items()):
                # 공고 분류: 맞춤(기업 전용) / 공통(여러 기업 공통)
                notices_custom = [n for n in notices if n.get('공고유형','맞춤') == '맞춤']
                notices_common = [n for n in notices if n.get('공고유형','맞춤') == '공통']
                # 맞춤 내에서 별점 구분
                notices_sss = [n for n in notices_custom if n.get('관련도','')=='★★★']
                notices_ss  = [n for n in notices_custom if n.get('관련도','')=='★★']

                def notice_card_simple(n, idx):
                    """공통 공고용 심플 카드 (작고 간결하게)"""
                    dl_raw = n.get('마감일','')
                    if not dl_raw and '~' in n.get('접수기간',''):
                        dl_raw = n.get('접수기간','').split('~')[-1].strip()
                    return f"""
                    <table width="100%" cellpadding="0" cellspacing="0"
                           style="margin-bottom:6px;">
                      <tr>
                        <td style="padding:10px 14px;
                                   background:rgba(255,255,255,0.02);
                                   border:1px solid rgba(255,255,255,0.05);
                                   border-radius:6px;">
                          <a href="{n.get('공고링크','#')}"
                             style="font-size:13px;font-weight:500;color:rgba(255,255,255,0.6);
                                    text-decoration:none;display:block;">
                            {n.get('공고명','')}
                          </a>
                          <p style="margin:3px 0 0;font-size:11px;color:rgba(255,255,255,0.25);">
                            {n.get('주관기관','')} &nbsp;·&nbsp; 마감 {dl_raw}
                          </p>
                        </td>
                      </tr>
                    </table>"""

                def notice_card(n, idx):
                    star = n.get('관련도','')
                    dl_raw = n.get('마감일','')
                    if not dl_raw and '~' in n.get('접수기간',''):
                        dl_raw = n.get('접수기간','').split('~')[-1].strip()
                    # 매칭근거 → 자연어 한 줄 변환
                    reason_sentence = reason_to_sentence(n.get('매칭근거',''))
                    reason_html = f"""
                          <p style="margin:6px 0 0;font-size:11px;color:#10B981;
                                     font-style:normal;line-height:1.5;">
                            ↳ {reason_sentence}
                          </p>""" if reason_sentence else ""
                    return f"""
                    <table width="100%" cellpadding="0" cellspacing="0"
                           style="margin-bottom:10px;background:#FFFFFF;
                                  border:1px solid #E2E8F0;
                                  border-radius:10px;overflow:hidden;
                                  box-shadow:0 1px 3px rgba(0,0,0,0.04);">
                      <tr>
                        <td style="padding:14px 16px;">
                          <a href="{n.get('공고링크','#')}"
                             style="font-size:14px;font-weight:600;color:#0F172A;
                                    text-decoration:none;line-height:1.5;display:block;">
                            {n.get('공고명','')}
                          </a>
                          <p style="margin:4px 0 0;font-size:12px;color:#94A3B8;">
                            {n.get('주관기관','')}
                            &nbsp;·&nbsp;
                            마감 {dl_raw if dl_raw else '상시'}
                          </p>
                          {reason_html}
                        </td>
                        <td width="60" align="center" valign="middle"
                            style="padding:14px 12px;
                                   border-left:1px solid #F1F5F9;">
                          <a href="{n.get('공고링크','#')}"
                             style="display:inline-block;font-size:12px;font-weight:600;
                                    color:#10B981;text-decoration:none;white-space:nowrap;">
                            보기 →
                          </a>
                        </td>
                      </tr>
                    </table>"""

                rows_html = ""

                # ── 맞춤 공고 섹션 ──────────────────────
                if notices_sss:
                    rows_html += """
                    <p style="margin:0 0 10px;font-size:10px;font-weight:700;
                               color:#F59E0B;letter-spacing:2px;text-transform:uppercase;">
                      ★★★ &nbsp;직접 연계 추천
                    </p>"""
                    for i,n in enumerate(notices_sss): rows_html += notice_card(n, i)
                    rows_html += """<div style="height:20px;"></div>"""

                if notices_ss:
                    rows_html += """
                    <p style="margin:0 0 10px;font-size:10px;font-weight:700;
                               color:#10B981;letter-spacing:2px;text-transform:uppercase;">
                      ★★ &nbsp;참고 추천
                    </p>"""
                    for i,n in enumerate(notices_ss): rows_html += notice_card(n, i)

                # ── 공통 공고 섹션 (이런 공고도 있어요) ──
                if notices_common:
                    rows_html += """<div style="height:24px;"></div>"""
                    rows_html += """
                    <div style="border-top:1px solid rgba(255,255,255,0.08);
                                padding-top:16px;margin-top:4px;">
                      <p style="margin:0 0 10px;font-size:10px;font-weight:700;
                                 color:rgba(255,255,255,0.35);letter-spacing:2px;
                                 text-transform:uppercase;">
                        📌 &nbsp;이런 공고도 있어요
                      </p>"""
                    for i,n in enumerate(notices_common):
                        rows_html += notice_card_simple(n, i)
                    rows_html += "</div>"

                ind_link=""
                if company in ind_cals and ind_cals[company].get('calendar_id'):
                    ind_link=f"""<div style="margin-top:8px">
                      <a href="https://calendar.google.com/calendar?cid={ind_cals[company]['calendar_id']}"
                         style="color:#2E75B6;font-size:13px">📅 {company} 전용 캘린더 구독</a></div>"""

                cal_sec=f"""
                <div style="background:#F0FDF4;border-radius:10px;
                            padding:16px 18px;border:1px solid #A7F3D0;margin:16px 0;">
                  <p style="margin:0 0 4px;color:#059669;font-weight:700;font-size:11px;
                             letter-spacing:1.5px;text-transform:uppercase;">
                    📅 공고 마감일 캘린더 알림
                  </p>
                  <p style="margin:0 0 12px;font-size:12px;color:#374151;">
                    D-7 · D-3 자동 알림을 받아보세요. 구글 계정만 있으면 1회 클릭으로 설정됩니다.
                  </p>
                  {"<a href='"+CALENDAR_LINK+"' style='display:inline-block;background:#10B981;color:#fff;padding:8px 18px;border-radius:6px;text-decoration:none;font-size:12px;font-weight:700;'>📅 공통 캘린더 구독하기 →</a>" if CALENDAR_LINK else "<p style='margin:0;font-size:12px;color:#6B7280;'>※ 캘린더 링크는 별도 안내 예정입니다.</p>"}
                  {ind_link}
                  <p style="margin:10px 0 0;font-size:11px;color:#9CA3AF;">
                    구글 계정이 없으신 경우 담당자에게 문의 주시면 안내해드립니다.
                  </p>
                </div>"""

                today_str = datetime.today().strftime('%Y.%m.%d')
                html=f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width,initial-scale=1.0">
</head>
<body style="margin:0;padding:0;background:#F2F4F7;
             font-family:'Apple SD Gothic Neo','Malgun Gothic',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0"
       style="background:#F2F4F7;padding:36px 0 52px;">
<tr><td align="center">
<table width="560" cellpadding="0" cellspacing="0">

  <!-- ── 로고 헤더 (흰 배경) ── -->
  <tr>
    <td style="background:#ffffff;border-radius:14px 14px 0 0;
               padding:20px 28px;border-bottom:1px solid #E8ECF0;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td valign="middle">
            <img src="{LOGO_URL}"
                 alt="혁신제품지원센터"
                 width="160" height="auto"
                 style="display:block;height:auto;max-height:36px;
                        object-fit:contain;object-position:left;">
          </td>
          <td align="right" valign="middle">
            <p style="margin:0;font-size:11px;color:#A0ADB8;letter-spacing:0.3px;">
              {today_str}
            </p>
          </td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- ── 다크 메인 카드 ── -->
  <tr>
    <td style="background:#0F1D2E;
               box-shadow:0 8px 32px rgba(0,0,0,0.18);">

      <!-- 헤더존 -->
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="padding:30px 28px 24px;
                     background:linear-gradient(150deg,#0D1B2A 0%,#132B47 100%);
                     border-bottom:1px solid rgba(255,255,255,0.06);">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td>
                  <p style="margin:0 0 2px;font-size:10px;font-weight:700;
                             letter-spacing:2.5px;color:rgba(255,255,255,0.3);
                             text-transform:uppercase;">
                    Scale-Up Program
                  </p>
                  <h1 style="margin:6px 0 4px;font-size:26px;font-weight:800;
                             color:#FFFFFF;letter-spacing:-0.6px;line-height:1.2;">
                    원스톱 스케일업
                  </h1>
                  <p style="margin:0;font-size:12px;color:rgba(255,255,255,0.38);">
                    맞춤 지원사업 공고 안내
                  </p>
                </td>
                <td align="right" valign="middle" width="72">
                  <div style="background:rgba(74,158,255,0.12);
                              border:1px solid rgba(74,158,255,0.28);
                              border-radius:12px;padding:10px 0;width:60px;
                              text-align:center;">
                    <p style="margin:0;font-size:22px;font-weight:800;color:#10B981;
                               line-height:1;">{len(notices)}</p>
                    <p style="margin:3px 0 0;font-size:9px;letter-spacing:1.2px;
                               color:rgba(74,158,255,0.6);text-transform:uppercase;">picks</p>
                  </div>
                </td>
              </tr>
            </table>
            <!-- 기업명 카드 -->
            <div style="margin-top:20px;padding:14px 18px;
                        background:rgba(255,255,255,0.05);
                        border-radius:8px;border-left:3px solid #10B981;">
              <p style="margin:0 0 3px;font-size:15px;font-weight:700;color:#FFFFFF;">
                {company}
                <span style="font-size:13px;font-weight:400;
                             color:rgba(255,255,255,0.45);margin-left:4px;">담당자님</span>
              </p>
              <p style="margin:0;font-size:12px;color:rgba(255,255,255,0.38);line-height:1.6;">
                기술 키워드 분석을 통해 선별된 공고를 안내드립니다.
              </p>
            </div>
          </td>
        </tr>

        <!-- 공고 목록 -->
        <tr>
          <td style="padding:24px 28px 20px;background:#0F1D2E;">
            {rows_html}
          </td>
        </tr>

        <!-- 캘린더 -->
        {f'''<tr><td style="padding:0 28px 24px;background:#0F1D2E;">{cal_sec}</td></tr>''' if cal_sec else ''}
      </table>
    </td>
  </tr>

  <!-- ── 흰 푸터 ── -->
  <tr>
    <td style="background:#ffffff;border-radius:0 0 14px 14px;
               padding:18px 28px;border-top:1px solid #E8ECF0;">
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td>
            <p style="margin:0;font-size:12px;color:#8A96A3;line-height:1.9;">
              혁신제품지원센터 원스톱 스케일업 운영팀<br>
              <a href="mailto:onestop.kipcc@gmail.com"
                 style="color:#1F4E79;text-decoration:none;font-weight:600;">
                onestop.kipcc@gmail.com
              </a>
            </p>
          </td>
          <td align="right" valign="middle">
            <p style="margin:0;font-size:10px;color:#C5CDD6;letter-spacing:0.5px;">
              수신 동의 기업 대상 발송
            </p>
          </td>
        </tr>
      </table>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body></html>"""

                co_email=""
                if not df_c_cur.empty and '이메일' in df_c_cur.columns:
                    m = df_c_cur[df_c_cur['기업명']==company]
                    if not m.empty: co_email = m.iloc[0].get('이메일','')

                recipients = get_test_recipients() if test_mode else ([co_email] if co_email else [])
                for to in recipients:
                    msg = MIMEMultipart('alternative')
                    msg['From']    = "onestop.kipcc@gmail.com"
                    msg['To']      = to
                    msg['Subject'] = f"[원스톱 스케일업] 맞춤 지원공고 {len(notices)}건 안내 — {company}"
                    msg.attach(MIMEText(html,'html','utf-8'))
                    gmail_send(base64.urlsafe_b64encode(msg.as_bytes()).decode())

                for n in notices:
                    dl  = parse_deadline(n.get('접수기간','')); pid = n.get('공고ID','')
                    if not dl or not pid: continue
                    desc  = f"주관기관: {n.get('주관기관','')}\n공고링크: {n.get('공고링크','')}\n안내기업: {company}"
                    cids  = [cal_id] if cal_id else []
                    if company in ind_cals and ind_cals[company].get('calendar_id'):
                        cids.append(ind_cals[company]['calendar_id'])
                    for cid in cids:
                        try:
                            if cal_list_events(cid, f"pblancId={pid}").get('items'): continue
                        except: pass
                        for days,label in [(0,"마감"),(7,"D-7"),(3,"D-3")]:
                            d = (dl-timedelta(days=days)).strftime('%Y-%m-%d')
                            cal_insert_event(cid, {
                                'summary':f"[{label}] {n.get('공고명','')}",
                                'description':desc,
                                'start':{'date':d,'timeZone':'Asia/Seoul'},
                                'end':  {'date':d,'timeZone':'Asia/Seoul'},
                                'extendedProperties':{'private':{'pblancId':pid}},
                            })

                for n in notices:
                    history_records.append({"기업명":company,"pblancId":n.get('공고ID',''),
                        "공고명":n.get('공고명',''),"발송일":datetime.today().strftime("%Y-%m-%d"),
                        "매칭점수":n.get('점수',''),"담당자검토":"○",
                        "검토의견":n.get('검토의견',''),"신청여부":"","선정결과":""})

                logs.append(f"✅ {company} — {len(notices)}건 발송 완료")
                log.code("\n".join(logs)); prog.progress((idx+1)/len(grouped))

            with st.spinner("발송 이력 드라이브 저장 중..."):
                df_h   = load_excel(drive, HISTORY_FILE)
                df_new = pd.DataFrame(history_records)
                df_fin = pd.concat([df_h,df_new],ignore_index=True) if not df_h.empty else df_new
                save_excel(drive, df_fin, HISTORY_FILE, "발송이력", "375623")

            prog.progress(1.0)
            st.success(f"발송 완료 — {len(history_records)}건 → send_history.xlsx 저장")
            st.session_state['match_results']=[]; st.session_state['review_state']={}


# ══════════════════════════════════════════════════════
# 안내 메일
# ══════════════════════════════════════════════════════
elif page == "안내 메일":
    drive = _get_drive()
    st.title("안내 메일")

    mail_tab1, mail_tab2 = st.tabs(["📨 메일 발송", "📬 회신 확인 (키워드 수집)"])

    with mail_tab2:
        st.subheader("회신 메일 확인 — 키워드 자동 추출")
        st.caption("기업 담당자가 안내 메일에 답장한 내용을 읽어 키워드를 자동 추출합니다.")

        df_c_reply = load_excel(drive, SELECTED_FILE)

        c1, c2 = st.columns(2)
        with c1:
            reply_label = st.text_input("회신 검색 키워드", value="원스톱 스케일업",
                placeholder="발송 메일 제목에 포함된 키워드")
        with c2:
            reply_days = st.number_input("최근 N일 회신 검색", value=30, min_value=1, max_value=90)

        st.caption("💡 기업 담당자가 안내 메일에 **답장(Reply)** 하면 자동으로 여기서 확인됩니다.")

        if st.button("📬 Gmail 회신 검색", type="primary", key="fetch_replies"):
            with st.spinner("Gmail 회신 검색 중..."):
                try:
                    after_date = (datetime.today() - timedelta(days=int(reply_days))).strftime('%Y/%m/%d')
                    query = f'in:inbox after:{after_date}'
                    resp = gapi('GET', 'https://gmail.googleapis.com/gmail/v1/users/me/messages',
                                params={'q': query, 'maxResults': 100})

                    # 응답 상태 디버그 표시
                    if not resp.ok:
                        st.error(f"Gmail API 오류 {resp.status_code}: {resp.text[:200]}")
                    else:
                        all_msgs = resp.json().get('messages', [])
                        st.session_state['reply_msgs'] = all_msgs
                        st.session_state['reply_label_filter'] = reply_label

                        if not all_msgs:
                            st.info(f"최근 {reply_days}일 내 받은 메일이 없습니다.")
                            st.caption(f"검색 쿼리: {query}")
                        else:
                            st.success(f"메일 {len(all_msgs)}건 로드 — 아래에서 확인하세요.")
                except Exception as e:
                    st.error(f"Gmail 검색 실패: {e}")

        def extract_body(payload):
            """멀티파트 메일에서 본문 텍스트 재귀 추출"""
            import base64
            mime = payload.get('mimeType', '')
            if mime == 'text/plain':
                data = payload.get('body', {}).get('data', '')
                if data:
                    return base64.urlsafe_b64decode(data + '==').decode('utf-8', errors='ignore')
            elif mime == 'text/html':
                data = payload.get('body', {}).get('data', '')
                if data:
                    html = base64.urlsafe_b64decode(data + '==').decode('utf-8', errors='ignore')
                    # 태그 제거 후 텍스트만
                    import re as _re
                    return _re.sub(r'<[^>]+>', ' ', html)
            # 멀티파트 재귀
            for part in payload.get('parts', []):
                result = extract_body(part)
                if result and len(result) > 50:
                    return result
            return ''

        # 회신 목록 표시
        if 'reply_msgs' in st.session_state and st.session_state['reply_msgs']:
            st.divider()
            total_msgs = len(st.session_state['reply_msgs'])
            st.subheader(f"회신 목록 ({total_msgs}건)")

            # 선정기업 이메일 목록 미리 구성
            co_emails = {}
            if not df_c_reply.empty and '이메일' in df_c_reply.columns:
                for _, row in df_c_reply.iterrows():
                    em = str(row.get('이메일','')).strip().lower()
                    if em:
                        co_emails[em] = row.get('기업명','')

            # ── 일괄 추출 버튼 ─────────────────────────
            st.info("💡 개별 메일을 열어 '🤖 키워드 자동 추출' 후 저장하거나, 아래 버튼으로 전체 일괄 추출할 수 있습니다.")
            if st.button("🤖 전체 일괄 추출 + 자동 저장", type="primary", key="bulk_extract"):
                import json as _json, re as _re2
                df_bulk = load_excel(drive, SELECTED_FILE)
                prog_bulk = st.progress(0, text="일괄 추출 중...")
                bulk_ok = 0; bulk_fail = 0; bulk_log = []

                for bi, msg_ref in enumerate(st.session_state['reply_msgs'][:50]):
                    try:
                        resp = gapi('GET',
                            f'https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg_ref["id"]}',
                            params={'format': 'full'})
                        if not resp.ok:
                            bulk_fail += 1; continue
                        msg_b = resp.json()
                        hdrs_b = {h['name']: h['value'] for h in msg_b.get('payload',{}).get('headers',[])}
                        sender_b = hdrs_b.get('From','')
                        em_b = _re2.search(r'<(.+?)>', sender_b)
                        if em_b:
                            em_b = em_b.group(1).lower().strip()
                        else:
                            # <> 없으면 전체가 이메일
                            em_b = sender_b.lower().strip().strip('"')
                        co_b = co_emails.get(em_b, '')
                        if not co_b:
                            bulk_log.append(f"⏭ {sender_b[:30]} — 미매칭 스킵")
                            prog_bulk.progress((bi+1)/min(total_msgs,50))
                            continue

                        body_b = extract_body(msg_b.get('payload',{}))
                        for sep in ['On ', '-----', '보낸 사람', 'From:']:
                            idx = body_b.find(sep)
                            if idx > 100: body_b = body_b[:idx]; break
                        body_b = body_b.strip()
                        if not body_b:
                            bulk_log.append(f"⏭ {co_b} — 본문 없음")
                            prog_bulk.progress((bi+1)/min(total_msgs,50))
                            continue

                        prompt_b = f"""기업 담당자 회신에서 키워드를 추출하세요.

회신:
{body_b[:1000]}

JSON만 응답 (코드블록 없이):
{{"기술키워드": ["키워드1", "키워드2"], "필요지원": ["지원유형"], "수출관심국": [], "요약": "요약"}}"""

                        res_b = claude_call_raw(prompt_b)
                        if not res_b or not res_b.strip():
                            bulk_fail += 1
                            bulk_log.append(f"❌ {co_b} — Claude 응답 없음 (API 오류)")
                            prog_bulk.progress((bi+1)/min(total_msgs,50))
                            continue
                        res_b = _re2.sub(r'```(?:json)?\s*', '', res_b).strip().rstrip('`').strip()
                        # { } 사이 내용 추출 (중첩 포함)
                        jm = _re2.search(r'\{[^}]*\}', res_b, _re2.DOTALL)
                        if not jm:
                            jm = _re2.search(r'\{.*\}', res_b, _re2.DOTALL)
                        if not jm:
                            bulk_fail += 1
                            bulk_log.append(f"❌ {co_b} — JSON 파싱 실패: [{res_b[:80]}]")
                            prog_bulk.progress((bi+1)/min(total_msgs,50))
                            continue

                        ext_b = _json.loads(jm.group())
                        kws_b = ext_b.get('기술키워드', [])
                        sup_b = ext_b.get('필요지원', [])
                        exp_b = ext_b.get('수출관심국', [])

                        # 드라이브 저장
                        mask_b = df_bulk['기업명'] == co_b
                        if mask_b.any():
                            existing_b = str(df_bulk.loc[mask_b, '키워드보완'].values[0] or '')
                            new_kws_b  = ', '.join(kws_b + sup_b)
                            df_bulk.loc[mask_b, '키워드보완'] = ', '.join(filter(None, [existing_b, new_kws_b]))
                            if exp_b:
                                existing_exp = str(df_bulk.loc[mask_b, '수출국가'].values[0] or '')
                                df_bulk.loc[mask_b, '수출국가'] = ', '.join(filter(None, [existing_exp, ', '.join(exp_b)]))
                            bulk_ok += 1
                            bulk_log.append(f"✅ {co_b} — {', '.join(kws_b[:3])}")
                        prog_bulk.progress((bi+1)/min(total_msgs,50))

                    except Exception as e_b:
                        bulk_fail += 1
                        bulk_log.append(f"❌ {sender_b[:20]}: {str(e_b)[:30]}")

                # 일괄 저장
                if bulk_ok > 0:
                    save_excel(drive, df_bulk, SELECTED_FILE, "선정기업명단", "1F4E79")
                    st.success(f"✅ {bulk_ok}개사 키워드 저장 완료 / 실패 {bulk_fail}건")
                else:
                    st.warning(f"저장된 기업 없음 (실패 {bulk_fail}건)")

                with st.expander("처리 로그"):
                    for log in bulk_log:
                        st.caption(log)

            st.divider()

            shown = 0
            for i, msg_ref in enumerate(st.session_state['reply_msgs'][:50]):
                try:
                    resp = gapi('GET',
                        f'https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg_ref["id"]}',
                        params={'format': 'full'})
                    if not resp.ok: continue
                    msg = resp.json()

                    hdrs    = {h['name']: h['value'] for h in msg.get('payload', {}).get('headers', [])}
                    subject = hdrs.get('Subject', '')
                    sender  = hdrs.get('From', '')
                    date_str= hdrs.get('Date', '')[:16]

                    import re as _re
                    sender_email_clean = _re.search(r'<(.+?)>', sender)
                    sender_email_clean = sender_email_clean.group(1).lower() if sender_email_clean else sender.lower()

                    # 선정기업 여부 확인
                    matched_co = co_emails.get(sender_email_clean, '')
                    badge = f"✅ {matched_co}" if matched_co else "❓ 미매칭"

                    body_text = extract_body(msg.get('payload', {}))
                    # 이메일 인용 구분선 이전 내용만 (답장 본문)
                    import re as _re
                    for sep in ['On ', '---- ', '-----', '____', '보낸 사람', 'From:']:
                        idx = body_text.find(sep)
                        if idx > 100:
                            body_text = body_text[:idx]
                            break
                    body_text = body_text.strip()

                    shown += 1
                    with st.expander(f"{badge} | **{sender[:30]}** — {subject[:35]} ({date_str})"):

                        # 기업명 자동 매칭 (이미 위에서 계산됨)
                        if matched_co:
                            st.success(f"✅ 매칭 기업: **{matched_co}**")
                        else:
                            st.caption(f"발신: {sender_email_clean}")
                            matched_co = st.selectbox("기업 선택 (수동 매칭)",
                                [''] + (df_c_reply['기업명'].tolist() if not df_c_reply.empty else []),
                                key=f"reply_co_{i}")

                        # 회신 본문 표시
                        if body_text:
                            st.text_area("회신 내용 (답장 본문)", value=body_text[:800],
                                         height=130, key=f"reply_body_{i}", disabled=True)
                        else:
                            st.warning("본문을 읽을 수 없습니다.")

                        # 키워드 추출
                        if body_text and st.button("🤖 키워드 자동 추출", key=f"extract_{i}"):
                            with st.spinner("Claude 분석 중..."):
                                prompt = f"""아래는 지원사업 안내 메일에 대한 기업 담당자의 회신입니다.
메일 본문에 [Q1], [Q2] 형식의 답변이 포함되어 있을 수 있습니다.

회신 내용:
{body_text[:1500]}

위 내용에서 다음을 추출해주세요.
- [Q1] 답변(→ 답변: 이후 텍스트)에서 기업의 주요 기술/제품 키워드
- [Q2] 답변(→ 답변: 이후 텍스트)에서 필요 지원 유형과 목표
- Q1/Q2 형식이 없으면 전체 내용에서 추출

반드시 아래 JSON 형식으로만 응답하세요. 코드블록이나 설명 텍스트 없이 JSON만:
{{"기술키워드": ["키워드1", "키워드2"], "필요지원": ["지원유형1"], "수출관심국": [], "요약": "한 줄 요약"}}"""
                                try:
                                    import json as _json, re as _re2
                                    res = claude_call_raw(prompt)
                                    # 코드블록 제거
                                    res = _re2.sub(r'```(?:json)?\s*', '', res).strip().rstrip('`').strip()
                                    # 중첩 JSON 추출
                                    json_match = _re2.search(r'\{[^{}]*\}', res, _re2.DOTALL)
                                    if not json_match:
                                        json_match = _re2.search(r'\{.*\}', res, _re2.DOTALL)
                                    if json_match:
                                        extracted = _json.loads(json_match.group())
                                        st.session_state[f'extracted_{i}'] = extracted
                                    else:
                                        st.error(f"JSON 파싱 실패 — Claude 응답: {res[:100]}")
                                except Exception as e:
                                    st.error(f"추출 실패: {e}")

                        # 추출 결과
                        if f'extracted_{i}' in st.session_state:
                            ext = st.session_state[f'extracted_{i}']
                            kws = ext.get('기술키워드', [])
                            sup = ext.get('필요지원', [])
                            exp = ext.get('수출관심국', [])

                            col_a, col_b, col_c = st.columns(3)
                            with col_a:
                                st.caption("**기술 키워드**")
                                st.write(', '.join(kws) if kws else '—')
                            with col_b:
                                st.caption("**필요 지원**")
                                st.write(', '.join(sup) if sup else '—')
                            with col_c:
                                st.caption("**수출 관심국**")
                                st.write(', '.join(exp) if exp else '—')
                            st.caption(f"요약: {ext.get('요약','')}")

                            if matched_co and (kws or sup):
                                if st.button(f"💾 {matched_co} 키워드 저장",
                                             key=f"save_reply_{i}", type="primary"):
                                    df_c2 = load_excel(drive, SELECTED_FILE)
                                    mask  = df_c2['기업명'] == matched_co
                                    if mask.any():
                                        existing = str(df_c2.loc[mask, '키워드보완'].values[0] or '')
                                        new_kws  = ', '.join(kws + sup)
                                        merged   = ', '.join(filter(None, [existing, new_kws]))
                                        df_c2.loc[mask, '키워드보완'] = merged
                                        if exp:
                                            # 수출국가 컬럼도 업데이트
                                            existing_exp = str(df_c2.loc[mask, '수출국가'].values[0] or '')
                                            df_c2.loc[mask, '수출국가'] = ', '.join(
                                                filter(None, [existing_exp, ', '.join(exp)]))
                                        if save_excel(drive, df_c2, SELECTED_FILE, "선정기업명단", "1F4E79"):
                                            st.success(f"✅ {matched_co} 키워드 저장 완료 → 다음 매칭에 반영됩니다")
                                        else:
                                            st.error("저장 실패")
                            elif not matched_co:
                                st.warning("저장하려면 위에서 기업을 선택하세요.")

                except Exception as e:
                    continue

            if shown == 0:
                st.info("메일 내용을 불러오는 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요.")

    with mail_tab1:
        info_box("안내 메일",
            """
공고 매칭과 무관한 일반 안내 메일을 선정/예비 기업에게 발송합니다.

**활용 예시**
- 첫 안내 메일 (선정 축하 + 회신 요청)
- 교육 프로그램 수요조사
- 성과집계 조사 요청
- 서류 제출 안내
            """,
            "발송 대상 선택 → 제목/내용 작성 → 미리보기 확인 → 발송")

        with st.spinner("기업 명단 로딩 중..."):
            df_c = load_excel(drive, SELECTED_FILE)

        if df_c.empty:
            st.warning("선정기업 명단이 없습니다. 기업 관리 탭에서 먼저 업로드하세요.")
            st.stop()

        for col in ['수신거부','선정구분']:
            if col not in df_c.columns: df_c[col] = ''

        st.subheader("① 발송 대상 선택")
        has_status = '선정구분' in df_c.columns and df_c['선정구분'].str.strip().ne('').any()

        col1, col2 = st.columns(2)
        with col1:
            if has_status:
                target_group = st.radio(
                    "발송 그룹",
                    ["선정 50개사", "예비 20개사", "전체 70개사", "직접 선택"],
                    horizontal=False, key="notice_mail_group"
                )
            else:
                target_group = "전체"

        # 대상 기업 필터링
        df_active = df_c[df_c['수신거부'] != 'Y'].copy()
        if has_status:
            if target_group == "선정 50개사":
                df_target = df_active[df_active['선정구분'] == '선정']
            elif target_group == "예비 20개사":
                df_target = df_active[df_active['선정구분'] == '예비']
            elif target_group == "직접 선택":
                selected_names = st.multiselect(
                    "기업 직접 선택", df_active['기업명'].tolist(), key="notice_mail_select"
                )
                df_target = df_active[df_active['기업명'].isin(selected_names)]
            else:
                df_target = df_active
        else:
            df_target = df_active

        with col2:
            st.metric("발송 대상", f"{len(df_target)}개사")
            if not df_target.empty:
                email_count = df_target['이메일'].str.strip().ne('').sum()
                st.metric("이메일 보유", f"{email_count}개사")

        st.divider()
        st.subheader("② 메일 내용 작성")

        # 템플릿 선택
        template_choice = st.selectbox("템플릿 선택 (직접 수정 가능)", [
            "직접 작성",
            "첫 안내 메일 (선정 축하 + 회신 요청)",
            "선정 기업 축하 및 프로그램 안내",
            "교육 프로그램 수요조사",
            "성과집계 조사 요청",
            "서류 제출 안내",
        ], key="notice_mail_template")

        TEMPLATES = {
            "첫 안내 메일 (선정 축하 + 회신 요청)": {
                "subject": "[원스톱 스케일업] 2026년 선정을 축하드립니다",
                "body": """안녕하세요, 혁신제품지원센터 원스톱 스케일업 운영팀입니다.

2026년 원스톱 스케일업 프로그램 선정 기업으로 확정되신 것을
진심으로 축하드립니다.

─────────────────────────────────
▣ 원스톱 스케일업이란?
─────────────────────────────────

귀사의 성장 단계와 수요에 맞춰 지원사업 공고 안내부터
역량강화 교육, 개별 컨설팅, 기업 간 협업 매칭까지
필요한 것을 한 곳에서 지원받으실 수 있는 프로그램입니다.

─────────────────────────────────
▣ 주요 프로그램 (7월~11월)
─────────────────────────────────

① 맞춤 공고 정기 안내 (격주)
   귀사 분야에 맞는 지원사업 공고를 선별해 격주로 보내드립니다.
   공고 마감일은 구글 캘린더에 자동 등록되며 D-7, D-3 알림을 받으실 수 있습니다.

② 역량강화 교육 (월 1회, 7월~12월)
   온라인마케팅, MAS 실무, 특허·상표권, 해외ODA 진출 등
   월별 전문가 강의를 안내해드립니다.

③ 개별 담당자 배정
   담당자가 배정되어 목표 확인 및 분기별 체크인을 진행합니다.

④ 기업 간 협업 매칭
   기술협력·마케팅 제휴·해외진출 파트너 연결을 지원합니다.

─────────────────────────────────
▣ 한 가지 부탁드립니다
─────────────────────────────────

귀사에 꼭 맞는 공고를 선별해드리기 위해
이 메일에 답장으로 아래 내용을 채워서 보내주시면 감사하겠습니다.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
[Q1] 귀사 주요 제품이나 기술을 한 줄로 소개해주세요.

예) 실내 공기질 측정 IoT 센서를 제조합니다
예) AI 기반 의료 영상 분석 소프트웨어를 개발합니다
예) 친환경 수처리 필터 소재를 생산합니다

→ 답변:


[Q2] 올해 가장 집중하고 싶은 것을 알려주세요.

예) G-PASS 등록 후 해외조달 시장 진출
예) 수출바우처로 동남아 시장 마케팅
예) 혁신제품 재지정 및 MAS 등록 준비
예) R&D 과제 발굴 및 기술개발

→ 답변:

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

짧게 적어주셔도 충분합니다.
회신 내용을 바탕으로 더 정확한 공고를 선별해드리겠습니다.

─────────────────────────────────

앞으로 귀사의 성장을 함께 돕겠습니다.
궁금한 점은 언제든 이 메일로 답장해주세요.

감사합니다.

혁신제품지원센터 원스톱 스케일업 운영팀
담당: [담당자명] / [연락처] / [이메일]""",
            },
            "선정 기업 축하 및 프로그램 안내": {
                "subject": "[원스톱 스케일업] 2026년 선정 기업 안내 드립니다",
                "body": """안녕하세요, 혁신제품지원센터 원스톱 스케일업 운영팀입니다.

    귀사가 2026년 원스톱 스케일업 프로그램 참여 기업으로 선정되셨음을 진심으로 축하드립니다.

    본 프로그램은 혁신제품 지정기업, G-PASS 기업, 우수조달기업 등 조달·수출 역량을 갖춘 기업의 해외 판로 개척 및 공공조달 시장 진출을 집중 지원하기 위해 기획되었습니다.

    ▣ 주요 지원 내용
    - 맞춤형 지원사업 공고 발굴 및 정기 안내
    - 해외조달시장(G-PASS, KONEPS 연계 등) 진출 컨설팅
    - 조달·수출 분야 역량강화 교육 프로그램
    - 성과 분석 및 지속적 피드백 제공

    향후 운영 일정과 세부 안내는 순차적으로 말씀드릴 예정입니다. 프로그램 운영 기간 동안 적극적인 참여와 관심 부탁드립니다.

    문의사항이 있으시면 아래 연락처로 편하게 연락 주시기 바랍니다.
    감사합니다.""",
            },
            "교육 프로그램 수요조사": {
                "subject": "[원스톱 스케일업] 역량강화 교육 프로그램 수요조사 안내",
                "body": """안녕하세요, 혁신제품지원센터 원스톱 스케일업 운영팀입니다.

    2026년 하반기 역량강화 교육 프로그램 편성을 위해 수요조사를 진행합니다.
    귀사에 실질적으로 도움이 되는 교육을 기획하고자 하오니, 3분 내외의 짧은 응답 부탁드립니다.

    📋 수요조사 참여하기: [링크]
    응답 기한: [응답 기한]

    ▣ 주요 조사 항목
    - 희망 교육 분야 (해외조달, 수출바우처 활용, 혁신제품 지정, IP·인증 등)
    - 선호 교육 방식 (온라인/오프라인, 집합교육/1:1 컨설팅)
    - 희망 교육 시간 및 일정

    수요조사 결과는 교육 프로그램 편성에 직접 반영될 예정입니다.
    바쁘신 와중에도 귀한 시간 내어 주셔서 감사합니다.""",
            },
            "성과집계 조사 요청": {
                "subject": "[원스톱 스케일업] 프로그램 참여 성과 자료 제출 요청",
                "body": """안녕하세요, 혁신제품지원센터 원스톱 스케일업 운영팀입니다.

    2026년 원스톱 스케일업 프로그램 운영 성과 집계를 위해 참여 기업의 성과 현황 파악이 필요합니다.
    아래 안내에 따라 기한 내 자료 제출을 부탁드립니다.

    ▣ 제출 항목
    - 프로그램 참여 후 지원사업 신청·선정 현황
    - 해외 수출 계약 또는 해외조달시장 진입 실적
    - 혁신제품 지정, G-PASS 등 인증 취득 현황
    - 기타 프로그램 활용 성과

    📋 성과 입력 링크: [링크]
    제출 기한: [제출 기한]

    제출해 주신 자료는 프로그램 개선 및 정책 보고 목적으로만 활용되며, 개별 정보는 외부에 공개되지 않습니다.
    협조해 주셔서 감사합니다.""",
            },
            "서류 제출 안내": {
                "subject": "[원스톱 스케일업] 참여 확약서 및 필수 서류 제출 안내",
                "body": """안녕하세요, 혁신제품지원센터 원스톱 스케일업 운영팀입니다.

    2026년 원스톱 스케일업 프로그램 참여를 위한 필수 서류 제출을 안내드립니다.
    아래 항목을 확인하시고 기한 내에 제출하여 주시기 바랍니다.

    ▣ 제출 서류
    1. 프로그램 참여 확약서 (서명 후 스캔 또는 사진 첨부)
    2. 사업자등록증 사본
    3. 기업 소개자료 (제품 및 기술 개요 포함, A4 2페이지 이내)

    📋 서류 제출 방법: [제출 링크 또는 이메일 안내]
    제출 기한: [제출 기한]
    제출처: onestop.kipcc@gmail.com

    기한 내 미제출 시 프로그램 참여가 제한될 수 있으니 반드시 기한을 준수해 주시기 바랍니다.
    서류 관련 문의사항은 아래 운영팀으로 연락 주시기 바랍니다.
    감사합니다.""",
            },
        }

        # 템플릿 선택 시 session_state를 직접 갱신 → text_input/text_area에 즉시 반영
        # (Streamlit에서 value= 파라미터는 최초 렌더링에만 적용되므로 session_state 방식 필요)
        prev_template = st.session_state.get('_prev_notice_template', '')
        if template_choice != prev_template:
            st.session_state['_prev_notice_template'] = template_choice
            if template_choice != "직접 작성" and template_choice in TEMPLATES:
                st.session_state['notice_mail_subject'] = TEMPLATES[template_choice]["subject"]
                st.session_state['notice_mail_body']    = TEMPLATES[template_choice]["body"]
            else:
                st.session_state['notice_mail_subject'] = ""
                st.session_state['notice_mail_body']    = ""
            st.rerun()

        mail_subject = st.text_input("제목", key="notice_mail_subject")
        mail_body    = st.text_area("본문", height=300, key="notice_mail_body",
                                    help="수신자 이름은 자동으로 '[기업명] 담당자님'으로 삽입됩니다.")
        form_link    = st.text_input("📋 구글 폼 링크 (선택사항)", placeholder="https://forms.gle/...",
                                     key="notice_mail_form")

        st.divider()
        st.subheader("③ 발송 미리보기")

        if df_target.empty:
            st.warning("발송 대상 기업이 없습니다.")
        elif not mail_subject or not mail_body:
            st.info("제목과 본문을 작성하면 미리보기가 표시됩니다.")
        else:
            sample_company = df_target.iloc[0]['기업명'] if not df_target.empty else "샘플기업"
            today_str = datetime.today().strftime('%Y.%m.%d')

            form_section = ""
            if form_link:
                form_section = f"""
                <div style="background:rgba(74,158,255,0.1);border:1px solid rgba(74,158,255,0.3);
                            border-radius:8px;padding:16px 18px;margin:16px 0;">
                  <p style="margin:0 0 8px;font-size:11px;font-weight:700;color:#10B981;letter-spacing:1.5px;">
                    📋 설문/폼 링크
                  </p>
                  <a href="{form_link}" style="font-size:13px;color:#10B981;">{form_link}</a>
                </div>"""

            body_html = mail_body.replace('\n', '<br>')

            sample_html = f"""<!DOCTYPE html>
    <html lang="ko">
    <head><meta charset="UTF-8"></head>
    <body style="margin:0;padding:0;background:#F2F4F7;
                 font-family:'Apple SD Gothic Neo','Malgun Gothic',Arial,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#F2F4F7;padding:36px 0 52px;">
    <tr><td align="center">
    <table width="560" cellpadding="0" cellspacing="0">
      <tr>
        <td style="background:#ffffff;border-radius:14px 14px 0 0;
                   padding:20px 28px;border-bottom:1px solid #E8ECF0;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td valign="middle">
                <span style="font-size:15px;font-weight:700;color:#1F4E79;">혁신제품지원센터</span>
              </td>
              <td align="right" valign="middle">
                <p style="margin:0;font-size:11px;color:#A0ADB8;">{today_str}</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="background:#0F1D2E;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:28px 28px 20px;
                         background:linear-gradient(150deg,#0D1B2A 0%,#132B47 100%);
                         border-bottom:1px solid rgba(255,255,255,0.06);">
                <p style="margin:0 0 2px;font-size:10px;font-weight:700;letter-spacing:2.5px;
                           color:rgba(255,255,255,0.3);text-transform:uppercase;">
                  원스톱 스케일업 안내
                </p>
                <h1 style="margin:6px 0 0;font-size:22px;font-weight:800;color:#FFFFFF;line-height:1.3;">
                  {mail_subject}
                </h1>
              </td>
            </tr>
            <tr>
              <td style="padding:24px 28px;background:#0F1D2E;">
                <div style="padding:14px 18px;background:rgba(255,255,255,0.05);
                            border-radius:8px;border-left:3px solid #10B981;margin-bottom:20px;">
                  <p style="margin:0;font-size:14px;font-weight:600;color:#FFFFFF;">
                    {sample_company}
                    <span style="font-size:13px;font-weight:400;color:rgba(255,255,255,0.45);margin-left:4px;">담당자님</span>
                  </p>
                </div>
                <div style="font-size:13px;color:rgba(255,255,255,0.75);line-height:1.9;">
                  {body_html}
                </div>
                {form_section}
              </td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="background:#ffffff;border-radius:0 0 14px 14px;
                   padding:18px 28px;border-top:1px solid #E8ECF0;">
          <p style="margin:0;font-size:12px;color:#8A96A3;line-height:1.9;">
            혁신제품지원센터 원스톱 스케일업 운영팀<br>
            <a href="mailto:onestop.kipcc@gmail.com"
               style="color:#1F4E79;text-decoration:none;font-weight:600;">
              onestop.kipcc@gmail.com
            </a>
          </p>
        </td>
      </tr>
    </table>
    </td></tr></table>
    </body></html>"""

            with st.expander(f"📧 미리보기 — {sample_company}", expanded=True):
                st.components.v1.html(sample_html, height=500, scrolling=True)

            st.divider()
            st.subheader("④ 발송 실행")

            c1, c2 = st.columns(2)
            with c1:
                st.info(f"**발송 대상:** {target_group if has_status else '전체'} ({len(df_target)}개사)")
            with c2:
                no_email = df_target[df_target['이메일'].str.strip() == '']
                if not no_email.empty:
                    st.warning(f"이메일 없는 기업 {len(no_email)}개사는 발송 제외됩니다.")

            # ── 첨부파일 + 기업 키워드 옵션 ──────────────
            opt1, opt2 = st.columns(2)
            with opt1:
                attach_file = st.file_uploader(
                    "📎 첨부파일 (선택)",
                    type=["pdf", "docx", "xlsx", "hwp"],
                    help="공고문 등 첨부할 파일을 선택하세요. 전체 기업에게 동일하게 첨부됩니다."
                )
                if attach_file:
                    st.caption(f"✅ {attach_file.name} ({attach_file.size//1024}KB)")
            with opt2:
                insert_kw = st.checkbox(
                    "📊 기업별 키워드 본문 삽입",
                    value=True,
                    help="각 기업의 관심분야·기술키워드를 메일 본문에 자동 삽입합니다."
                )
                if insert_kw:
                    st.caption("관심분야 / 기술키워드 / 기업유형이 메일에 표시됩니다.")

            if test_mode:
                st.warning("⚠️ 테스트 모드 — 본인 메일로만 발송됩니다.")

            if st.button("📤 안내 메일 발송", type="primary", key="notice_mail_send"):
                from email.mime.multipart import MIMEMultipart
                from email.mime.text import MIMEText
                from email.mime.base import MIMEBase
                from email import encoders
                import base64

                # 첨부파일 미리 읽기
                attach_data = None
                attach_name = None
                if attach_file:
                    attach_data = attach_file.read()
                    attach_name = attach_file.name

                df_send = df_target[df_target['이메일'].str.strip() != '']
                prog = st.progress(0); log_area = st.empty(); logs = []
                ok_count = fail_count = 0

                for i, (_, row) in enumerate(df_send.iterrows()):
                    company  = row['기업명']
                    to_email  = row['이메일'].strip() if not test_mode else get_test_recipients()[0]
                    today_str = datetime.today().strftime('%Y.%m.%d')

                    # 기업별 키워드 섹션 생성
                    kw_section_html = ""
                    kw_section_text = ""
                    if insert_kw:
                        co_interest = str(row.get('관심사업분야', '') or row.get('관심분야', '') or '')
                        co_keywords = str(row.get('기술키워드', '') or '')
                        co_biztype  = str(row.get('기업유형', '') or '')
                        co_kw_extra = str(row.get('키워드보완', '') or '')
                        if any([co_interest, co_keywords, co_biztype]):
                            kw_section_html = f"""
                        <div style="background:#F0FDF4;border:1px solid #BBF7D0;
                                    border-left:4px solid #10B981;border-radius:8px;
                                    padding:16px 18px;margin:16px 0;">
                          <p style="margin:0 0 10px;font-size:11px;font-weight:700;
                                    color:#059669;letter-spacing:1px;">
                            📊 저희가 파악한 귀사 정보
                          </p>
                          {'<p style="margin:4px 0;font-size:13px;color:#0F172A;"><b>관심분야:</b> ' + co_interest + '</p>' if co_interest else ''}
                          {'<p style="margin:4px 0;font-size:13px;color:#0F172A;"><b>기술키워드:</b> ' + co_keywords + '</p>' if co_keywords else ''}
                          {'<p style="margin:4px 0;font-size:13px;color:#0F172A;"><b>기업유형:</b> ' + co_biztype + '</p>' if co_biztype else ''}
                          {'<p style="margin:4px 0;font-size:13px;color:#0F172A;"><b>추가 키워드:</b> ' + co_kw_extra + '</p>' if co_kw_extra else ''}
                          <p style="margin:10px 0 0;font-size:12px;color:#64748B;">
                            위 정보를 기반으로 맞춤 공고를 선별하고 있습니다.<br>
                            다르거나 추가하고 싶은 내용은 Q1, Q2 답변에 자유롭게 적어주세요.
                          </p>
                        </div>"""
                            kw_section_text = f"""
─────────────────────────────────
▣ 저희가 파악한 귀사 정보
─────────────────────────────────
{'관심분야: ' + co_interest if co_interest else ''}
{'기술키워드: ' + co_keywords if co_keywords else ''}
{'기업유형: ' + co_biztype if co_biztype else ''}

위 정보를 기반으로 맞춤 공고를 선별하고 있습니다.
다르거나 추가하고 싶은 내용은 Q1, Q2 답변에 자유롭게 적어주세요.
"""

                    # HTML 본문 생성
                    body_html_co = mail_body.replace('\n', '<br>')

                    # 기업 키워드 카드를 "▣ 한 가지 부탁드립니다" 앞에 삽입
                    if kw_section_html and '▣ 한 가지 부탁드립니다' in body_html_co:
                        body_html_co = body_html_co.replace(
                            '▣ 한 가지 부탁드립니다',
                            f'</div>{kw_section_html}<div style="font-size:13px;color:rgba(255,255,255,0.75);line-height:1.9;">▣ 한 가지 부탁드립니다'
                        )
                        kw_section_html = ""  # 이미 삽입했으므로 하단 중복 방지

                    form_sec_co = ""
                    if form_link:
                        form_sec_co = f"""
                        <div style="background:rgba(74,158,255,0.1);border:1px solid rgba(74,158,255,0.3);
                                    border-radius:8px;padding:16px 18px;margin:16px 0;">
                          <p style="margin:0 0 8px;font-size:11px;font-weight:700;color:#10B981;letter-spacing:1.5px;">
                            📋 설문/폼 링크
                          </p>
                          <a href="{form_link}" style="font-size:13px;color:#10B981;">{form_link}</a>
                        </div>"""

                    html_body = f"""<!DOCTYPE html>
    <html lang="ko"><head><meta charset="UTF-8"></head>
    <body style="margin:0;padding:0;background:#F2F4F7;
                 font-family:'Apple SD Gothic Neo','Malgun Gothic',Arial,sans-serif;">
    <table width="100%" cellpadding="0" cellspacing="0" style="background:#F2F4F7;padding:36px 0 52px;">
    <tr><td align="center">
    <table width="560" cellpadding="0" cellspacing="0">
      <tr>
        <td style="background:#ffffff;border-radius:14px 14px 0 0;
                   padding:20px 28px;border-bottom:1px solid #E8ECF0;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td valign="middle">
                <span style="font-size:15px;font-weight:700;color:#1F4E79;">혁신제품지원센터</span>
              </td>
              <td align="right">
                <p style="margin:0;font-size:11px;color:#A0ADB8;">{today_str}</p>
              </td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="background:#0F1D2E;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td style="padding:28px 28px 20px;
                         background:linear-gradient(150deg,#0D1B2A 0%,#132B47 100%);
                         border-bottom:1px solid rgba(255,255,255,0.06);">
                <p style="margin:0 0 2px;font-size:10px;font-weight:700;letter-spacing:2.5px;
                           color:rgba(255,255,255,0.3);text-transform:uppercase;">
                  원스톱 스케일업 안내
                </p>
                <h1 style="margin:6px 0 0;font-size:22px;font-weight:800;color:#FFFFFF;line-height:1.3;">
                  {mail_subject}
                </h1>
              </td>
            </tr>
            <tr>
              <td style="padding:24px 28px;background:#0F1D2E;">
                <div style="padding:14px 18px;background:rgba(255,255,255,0.05);
                            border-radius:8px;border-left:3px solid #10B981;margin-bottom:20px;">
                  <p style="margin:0;font-size:14px;font-weight:600;color:#FFFFFF;">
                    {company}
                    <span style="font-size:13px;font-weight:400;color:rgba(255,255,255,0.45);margin-left:4px;">담당자님</span>
                  </p>
                </div>
                <div style="font-size:13px;color:rgba(255,255,255,0.75);line-height:1.9;">
                  {body_html_co}
                </div>
                {kw_section_html}
                {form_sec_co}
              </td>
            </tr>
          </table>
        </td>
      </tr>
      <tr>
        <td style="background:#ffffff;border-radius:0 0 14px 14px;
                   padding:18px 28px;border-top:1px solid #E8ECF0;">
          <p style="margin:0;font-size:12px;color:#8A96A3;line-height:1.9;">
            혁신제품지원센터 원스톱 스케일업 운영팀<br>
            <a href="mailto:onestop.kipcc@gmail.com"
               style="color:#1F4E79;text-decoration:none;font-weight:600;">
              onestop.kipcc@gmail.com
            </a>
          </p>
        </td>
      </tr>
    </table>
    </td></tr></table>
    </body></html>"""

                    try:
                        # 첨부파일 있으면 mixed, 없으면 alternative
                        if attach_data:
                            msg = MIMEMultipart('mixed')
                            alt = MIMEMultipart('alternative')
                            alt.attach(MIMEText(html_body, 'html', 'utf-8'))
                            msg.attach(alt)
                            # 첨부파일명 인코딩 (한글/특수문자 대응)
                            from email.utils import encode_rfc2231
                            part = MIMEBase('application', 'octet-stream')
                            part.set_payload(attach_data)
                            encoders.encode_base64(part)
                            # RFC2231 방식으로 파일명 인코딩
                            try:
                                attach_name.encode('ascii')
                                # ASCII만 있으면 그대로
                                part.add_header('Content-Disposition',
                                                f'attachment; filename="{attach_name}"')
                            except UnicodeEncodeError:
                                # 한글 등 비ASCII 파일명 → RFC2231
                                encoded_name = encode_rfc2231(attach_name, charset='utf-8')
                                part.add_header('Content-Disposition',
                                                f"attachment; filename*={encoded_name}")
                            # Content-Type에도 파일명 명시
                            import mimetypes
                            mime_type, _ = mimetypes.guess_type(attach_name)
                            if mime_type:
                                main, sub = mime_type.split('/')
                                part.replace_header('Content-Type',
                                                    f'{mime_type}; name="{attach_name}"')
                            msg.attach(part)
                        else:
                            msg = MIMEMultipart('alternative')
                            msg.attach(MIMEText(html_body, 'html', 'utf-8'))

                        msg['Subject'] = mail_subject if not test_mode else f"[TEST] {mail_subject}"
                        msg['From']    = "onestop.kipcc@gmail.com"
                        msg['To']      = to_email
                        raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
                        gmail_send(raw)
                        ok_count += 1
                        logs.append(f"✅ {company} → {to_email}"
                                    + (" 📎" if attach_data else ""))
                    except Exception as e:
                        fail_count += 1
                        logs.append(f"❌ {company} — {str(e)[:40]}")

                    prog.progress((i+1)/len(df_send))
                    log_area.code("\n".join(logs[-10:]))

                if ok_count > 0:
                    st.success(f"✅ 발송 완료 — 성공 {ok_count}건 / 실패 {fail_count}건")
                if fail_count > 0:
                    st.error(f"실패 {fail_count}건 — 로그를 확인하세요.")



        drive = _get_drive()
        st.title("발송 이력")
        info_box("발송 이력",
            """
    발송 완료 건 기록 + 성과 추적

    **자동 기록** — 기업명·공고명·발송일·매칭점수
    **담당자 직접 입력** — 신청여부(Y/N)·선정결과(선정/미선정/대기)

    **중복 발송 방지** — 이 이력 기반으로 이미 발송한 공고는 다음 매칭에서 자동 제외
    재발송 필요 시 해당 행 삭제 후 저장
            """,
            "특정 행 삭제 → 표에서 행 선택 후 Delete → '드라이브 저장' 클릭")

        with st.spinner("드라이브 이력 로딩 중..."):
            df_h = load_excel(drive, HISTORY_FILE)

        if df_h.empty:
            st.info("발송 이력 없음")
        else:
            c1,c2,c3 = st.columns(3)
            c1.metric("총 발송", len(df_h))
            c2.metric("신청 건", (df_h['신청여부']=='Y').sum() if '신청여부' in df_h.columns else 0)
            c3.metric("선정 건", (df_h['선정결과']=='선정').sum() if '선정결과' in df_h.columns else 0)
            st.divider()

            edited = st.data_editor(df_h, use_container_width=True, hide_index=True,
                column_config={
                    "신청여부":st.column_config.SelectboxColumn("신청여부",options=["","Y","N"]),
                    "선정결과":st.column_config.SelectboxColumn("선정결과",options=["","선정","미선정","대기"]),
                })
            if st.button("💾 드라이브 저장"):
                with st.spinner("저장 중..."):
                    save_excel(drive, edited, HISTORY_FILE, "발송이력", "375623")
                st.success("저장 완료")

            st.divider()
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as w: df_h.to_excel(w, index=False)
            st.download_button("📥 엑셀 다운로드", buf.getvalue(), HISTORY_FILE,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

elif page == "설정":
    drive = _get_drive()
    st.title("설정")
    st.caption("키워드·매칭 설정 / 드라이브 연동 / 인증 관리")

    tab_s1, tab_s2, tab_s3, tab_s4, tab_s5 = st.tabs([
        "🎯 축1 — 지원대상", "📋 축2 — 사업성격",
        "🏢 기업별 키워드", "⚖️ 매칭 가중치", "🔧 시스템 설정"
    ])
    kw_data = load_json(drive, KEYWORDS_FILE) or {}
    HIGH_kw, MID_kw = load_keywords(drive)
    df_c = load_excel(drive, SELECTED_FILE)

    # ── 탭1: 축1 지원대상 ──────────────────────────────
    with tab_s1:
        st.subheader("축1 — 지원대상 키워드")
        st.caption("공고가 어떤 기업을 대상으로 하는지 판단합니다. 이 키워드가 공고에 있을 때 해당 카테고리로 분류됩니다.")

        current_target = kw_data.get("TARGET_KW", TARGET_KW)
        t_inputs = {}
        label_map1 = {
            "조달기업특화": "🏆 조달기업특화 (G-PASS·조달청·혁신제품 대상 공고) → ★★★",
            "수출기업특화": "🌏 수출기업특화 (수출·글로벌 대상 공고) → ★★★",
            "중소벤처일반": "🏢 중소벤처일반 (일반 중소·벤처 대상 공고) → ★★",
        }
        for cat, kws in current_target.items():
            st.markdown(f"**{label_map1.get(cat, cat)}**")
            st.caption(f"현재 {len(kws)}개 키워드")
            t_inputs[cat] = st.text_area(
                f"{cat}", value=", ".join(kws), height=80,
                key=f"tkw_{cat}", label_visibility="collapsed",
                placeholder="쉼표로 구분하여 입력"
            )
            st.divider()

        if st.button("💾 축1 키워드 저장", type="primary", key="save_target_kw"):
            new_target = {}
            for cat in current_target:
                new_target[cat] = [k.strip() for k in t_inputs[cat].split(',') if k.strip()]
            kw_data["TARGET_KW"] = new_target
            if save_json(drive, kw_data, KEYWORDS_FILE):
                total = sum(len(v) for v in new_target.values())
                st.success(f"축1 저장 완료 — 총 {total}개 키워드")
            else:
                st.error("저장 실패")

    # ── 탭2: 축2 사업성격 ──────────────────────────────
    with tab_s2:
        st.subheader("축2 — 사업성격 키워드")
        st.caption("공고가 어떤 종류의 지원인지 판단합니다. 축1과 교차하여 별점이 결정됩니다.")

        current_type = kw_data.get("TYPE_KW", TYPE_KW)
        ty_inputs = {}
        label_map2 = {
            "공공조달":   "🏛 공공조달 (조달청·시범구매·MAS) → ★★★",
            "해외진출":   "🌏 해외진출 (수출바우처·해외판로) → ★★★",
            "마케팅홍보": "📢 마케팅·홍보 (전시회·박람회·홍보) → ★★",
            "인증특허":   "📋 인증·특허 (해외인증·IP) → ★★",
            "기술개발":   "🔬 기술개발 (R&D·기술사업화) → ★★",
            "금융융자":   "💰 금융·융자 (정책자금·보증) → ★★",
            "내수판로":   "🛒 내수판로 (온라인몰·유통) → ★★",
            "인력채용":   "👥 인력·채용 (청년채용·인재양성) → ★★",
        }
        for cat, kws in current_type.items():
            st.markdown(f"**{label_map2.get(cat, cat)}**")
            st.caption(f"현재 {len(kws)}개 키워드")
            ty_inputs[cat] = st.text_area(
                f"{cat}", value=", ".join(kws), height=80,
                key=f"tykw_{cat}", label_visibility="collapsed",
                placeholder="쉼표로 구분하여 입력"
            )
            st.divider()

        if st.button("💾 축2 키워드 저장", type="primary", key="save_type_kw"):
            new_type = {}
            for cat in current_type:
                new_type[cat] = [k.strip() for k in ty_inputs[cat].split(',') if k.strip()]
            kw_data["TYPE_KW"] = new_type
            if save_json(drive, kw_data, KEYWORDS_FILE):
                total = sum(len(v) for v in new_type.values())
                st.success(f"축2 저장 완료 — 총 {total}개 키워드")
            else:
                st.error("저장 실패")

    # ── 탭3: 기업별 키워드 ────────────────────────────
    with tab_s3:
        st.subheader("기업별 키워드 보완")
        st.caption("각 기업에 추가 키워드를 설정합니다. 기업 관리 탭과 동일한 필드이며 여기서 일괄 편집 가능합니다.")

        # 매칭결과 기반 기본 키워드 (분석 결과로 자동 생성)
        DEFAULT_CO_KW = {
            "(주)더세이프": "친환경, 환경, 에너지, 운송, 대기, 탄소중립",
            "(주)빛가람시스템": "AI, 정보통신, 통신, 네트워크, ICT, 클라우드",
            "(주)새영테크놀로지": "환경, 에너지, 탄소, 저탄소, ESG, 대기",
            "(주)세이프텍리서치": "국방, 방어, 방산, AI, 데이터, 데이터분석",
            "(주)아이서티": "정보통신, 통신, 인공지능, 데이터플랫폼, AI, 데이터",
            "(주)아이쉐어넷": "모빌리티, 운송, 자율주행, 차량, 전자, 진단",
            "(주)에코라이프": "탄소, 대기, 녹색, 환경, 에너지, 탄소중립",
            "(주)엠피웨이브": "정보통신, 통신, IoT, ICT, AI, 스마트",
            "(주)오썸피아": "해외조달, 정보통신, 통신, AI, IoT, 교육",
            "(주)오투엔비": "탄소, 대기, 환경, 에너지, 녹색, 제조",
            "(주)케이씨티이엔씨": "건설, 건설자재, 통신, 보안, 정보통신",
            "(주)테스토닉": "전자, 환경, 헬스, 디지털헬스, 탄소, 대기",
            "(주)피투에스지글로벌": "AI, 전자, 반도체, 물류, 글로벌",
            "대성": "소재, 생산, 장비, 제조, 수산, 수산물",
            "블루센 주식회사": "전자, 배터리, 센서, 대기, 탄소, 녹색",
            "비웨이브(주)": "바이오, 진단, 의료, 헬스케어, 헬스, 정보통신",
            "스마트이앤씨": "정보통신, 통신, 소프트웨어, SW, AI",
            "아이씨피(주)": "방산, 운송, 물류, 공급망, 국방, 인프라",
            "에스엠테크": "탄소, 대기, 환경, 에너지, 로봇, 장비",
            "에이트스튜디오 주식회사": "AI, 네트워크, 헬스, 디지털헬스, IoT, 스마트",
            "유비컨트롤 주식회사": "제조, 소재, 기계, 장비, 설비, 생산",
            "이노넷(주)": "AI, 드론, 건설, 보안, 물류, 공급망",
            "인플랩 주식회사": "금융, 인프라, 정보통신, 통신, 물류, 물류센터",
            "주식회사 동성이에스": "탄소중립, 탄소, 대기, 환경, 에너지, 전자",
            "주식회사 리얼디자인테크": "AI, 헬스, 디지털헬스, IoT, 진단, 의료기기",
            "주식회사 바르미페인트": "모빌리티, 자율주행, 대기, 해외조달, 친환경, 환경",
            "주식회사 바이오바이츠": "바이오, 헬스케어, 진단, 헬스, 정보통신, 통신",
            "주식회사 아스팔트아트": "정보통신, 통신, 5G, 인프라, 스마트물류",
            "주식회사 애니락": "건축, 환경, 대기",
            "주식회사 에어딥": "환경, 대기, 데이터센터, 5G, 에너지, ESG",
            "주식회사 엠쓰리시스템즈": "장비, 제조, 소재, 건설, 건설자재, 제조업",
            "주식회사 옥스": "전자, 배터리, 센서, AI, 정보통신, 통신",
            "주식회사 와트": "해외전시, AI솔루션, 지능형, 데이터분석, 통신",
            "주식회사 이지서티": "정보통신, 통신, 금융, AI, AI솔루션, 지능형",
            "주식회사 제이씨에프테크놀러지": "해외조달, 교육, 인재, AI, 헬스, 디지털헬스",
            "주식회사 칼렛바이오": "친환경, 환경, 에너지, 대기, 모빌리티, 자율주행",
            "주식회사 케이스마트피아": "대기, 환경, 에너지",
            "주식회사 퀀텀바이오닉스": "환경, 탄소중립, 온실가스, 탄소, 녹색, 기후",
            "주식회사 클리어창": "대기, 탄소, 친환경",
            "주식회사 트레시스": "전자, 대기, 탄소, 배터리, 센서, 이차전지",
            "주식회사 필메디": "바이오, 의료, 의료기기, 헬스, 디지털헬스, 의약",
            "주식회사 휴마스터": "수소, 탄소중립, 탄소, 대기, 온실가스, 녹색",
            "창조 엔지니어링": "수산, 수산물, 식품, 축산, 해양수산, 환경",
            "탐투스(주)": "해외조달, 전자, 환경, 탄소, 대기, 배터리",
            "파이어버스터Lab": "모빌리티, 자율주행, 차량, 배터리, 자동차부품",
            "팔수": "헬스, 디지털헬스, 진단, 바이오, 의료, 의약",
            "피아스페이스(주)": "AI, 정보통신, 통신, 데이터분석, 디지털전환",
            "주식회사 쎈인더스트리": "제조, 장비, 설비, 기계",
            "주식회사 애니락": "건축, 건설자재, 환경",
            "주식회사 쎈인더스트리": "제조, 장비, 설비, 스마트제조",
        }

        if df_c.empty:
            st.warning("선정기업 명단이 없습니다.")
        else:
            if '키워드보완' not in df_c.columns:
                df_c['키워드보완'] = ''

            # 키워드 없는 기업에 DEFAULT_CO_KW 자동 채우기 버튼
            no_kw_count = (df_c['키워드보완'].fillna('') == '').sum()
            if no_kw_count > 0:
                c1, c2 = st.columns([2, 3])
                with c1:
                    st.warning(f"⚠️ {no_kw_count}개사 키워드 미입력")
                with c2:
                    if st.button("🤖 분석 기반 키워드 자동 채우기", key="auto_fill_kw"):
                        filled = 0
                        for idx, row in df_c.iterrows():
                            co = row.get('기업명','')
                            if not row.get('키워드보완','') and co in DEFAULT_CO_KW:
                                df_c.at[idx, '키워드보완'] = DEFAULT_CO_KW[co]
                                filled += 1
                        if filled > 0:
                            if save_excel(drive, df_c, SELECTED_FILE, "선정기업명단", "1F4E79"):
                                st.success(f"✅ {filled}개사 키워드 자동 입력 완료")
                                st.rerun()
            else:
                st.success("✅ 전체 기업 키워드 입력 완료")

            search = st.text_input("🔍 기업명 검색", key="kw_co_search")
            df_show = df_c[df_c['기업명'].str.contains(search, na=False)] if search else df_c

            changed = {}
            for idx, row in df_show.iterrows():
                c1, c2 = st.columns([2, 3])
                with c1:
                    nm = row.get('기업명', '')
                    st.write(f"**{nm}**")
                    st.caption(f"{row.get('제품분야','—')[:20]} | {row.get('소재지','—')}")
                with c2:
                    cur_kw = row.get('키워드보완','')
                    hint   = DEFAULT_CO_KW.get(nm, '')
                    new_kw = st.text_input(
                        "추가 키워드",
                        value=cur_kw,
                        key=f"co_kw_{idx}",
                        placeholder=hint if hint else "예: 스마트팜, IoT 센서",
                        label_visibility="collapsed"
                    )
                    if new_kw != cur_kw:
                        changed[idx] = new_kw

            if changed:
                if st.button(f"💾 변경된 {len(changed)}개사 키워드 저장",
                             type="primary", key="save_co_kw"):
                    for idx, val in changed.items():
                        df_c.at[idx, '키워드보완'] = val
                    if save_excel(drive, df_c, SELECTED_FILE, "선정기업명단", "1F4E79"):
                        st.success(f"{len(changed)}개사 키워드 저장 완료")
                        st.rerun()
                    else:
                        st.error("저장 실패")

        # 피드백 패턴 (이전 검토에서 쌓인 데이터)
        feedback = kw_data.get('feedback', {})
        if feedback:
            st.divider()
            st.subheader("📋 피드백 패턴 (제외 이력)")
            st.caption("이전 검토에서 제외한 공고의 사업성격 패턴입니다. 다음 매칭 시 자동 감점됩니다.")
            for co, patterns in list(feedback.items())[:10]:
                st.write(f"**{co}**: " + ", ".join([f"{k}({v}회)" for k,v in sorted(patterns.items(), key=lambda x:-x[1])]))

    # ── 탭4: 매칭 가중치 ──────────────────────────────
    with tab_s4:
        st.subheader("매칭 축별 가중치 조정")
        st.caption("각 축의 점수 비중을 조정합니다. 합산 점수가 별점(★) 판정에 영향을 줍니다.")

        weights = kw_data.get("weights", {
            "지원대상": 3, "사업성격": 2, "기업키워드": 2,
            "핵심수요": 3, "업종역방향": 1, "소재지가산": 3, "세그먼트": 2
        })

        st.markdown("**각 매칭 축의 기본 점수 (건당)**")
        c1, c2 = st.columns(2)
        with c1:
            w_target  = st.slider("지원대상 매칭 (×건수)", 1, 6, weights.get("지원대상", 3), key="w1")
            w_type    = st.slider("사업성격 매칭 (×건수)", 1, 6, weights.get("사업성격", 2), key="w2")
            w_kw      = st.slider("기업키워드 매칭 (×건수)", 1, 6, weights.get("기업키워드", 2), key="w3")
            w_demand  = st.slider("핵심수요 매칭 (×건수)", 1, 6, weights.get("핵심수요", 3), key="w4")
        with c2:
            w_ind     = st.slider("업종 역방향 매칭 (×건수)", 1, 6, weights.get("업종역방향", 1), key="w5")
            w_loc     = st.slider("소재지 일치 가산점", 1, 6, weights.get("소재지가산", 3), key="w6")
            w_seg     = st.slider("세그먼트 부스트 (최대)", 1, 8, weights.get("세그먼트", 4), key="w7")

        st.divider()
        st.subheader("★ 별점 판정 기준")
        col1, col2 = st.columns(2)
        with col1:
            star3_threshold = st.slider("★★★ 최소 점수", 10, 40, kw_data.get("star3_threshold", 20), key="s3")
        with col2:
            star2_threshold = st.slider("★★ 최소 점수",  5, 30, kw_data.get("star2_threshold", 10), key="s2")
        st.caption(f"현재 설정: ★★★ ≥ {star3_threshold}점, ★★ ≥ {star2_threshold}점")

        if st.button("💾 가중치 저장", type="primary", key="save_weights"):
            kw_data["weights"] = {
                "지원대상": w_target, "사업성격": w_type,
                "기업키워드": w_kw, "핵심수요": w_demand,
                "업종역방향": w_ind, "소재지가산": w_loc, "세그먼트": w_seg
            }
            kw_data["star3_threshold"] = star3_threshold
            kw_data["star2_threshold"] = star2_threshold
            if save_json(drive, kw_data, KEYWORDS_FILE):
                st.success("가중치 저장 완료 — 다음 매칭 실행 시 반영됩니다.")
            else:
                st.error("저장 실패")

        with st.expander("ℹ️ 가중치 설명"):
            st.markdown("""
**지원대상 매칭** — 공고가 "수출기업", "벤처기업" 등 기업 유형을 명시한 경우

**사업성격 매칭** — 공고가 "해외진출", "기술개발" 등 사업 유형에 해당하는 경우

**기업키워드 매칭** — 기업의 기술키워드/제품분야가 공고 텍스트에 직접 등장하는 경우

**핵심수요 매칭** — 기업의 핵심수요태그(G-PASS, 수출바우처 등)가 공고와 일치하는 경우

**업종 역방향 매칭** — 기업 제품분야 → INDUSTRY_KW → 공고 텍스트로 간접 매칭된 경우

**소재지 일치** — 지역 제한 공고에서 기업 소재지가 일치하는 경우

**세그먼트 부스트** — 기업 세그먼트(해외진출형/조달강화형 등)에 맞는 공고 유형 보너스
            """)

    # ── 탭5: 시스템 설정 ──────────────────────────────
    with tab_s5:
        st.subheader("🔧 시스템 설정")

        # 드라이브 연동 상태
        st.markdown("**📁 드라이브 연동 현황**")
        st.markdown(f"[📂 드라이브 폴더 열기](https://drive.google.com/drive/folders/{DRIVE_FOLDER_ID})")
        files_check = {
            SELECTED_FILE: "선정기업 명단",
            NOTICES_FILE:  "공고 DB",
            HISTORY_FILE:  "발송 이력",
            KEYWORDS_FILE: "키워드 설정",
            AI_ANALYSIS_FILE: "AI 분석 결과",
            INDCAL_FILE:   "기업별 캘린더",
        }
        cols = st.columns(2)
        for i, (fname, label) in enumerate(files_check.items()):
            fid = drive_file_id(drive, fname)
            cols[i%2].write(f"{'✅' if fid else '❌'} {label} (`{fname}`)")

        st.divider()

        # 인증 상태
        st.markdown("**🔑 구글 인증 상태**")
        try:
            test_resp = gapi('GET', 'https://www.googleapis.com/drive/v3/about',
                             params={'fields': 'user'})
            if test_resp.ok:
                user_info = test_resp.json().get('user', {})
                st.success(f"✅ 인증 정상 — {user_info.get('emailAddress', '확인됨')}")
            else:
                st.error(f"❌ 인증 실패 ({test_resp.status_code}) — 토큰 갱신 필요")
        except Exception as e:
            st.error(f"❌ 인증 오류: {e}")

        st.divider()

        # 토큰 갱신 안내
        st.markdown("**🔄 토큰 갱신 방법**")
        st.info("""
토큰이 만료되면 드라이브/Gmail/캘린더 기능이 모두 중단됩니다.

**갱신 순서:**
1. 회사 PC에서 `python refresh_token.py` 실행
2. 브라우저 구글 로그인 → 권한 허용
3. `token_new.json` 생성 확인
4. Streamlit Cloud → Settings → Secrets → `token` 값 교체
5. 앱 Reboot
        """)

        st.divider()
        st.markdown("**ℹ️ 앱 정보**")
        st.caption(f"GitHub: onestopkipcc-web/scaleup-matching")
        st.caption(f"배포: Streamlit Cloud")
        st.caption(f"드라이브 폴더 ID: {DRIVE_FOLDER_ID}")



# ══════════════════════════════════════════════════════
# 캘린더 관리
# ══════════════════════════════════════════════════════
elif page == "캘린더":
    drive = _get_drive()
    st.title("캘린더 관리")
    info_box("캘린더 관리",
        """
기업별 구글 캘린더를 생성하고 담당자에게 공유합니다.

**운영 흐름**
1. 기업 담당자 구글계정 수집 (기업 관리 탭 또는 DB 업로드)
2. 캘린더 일괄 생성 → individual_calendars.json 저장
3. 담당자 이메일로 공유 초대 자동 발송
4. 이후 발송 시 공고 마감일이 캘린더에 자동 등록됨
        """,
        "테스트 먼저 → 전체 실행")

    tab_cal1, tab_cal2, tab_cal3 = st.tabs([
        "🧪 테스트 생성", "📋 기업별 현황", "⚙️ 전체 실행"
    ])

    # ── 공통 데이터 로드 ──────────────────────────────
    df_c       = load_excel(drive, SELECTED_FILE)
    ind_cals   = load_json(drive, INDCAL_FILE) or {}

    # ── 탭1: 테스트 생성 ──────────────────────────────
    with tab_cal1:
        st.subheader("테스트용 캘린더 생성")
        st.caption("실제 기업 DB와 무관하게 테스트 계정으로 먼저 동작 확인합니다.")

        with st.form("test_cal_form"):
            t_company  = st.text_input("테스트 기업명", value="테스트기업_홍길동",
                                        placeholder="캘린더 이름에 표시될 기업명")
            t_email    = st.text_input("공유할 구글 계정",
                                        placeholder="example@gmail.com",
                                        help="이 계정으로 캘린더 공유 초대가 발송됩니다")
            t_interest = st.text_input("관심분야 (선택)", placeholder="수출, 기술개발")
            t_keyword  = st.text_input("기술키워드 (선택)", placeholder="AI, IoT")
            submitted  = st.form_submit_button("🗓 테스트 캘린더 생성", type="primary")

        if submitted:
            if not t_email or '@' not in t_email:
                st.error("구글 계정 이메일을 올바르게 입력해주세요.")
            else:
                with st.spinner("캘린더 생성 중..."):
                    try:
                        cal_id = cal_create(
                            f"[원스톱] {t_company} 지원공고 D-day",
                            f"혁신제품지원센터 원스톱 스케일업 프로그램\n기업: {t_company}\n관심분야: {t_interest}\n기술키워드: {t_keyword}\n문의: onestop.kipcc@gmail.com"
                        )
                        if not cal_id:
                            st.error("캘린더 생성 실패 — 인증 상태를 확인하세요.")
                        else:
                            sub_link = f"https://calendar.google.com/calendar?cid={cal_id}"
                            st.success("✅ 캘린더 생성 완료!")
                            st.code(f"캘린더 ID: {cal_id}")

                            # 공유 초대
                            if cal_share(cal_id, t_email):
                                st.success(f"✅ {t_email} 으로 공유 초대 발송 완료!")
                            else:
                                st.warning("⚠️ 공유 초대 실패 — 이메일을 확인하세요.")
                            st.markdown(f"[📅 캘린더 구독 링크]({sub_link})")

                            # 테스트 이벤트 등록
                            from datetime import timedelta
                            test_date = (datetime.today() + timedelta(days=7)).strftime('%Y-%m-%d')
                            event = {
                                'summary': f"[D-7] 테스트공고 마감 — {t_company}",
                                'description': "원스톱 스케일업 테스트 이벤트입니다.",
                                'start': {'date': test_date, 'timeZone': 'Asia/Seoul'},
                                'end':   {'date': test_date, 'timeZone': 'Asia/Seoul'},
                            }
                            cal_insert_event(cal_id, event)
                            st.success(f"✅ 테스트 이벤트 등록 완료 ({test_date} D-7 알림)")
                            st.info("📱 구글 캘린더 앱에서 공유 초대를 수락하면 캘린더가 추가됩니다.")

                    except Exception as e:
                        st.error(f"오류 발생: {str(e)}")

    # ── 탭2: 기업별 현황 ──────────────────────────────
    with tab_cal2:
        st.subheader("기업별 캘린더 현황")

        if not ind_cals:
            st.info("아직 생성된 캘린더가 없습니다. '전체 실행' 탭에서 생성하세요.")
        else:
            total    = len(ind_cals)
            shared   = sum(1 for v in ind_cals.values() if v.get('shared'))
            noshare  = total - shared

            c1, c2, c3 = st.columns(3)
            c1.metric("전체 캘린더", f"{total}개")
            c2.metric("✅ 공유 완료", f"{shared}개")
            c3.metric("⏳ 공유 대기", f"{noshare}개")
            st.divider()

            for co, info in ind_cals.items():
                status = "✅" if info.get('shared') else "⏳"
                with st.expander(f"{status} {co}"):
                    c1, c2 = st.columns(2)
                    c1.caption("구글 계정")
                    c1.write(info.get('google_account', '미입력') or '미입력')
                    c2.caption("공유 상태")
                    c2.write("공유 완료" if info.get('shared') else "미공유")
                    if info.get('subscribe_link'):
                        st.markdown(f"[📅 구독 링크]({info['subscribe_link']})")

                    # 미공유 기업 수동 공유 버튼
                    if not info.get('shared') and info.get('google_account'):
                        if st.button(f"📨 {info['google_account']}에 공유", key=f"share_{co}"):
                            try:
                                ok = cal_share(info['calendar_id'], info['google_account'])
                                if ok:
                                    ind_cals[co]['shared'] = True
                                    import json as _json
                                    content = _json.dumps(ind_cals, ensure_ascii=False).encode('utf-8')
                                    drive_upload(drive, INDCAL_FILE, content, "application/json")
                                    st.success("공유 완료!")
                                    st.rerun()
                                else:
                                    st.error("공유 실패 — 구글 계정을 확인하세요.")
                            except Exception as e:
                                st.error(f"공유 실패: {e}")

        # DB 업로드 섹션
        st.divider()
        st.subheader("📁 individual_calendars.json 직접 업로드")
        st.caption("로컬에서 create_individual_calendars.py 실행 후 생성된 파일을 업로드하세요.")
        uploaded_cal = st.file_uploader("individual_calendars.json 업로드",
                                         type=["json"], key="cal_upload")
        if uploaded_cal:
            import json as _json
            cal_data = _json.loads(uploaded_cal.read().decode('utf-8'))
            content  = _json.dumps(cal_data, ensure_ascii=False).encode('utf-8')
            if drive_upload(drive, INDCAL_FILE, content, "application/json"):
                st.success(f"✅ {len(cal_data)}개사 캘린더 정보 업로드 완료")
                st.rerun()

    # ── 탭3: 전체 실행 ────────────────────────────────
    with tab_cal3:
        st.subheader("전체 기업 캘린더 일괄 생성")

        if df_c.empty:
            st.warning("선정기업 명단이 없습니다.")
        else:
            has_google = '구글계정' in df_c.columns
            google_filled = (df_c['구글계정'].fillna('') != '').sum() if has_google else 0
            no_google     = len(df_c) - google_filled if has_google else len(df_c)

            c1, c2, c3 = st.columns(3)
            c1.metric("전체 기업", f"{len(df_c)}개사")
            c2.metric("구글계정 입력", f"{google_filled}개사")
            c3.metric("계정 미입력", f"{no_google}개사")

            if no_google > 0:
                st.warning(f"⚠️ {no_google}개사는 구글계정이 없어 캘린더 공유가 불가합니다. "
                           f"기업 관리 탭에서 구글계정을 먼저 입력하세요.")
            already_created = len(ind_cals)
            remaining = len(df_c) - already_created
            if already_created > 0:
                st.info(f"이미 생성된 캘린더: {already_created}개사 / 새로 생성할: {remaining}개사")

            st.divider()
            st.warning("⚠️ 실행 전 테스트 탭에서 먼저 동작을 확인하세요.")

            col1, col2 = st.columns(2)
            with col1:
                confirm_text = st.text_input("확인코드 입력", placeholder="캘린더생성",
                                              key="cal_create_confirm")
            with col2:
                st.caption("확인코드: **캘린더생성**")

            if st.button("🗓 전체 캘린더 일괄 생성", type="primary", key="create_all_cal"):
                if confirm_text != "캘린더생성":
                    st.error("확인코드를 입력해주세요.")
                else:
                    prog = st.progress(0, text="캘린더 생성 중...")
                    logs = []; ok = 0; fail = 0

                    for i, (_, row) in enumerate(df_c.iterrows()):
                        co       = row.get('기업명', '')
                        g_email  = str(row.get('구글계정', '') or '').strip()
                        interest = str(row.get('관심사업분야', '') or '')[:30]
                        keywords = str(row.get('기술키워드', '') or '')[:40]

                        if co in ind_cals:
                            logs.append(f"⏭ {co} — 이미 생성됨")
                            prog.progress((i+1)/len(df_c), text=f"{i+1}/{len(df_c)} 처리 중...")
                            continue

                        try:
                            cal_id = cal_create(
                                f"[원스톱] {co} 지원공고 D-day",
                                f"혁신제품지원센터 원스톱 스케일업 프로그램\n기업: {co}\n관심분야: {interest}\n기술키워드: {keywords}\n문의: onestop.kipcc@gmail.com"
                            )
                            if not cal_id:
                                raise Exception("캘린더 생성 실패")

                            sub_link = f"https://calendar.google.com/calendar?cid={cal_id}"
                            shared   = False

                            if g_email and '@' in g_email:
                                shared = cal_share(cal_id, g_email)

                            ind_cals[co] = {
                                'calendar_id':    cal_id,
                                'google_account': g_email,
                                'shared':         shared,
                                'subscribe_link': sub_link,
                                'interest':       interest,
                            }
                            ok += 1
                            logs.append(f"✅ {co}" + (f" → {g_email} 공유완료" if shared else " (공유 대기)"))

                            # 5개마다 중간 저장
                            if ok % 5 == 0:
                                import json as _json
                                _c = _json.dumps(ind_cals, ensure_ascii=False).encode('utf-8')
                                drive_upload(drive, INDCAL_FILE, _c, "application/json")

                        except Exception as e:
                            fail += 1
                            logs.append(f"❌ {co}: {str(e)[:40]}")

                        prog.progress((i+1)/len(df_c), text=f"{i+1}/{len(df_c)} 처리 중...")

                    # 최종 저장
                    import json as _json
                    _c = _json.dumps(ind_cals, ensure_ascii=False).encode('utf-8')
                    drive_upload(drive, INDCAL_FILE, _c, "application/json")

                    st.success(f"✅ 완료 — 생성 {ok}개 / 실패 {fail}개")
                    with st.expander("처리 로그"):
                        for log in logs:
                            st.caption(log)
                    st.rerun()



elif page == "시스템 명세":
    st.title("시스템 명세")
    st.caption("원스톱 스케일업 공고 매칭 시스템 — 구조 및 운영 가이드")

    tab1, tab2, tab3, tab4 = st.tabs(["전체 흐름", "매칭 로직", "키워드 구조", "파일 구조"])

    # ── 탭1: 전체 흐름 ──────────────────────────────
    with tab1:
        st.subheader("원스톱 스케일업 시스템 전체 흐름")

        st.markdown("""
<div style="background:#F8FAFC;border:1px solid #E2E8F0;border-radius:12px;padding:24px;font-family:'Apple SD Gothic Neo',sans-serif;">

<h4 style="color:#0F172A;margin:0 0 20px;">📌 전체 파이프라인</h4>

<div style="display:flex;flex-direction:column;gap:4px;">

<!-- 단계 1 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#10B981;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 1</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">기업 DB 구축</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">기수 시작 시 1회</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    WALLA 신청서 → <code>walla_to_selected.py</code> → <code>선정기업_명단.xlsx</code> → 구글 드라이브<br>
    포함 정보: 기업명, 이메일, 소재지, 기업유형, TRL, 관심사업분야, 기술키워드, 핵심수요태그, 수출실적
  </div>
</div>

<div style="text-align:center;color:#10B981;font-size:18px;margin:2px 0;">↓</div>

<!-- 단계 2 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#10B981;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 2</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">공고 수집</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">매주 월요일 (수동 실행)</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    bizinfo API → 8개 분야 공고 수집 → <code>notices_db.xlsx</code><br>
    이후 Playwright 크롤링 → 공고 전문 → <code>notices_detail.xlsx</code>
  </div>
</div>

<div style="text-align:center;color:#10B981;font-size:18px;margin:2px 0;">↓</div>

<!-- 단계 3 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#10B981;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 3</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">매칭 실행</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">격주 수요일 (수동 실행)</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    50개사 × 전체 공고 → 7개 축 점수 계산 → 별점 판정 → 상위 N건 추출<br>
    출력: 기업별 ★★★/★★ 공고 목록 (소재지 강등 · 쏠림 페널티 · 세그먼트 부스트 적용)
  </div>
</div>

<div style="text-align:center;color:#10B981;font-size:18px;margin:2px 0;">↓</div>

<!-- 단계 4 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#3B82F6;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 4</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">담당자 검토</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">격주 수~목 (수동)</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    기업별 검토 탭 → 공고 상세 확인 + AI 분석 → ✅ 승인 / ❌ 제외<br>
    검토 상태 드라이브 자동 저장 → 이전/다음 버튼으로 50개사 순회
  </div>
</div>

<div style="text-align:center;color:#10B981;font-size:18px;margin:2px 0;">↓</div>

<!-- 단계 5 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#10B981;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 5</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">발송</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">격주 목요일 (수동 실행)</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    승인된 공고 → 기업별 HTML 메일 발송 (Gmail API)<br>
    메일 내 공고카드: 공고명 + 마감일 + <b style="color:#10B981;">매칭근거 자연어 한 줄</b><br>
    동시에: 공고 마감 D-7·D-3 → 기업별 구글 캘린더 이벤트 등록
  </div>
</div>

<div style="text-align:center;color:#10B981;font-size:18px;margin:2px 0;">↓</div>

<!-- 단계 6 -->
<div style="background:#FFFFFF;border:1px solid #E2E8F0;border-radius:8px;padding:14px 18px;">
  <div style="display:flex;align-items:center;gap:12px;">
    <span style="background:#F59E0B;color:#fff;font-size:12px;font-weight:700;padding:3px 10px;border-radius:20px;">STEP 6</span>
    <span style="font-size:14px;font-weight:600;color:#0F172A;">피드백 수집 · 개선</span>
    <span style="font-size:11px;color:#94A3B8;margin-left:auto;">상시</span>
  </div>
  <div style="margin-top:10px;padding-left:4px;color:#475569;font-size:13px;line-height:1.8;">
    기업 회신 메일 → Gmail API 읽기 → Claude 키워드 추출 → 기업 키워드 보완 자동 저장<br>
    검토 제외 패턴 → 피드백 루프 → 다음 매칭 자동 감점<br>
    신청여부·선정결과 → 발송 이력 입력 → 분기 성과 집계
  </div>
</div>

</div>
</div>
        """, unsafe_allow_html=True)

        st.divider()
        st.subheader("📊 점수 계산 한눈에 보기")
        st.markdown("""
| 매칭 축 | 계산 방식 | 기본 점수 | 최대 점수 |
|---------|-----------|-----------|-----------|
| **축1 지원대상** | TARGET_KW 카테고리 매칭 건수 × 가중치 | ×3점 | ~15점 |
| **축2 사업성격** | TYPE_KW 카테고리 매칭 건수 × 가중치 | ×2점 | ~10점 |
| **축3 기업키워드** | 기업 키워드 → 공고 전문 직접 검색 | ×2점 | ~10점 |
| **축4 핵심수요** | 핵심수요태그 → 공고 직접 매칭 | ×3점 | ~9점 |
| **축5 업종역방향** | 제품분야 → INDUSTRY_KW → 공고 간접매칭 | ×1점 | ~6점 |
| **축6 소재지** | 지역 일치 +3점 / 불일치 -5점 | ±점 | +3 / -5 |
| **축7 세그먼트** | 기업 유형별 사업성격 부스트 | +1~4점 | +4점 |
| **★ 별점 보너스** | ★★★ 판정 시 추가 가산 | +5점 | +5점 |
| **쏠림 페널티** | 이미 많은 기업에 추천된 공고 감산 | -3~-20점 | — |
| **피드백 페널티** | 과거 자주 제외한 사업성격 감산 | -3~-6점 | — |

**별점 판정 기준**
- ★★★ : 핵심수요 직접매칭 / 조달+공공조달 / 조달+해외진출 / 수출+해외진출 / 기업키워드+공공조달
- ★★  : 기업키워드+사업성격 / 조달·수출+마케팅·인증 / 중소벤처+조달·해외진출
- 소재지 불일치 시 강제 강등: ★★★→★★, ★★→★ (★는 제외)
        """)

        st.divider()
        st.subheader("🔄 키워드 개선 루프")
        st.markdown("""
```
기업 DB 구축 (기술키워드 · 핵심수요태그)
    ↓
매칭 실행 → 점수 계산
    ↓
담당자 검토 → 제외 패턴 수집
    ↓ (피드백 반영 버튼)
keywords.json 업데이트 → 다음 매칭 자동 반영
    ↓
기업 회신 메일 → Claude 키워드 추출
    ↓
키워드보완 컬럼 업데이트 → 다음 매칭 정확도 향상
    ↓ (반복)
매칭 품질 점진적 개선
```
        """)

        st.divider()
        st.subheader("시스템 구조도")
        st.markdown("""
```
WALLA 신청서
    ↓ walla_to_selected.py (로컬 1회)
선정기업_명단.xlsx → 구글 드라이브

bizinfo API (매주 월요일, Streamlit 앱)
    ↓
notices_db.xlsx → 구글 드라이브

기업마당 크롤링 (공고 수집 탭에서 수동 실행)
    ↓
notices_detail.xlsx → 구글 드라이브

[Streamlit 웹앱] scaleup-matching.streamlit.app
    ↓ 매칭 실행 (전문 내용 활용)
    ↓ 담당자 검토 + AI 분석
    ↓
Gmail API → 기업별 HTML 메일 (맞춤 + 공통 공고)
Calendar API → D-7·D-3·마감 캘린더 등록
```
        """)



        st.divider()
        st.subheader("캘린더 구조")
        st.markdown("""
| 유형 | 대상 | 내용 |
|------|------|------|
| 공통 캘린더 1개 | 운영팀 + 전체 기업 | 전체 공고 마감 D-7·D-3 |
| 분야별 캘린더 4개 | 운영팀 내부 | 기술개발/수출/경영금융/혁신제품 |
| 기업별 개별 캘린더 | 해당 기업만 | 해당 기업 맞춤 공고만 |
        """)

    # ── 탭2: 매칭 로직 ──────────────────────────────
    with tab2:
        st.subheader("전체 매칭 점수 체계")
        st.markdown("""
**기본 구조: 공고 1건 × 기업 1개사 → 점수 계산 → 별점 판정**

최종 점수 = 지원대상 + 사업성격 + 기업키워드 + 핵심수요 + 업종역방향 + 소재지 + 세그먼트부스트 + 피드백페널티
        """)

        st.divider()
        st.subheader("① 사전 필터 (점수 계산 전 자동 제외)")
        st.markdown("""
| 조건 | 제외 이유 |
|------|-----------|
| 마감일 지난 공고 | 신청 불가 |
| 이미 발송한 공고 | 중복 발송 방지 (send_history 체크) |
| 수신거부 기업 | 기업 관리 탭 수신거부=Y |
| 수출실적 없음 + 수출 전용 분야 | 자격 미달 |
| 제외 키워드 포함 공고 | 키워드 관리 탭에서 설정 가능 |
| 업력 초과 (창업 N년 이내 공고) | 설립연도 기반 자동 계산 |
| 예비창업자 전용 공고 | "예비창업자", "창업팀" 등 명시 |
        """)

        st.divider()
        st.subheader("② 매칭 점수 계산 (7개 축)")
        st.markdown("""
**축1. 지원대상 매칭 (+W_TARGET점/건)**
- 공고 텍스트에서 TARGET_KW(조달기업특화/수출기업특화/중소벤처일반) 탐색
- 매칭된 키워드 수 × 가중치(기본 3점)

**축2. 사업성격 매칭 (+W_TYPE점/건)**
- 공고 텍스트에서 TYPE_KW(공공조달/해외진출/마케팅/인증/기술개발/금융/내수/인력) 탐색
- 매칭된 키워드 수 × 가중치(기본 2점)

**축3. 기업키워드 직접매칭 (+W_KW점/건)**
- 기업의 기술키워드 + 제품분야 + 키워드보완 + 핵심수요태그를 공고 전문에서 직접 검색
- 매칭된 키워드 수 × 가중치(기본 2점)
- **역매핑 확장**: 기업 기술키워드 → INDUSTRY_KW 카테고리 → 공고에서 유사어 검색 (3자 이상 포함관계)

**축4. 핵심수요 태그 매칭 (+W_DEMAND점/건)**
- 기업 핵심수요태그가 공고에 직접 등장하면 가산 (가중치 3점, 가장 높음)
- 예: "G-PASS" 기업 + "G-PASS 등록 지원" 공고 → +3점

**축5. 업종 역방향 매칭 (+W_IND점)**
- 기업의 제품분야(WALLA 15개 카테고리) → INDUSTRY_KW → 공고에서 업종 키워드 매칭
- 카테고리 단위로 최대 +6점

**축6. 소재지 (+W_LOC점 또는 -5점)**
- 공고명 [] 패턴 → 주관기관 지자체 → 전문내용 "○○ 소재" 패턴 순으로 감지
- 일치: +2~3점 / 불일치: -3~5점
- **불일치 시 별점 강등**: ★★★→★★, ★★→★ (★는 최종 제외)

**축7. 세그먼트 부스트 (+최대 W_SEG점)**
- 기업을 4개 세그먼트로 자동 분류: 조달강화형/해외진출형/기술개발형/일반형
- 세그먼트에 맞는 사업성격 카테고리에 추가 점수
        """)
        with st.expander("세그먼트 분류 기준"):
            st.markdown("""
| 세그먼트 | 분류 조건 | 부스트 |
|---------|-----------|--------|
| 조달강화형 | 혁신제품·G-PASS·우수조달 태그 있거나, TRL 8~9이면서 수출실적 없음 | 공공조달 +4, 해외진출 +2 |
| 해외진출형 | 수출실적 있거나 해외조달·수출바우처 태그 있음 | 해외진출 +4, 마케팅홍보 +2 |
| 기술개발형 | TRL 4~6이거나 기술사업화·특허·R&D 태그 있음 | 기술개발 +4, 인증특허 +2 |
| 일반형 | 위 해당 없음 | 전 유형 균등 +1 |
            """)

        st.divider()
        st.subheader("③ 별점 판정 기준")
        st.markdown("""
| 별점 | 조건 (아래 중 하나 이상 해당) |
|------|-------------------------------|
| ★★★ | 핵심수요태그 직접매칭 |
| ★★★ | 조달기업특화 + 공공조달 |
| ★★★ | 조달기업특화 + 해외진출 |
| ★★★ | 수출기업특화 + 해외진출 |
| ★★★ | 기업키워드 매칭 + 공공조달 |
| ★★ | 기업키워드 매칭 + 사업성격 |
| ★★ | 조달·수출기업 + 마케팅·인증·기술개발 |
| ★★ | 중소벤처일반 + 공공조달 또는 해외진출 |
| ★★ | 수출국가 매칭 + 조달 또는 해외 |
        """)

        st.divider()
        st.subheader("④ 공고 쏠림 페널티")
        st.markdown("""
같은 공고가 너무 많은 기업에게 추천되면 자동 감산:

| 추천 기업 수 | 점수 감산 |
|-------------|-----------|
| 2개사 이상 | -3점 |
| 4개사 이상 | -6점 |
| 7개사 이상 | -10점 |
| 10개사 이상 | -15점 |
| 15개사 이상 | -20점 (사실상 차단) |
        """)

        st.divider()
        st.subheader("⑤ 소재지 판단 우선순위")
        st.markdown("""
| 우선순위 | 감지 방법 | 예시 | 신뢰도 |
|----------|-----------|------|--------|
| 1순위 | 공고명 [] 패턴 | [경기] 중소기업 수출지원 | 최고 |
| 2순위 | 주관기관이 지자체 | 충청남도, 경기도청 | 높음 |
| 3순위 | 전문내용 "○○ 소재 기업" | "충남 소재 중소기업 대상" | 중간 |
| 4순위 | 전문내용 지역 언급 패턴 | "광주에 위치한", "경남 관내" | 보조 |
        """)

        st.divider()
        st.subheader("⑥ 피드백 루프")
        st.markdown("""
검토 완료 → "피드백 반영" 버튼 클릭 → 제외한 공고의 사업성격 패턴 집계 → keywords.json 저장

다음 매칭 실행 시 해당 기업에게 자주 제외된 사업성격 유형 감산:
- 2회 이상 제외 → -3점
- 3회 이상 제외 → -6점

**효과**: 사이클이 반복될수록 각 기업에게 맞지 않는 유형의 공고가 자동으로 걸러집니다.
        """)

    # ── 탭3: 키워드 구조 ────────────────────────────
    with tab3:
        st.subheader("축1 — 지원대상 키워드")
        for cat, kws in TARGET_KW.items():
            with st.expander(f"**{cat}** ({len(kws)}개)"):
                st.markdown(", ".join(f"`{k}`" for k in kws))

        st.divider()
        st.subheader("축2 — 사업성격 키워드")
        for cat, kws in TYPE_KW.items():
            with st.expander(f"**{cat}** ({len(kws)}개)"):
                st.markdown(", ".join(f"`{k}`" for k in kws))

        st.divider()
        st.subheader("축3 — 제품분야 역방향 키워드 (WALLA 15개 카테고리)")
        for cat, kws in INDUSTRY_KW.items():
            with st.expander(f"**{cat}** ({len(kws)}개)"):
                st.markdown(", ".join(f"`{k}`" for k in kws))

        # 총 키워드 수 집계
        total = (sum(len(v) for v in TARGET_KW.values()) +
                 sum(len(v) for v in TYPE_KW.values()) +
                 sum(len(v) for v in INDUSTRY_KW.values()))
        st.info(f"전체 키워드 총 {total}개")

    # ── 탭4: 파일 구조 ──────────────────────────────
    with tab4:
        st.subheader("드라이브 파일")
        st.markdown("""
| 파일명 | 역할 | 생성 방법 |
|--------|------|-----------|
| 선정기업_명단.xlsx | 기업 DB | walla_to_selected.py 로컬 실행 |
| notices_db.xlsx | 공고 DB | 공고 수집 버튼 (매주 월요일) |
| **notices_detail.xlsx** | **공고 전문 DB** | **공고 수집 탭에서 크롤링 실행** |
| send_history.xlsx | 발송 이력 | 발송 실행 시 자동 기록 |
| keywords.json | 매칭 키워드 | 설정 탭에서 저장 시 자동 생성 |
| calendar_id.txt | 공통 캘린더 ID | create_common_calendar.py 실행 |
| individual_calendars.json | 기업별 캘린더 | create_individual_calendars.py 실행 |
        """)

        st.divider()
        st.subheader("깃허브 파일")
        st.markdown("""
| 파일명 | 역할 |
|--------|------|
| app.py | Streamlit 웹앱 메인 |
| requirements.txt | 라이브러리 목록 |
| **crawl_notices.py** | **공고 전문 크롤러 (앱 내 직접 실행)** |
| **walla_to_selected.py** | **WALLA → 선정기업 명단 변환 (선정 후 1회)** |

| create_individual_calendars.py | 기업별 캘린더 생성 (선정 후) |
| auth_test.py | 구글 인증 (최초 1회) |
| logo.png | 혁신제품지원센터 CI |
        """)

        st.divider()
        st.subheader("선정기업 명단 컬럼")
        st.markdown("""
| 컬럼 | 구분 | 매칭 활용 |
|------|------|-----------|
| 기업명 | 필수 | 기준키 |
| 소재지 | 필수 | 지역 필터 |
| 이메일 | 필수 | 발송 대상 |
| 관심사업분야 | 필수 | 분야 필터 |
| 기술키워드 | 필수 | 키워드 매칭 |
| 제품분야 | 필수 | 역방향 매칭 |
| 수출실적 | 필수 | 수출 공고 필터 |
| 수출국가 | 필수 | 국가 매칭 |
| TRL단계 | 매칭고도화 | R&D 공고 필터 |
| 핵심수요태그 | 매칭고도화 | 최우선 매칭 (+3점) |
| 매출규모 | 매칭고도화 | (추후 활용) |
| 키워드보완 | 운영관리 | 보완 키워드 |
| 수신거부 | 운영관리 | 발송 제외 |
        """)
