"""
평가결과 파일 → 선정기업_명단.xlsx 자동 변환 스크립트 (v4: 단일 파일 구조)
사용법:
  1. SOURCE_FILE에 평가 결과 파일 경로 입력
     (이 파일 하나에 '원본데이터(WALLA)' 시트와 '종합순위_내부협의' 시트가 모두 포함되어 있음)
  2. python walla_to_selected.py 실행
  3. 생성된 선정기업_명단.xlsx를 구글 드라이브에 업로드

v4 변경사항:
  - 별도 WALLA CSV 불필요 → 평가결과 xlsx 단일 파일로 전체 변환
    ('원본데이터(WALLA)' 시트가 WALLA 신청 원본을 그대로 담고 있음, 헤더 2행/데이터 3행부터)
  - '종합순위_내부협의' 시트에서 선정/예비 구분 + 정량점수/총점/내부논의/서비스요청 매핑
  - 사업자등록번호 기준 매칭 (동명/표기차 리스크 제거)
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re, os

# ── 설정 ─────────────────────────────────────────────
# 스크립트 파일이 있는 폴더를 기준으로 경로를 잡는다 (실행 위치가 달라도 안전)
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE     = os.path.join(BASE_DIR, "_한국조달연구원_2026_원스톱_스케일업_기업_선정_최종_20260615_.xlsx")
WALLA_SHEET     = "원본데이터(WALLA)"
SELECTION_SHEET = "종합순위_내부협의"
OUTPUT_FILE     = os.path.join(BASE_DIR, "선정기업_명단.xlsx")


def clean_biznum(v):
    """사업자등록번호를 비교 가능한 순수 숫자 문자열로 정규화"""
    return re.sub(r'\D', '', str(v))


def load_selection_roster(path, sheet):
    """종합순위_내부협의 시트에서 사업자등록번호 -> 선정/예비 + 평가정보 매핑을 만든다.
    컬럼 구조(3행 헤더 기준): A순위 B기업명 C사업자등록번호 D소재지 E지역
                              F기업유형 G키워드(참고) H정량(30) I총점(100)
                              J내부논의 K서비스요청(참고) L선정(O/X) M최종구분
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    roster = {}
    for row in ws.iter_rows(min_row=4, values_only=True):  # 4행부터 데이터
        if not row or len(row) < 13:
            continue
        (rank, name, biznum, addr, region, biztype, keywords,
         quant_score, total_score, internal_note, service_req,
         selected_flag, final_label) = row[:13]

        if not biznum:
            continue
        biznum_clean = clean_biznum(biznum)
        if not biznum_clean:
            continue

        if final_label and '선정' in str(final_label):
            status = '선정'
        elif final_label and '예비' in str(final_label):
            status = '예비'
        elif selected_flag == 'O':
            status = '선정'   # final_label이 비어있어도 O 표시가 있으면 선정으로 처리
        else:
            continue  # 70개사에 들지 못한 나머지 113개사는 제외

        roster[biznum_clean] = {
            '선정구분':   status,
            '순위':       rank,
            '기업명_원본': name,
            '평가_정량점수': quant_score,
            '평가_총점':    total_score,
            '평가_내부논의': internal_note,
            '평가_서비스요청': service_req,
        }

    return roster

# ── 기업유형 → 핵심수요태그 자동 변환 ─────────────────
def extract_demand_tags(row):
    tags = []

    # 기업유형에서 추출
    biztype = str(row.get('기업유형',''))
    if '혁신제품' in biztype: tags.append('혁신제품')
    if 'G-PASS'  in biztype: tags.append('G-PASS')
    if '우수조달' in biztype: tags.append('우수조달')
    if '벤처'    in biztype: tags.append('벤처기업')

    # 검토중인 지원사업에서 추출
    for col in ['검토사업1','검토사업2','검토사업3']:
        v = str(row.get(col,''))
        if v and v != 'nan':
            if '시범구매'  in v: tags.append('시범구매')
            if '해외조달'  in v: tags.append('해외조달')
            if '우수제품'  in v: tags.append('우수제품')
            if 'MAS'       in v: tags.append('MAS등록')
            if '혁신제품'  in v: tags.append('혁신제품지정')
            if '수출바우처' in v: tags.append('수출바우처')
            if '기술사업화' in v: tags.append('기술사업화')
            if '전시회'    in v or '박람회' in v: tags.append('해외전시')

    # 세미나 희망에서 추출
    seminar = str(row.get('세미나희망',''))
    if '조달'    in seminar: tags.append('조달컨설팅')
    if '마케팅'  in seminar: tags.append('온라인마케팅')
    if 'MAS'     in seminar or '다수공급자' in seminar: tags.append('MAS등록')
    if 'CES'     in seminar: tags.append('CES')
    if '특허'    in seminar: tags.append('특허')

    # 중복 제거
    return ', '.join(list(dict.fromkeys(tags)))

# ── TRL 단계 단순화 ───────────────────────────────────
def parse_trl(trl_str):
    if 'TRL 8-9' in str(trl_str): return '9'
    if 'TRL 6-7' in str(trl_str): return '7'
    if 'TRL 4-5' in str(trl_str): return '5'
    return ''

# ── 수출국가 합치기 ───────────────────────────────────
def merge_countries(row):
    countries = []
    for col in ['수출국가①','수출국가②','수출국가③','수출국가④','수출국가⑤']:
        v = str(row.get(col,''))
        if v and v != 'nan': countries.append(v.strip())
    return ', '.join(countries)

# ── 메인 변환 ─────────────────────────────────────────
def main():
    if not os.path.exists(SOURCE_FILE):
        print(f"\n❌ 파일을 찾을 수 없습니다: {SOURCE_FILE}")
        print(f"   스크립트 위치: {BASE_DIR}")
        print(f"   해당 폴더의 실제 파일 목록:")
        for f in os.listdir(BASE_DIR):
            if f.lower().endswith('.xlsx'):
                print(f"     - {f}")
        return
    print(f"WALLA 원본 읽는 중: {SOURCE_FILE} [{WALLA_SHEET}]")
    # 1행: 섹션 구분(기본정보/기업현황 등), 2행: 실제 문항 헤더, 3행부터 데이터
    df_w = pd.read_excel(SOURCE_FILE, sheet_name=WALLA_SHEET, header=1, dtype=str)

    # 컬럼 매핑 (위치 기반 — 헤더 텍스트가 길고 특수문자 포함이라 인덱스로 고정)
    idx_map = {
        '응답시간': 1,
        '기업명': 2, '사업자등록번호': 3, '소재지': 4, '이메일': 7, '기업유형': 8,
        '제품분야': 10, '매출규모': 16, 'TRL원본': 20, '수출실적': 22,
        '수출국가①': 23, '수출국가②': 24, '수출국가③': 25, '수출국가④': 26, '수출국가⑤': 27,
        '세미나희망': 28, '검토사업1': 36, '검토사업2': 37, '검토사업3': 38,
        '관심사업분야': 39, '기술키워드': 40, '희망서비스': 45, '정보수신이메일': 49,
    }

    df_mapped = pd.DataFrame()
    for new_col, ci in idx_map.items():
        df_mapped[new_col] = df_w.iloc[:, ci] if ci < df_w.shape[1] else ''

    df_mapped = df_mapped.fillna('')
    df_mapped['_biznum_clean'] = df_mapped['사업자등록번호'].apply(clean_biznum)

    # 평가 결과 시트 로드 (사업자등록번호 → 선정구분 + 평가정보 매핑)
    print(f"평가 결과 읽는 중: {SOURCE_FILE} [{SELECTION_SHEET}]")
    roster = load_selection_roster(SOURCE_FILE, SELECTION_SHEET)
    print(f"평가 결과 70개사 로드 (선정 {sum(1 for v in roster.values() if v['선정구분']=='선정')}개, "
          f"예비 {sum(1 for v in roster.values() if v['선정구분']=='예비')}개)")

    # 사업자등록번호로 매칭
    df_sel = df_mapped[df_mapped['_biznum_clean'].isin(roster.keys())].copy()
    for field in ['선정구분','순위','평가_정량점수','평가_총점','평가_내부논의','평가_서비스요청']:
        df_sel[field] = df_sel['_biznum_clean'].map(lambda b: roster[b][field])

    # 동일 사업자등록번호로 WALLA 중복 제출된 경우 → 최신 응답시간 1건만 채택
    dup_groups = df_sel.groupby('_biznum_clean').filter(lambda g: len(g) > 1)
    if not dup_groups.empty:
        print(f"\n⚠️ WALLA 중복 제출 감지 ({dup_groups['_biznum_clean'].nunique()}개사, {len(dup_groups)}건):")
        for biznum, g in dup_groups.groupby('_biznum_clean'):
            names = g['기업명'].tolist()
            times = g['응답시간'].tolist()
            print(f"   - {names} (사업자번호 {biznum}) 응답시간: {times} → 최신 제출 채택")
    df_sel = df_sel.sort_values('응답시간').drop_duplicates('_biznum_clean', keep='last')

    matched_biznums = set(df_sel['_biznum_clean'])
    not_found = [v['기업명_원본'] for k, v in roster.items() if k not in matched_biznums]
    if not_found:
        print(f"\n⚠️ WALLA에서 사업자등록번호로 찾을 수 없는 기업({len(not_found)}개): {not_found}")
        print("   → 사업자등록번호 오기입 가능성, 수동 확인 필요")

    print(f"\n변환 대상: {len(df_sel)}개사 (선정 {(df_sel['선정구분']=='선정').sum()}개, "
          f"예비 {(df_sel['선정구분']=='예비').sum()}개)")

    # 최종 컬럼 생성
    records = []
    for _, row in df_sel.iterrows():
        # 이메일: 정보수신 이메일 우선, 없으면 담당자 이메일
        email = row.get('정보수신이메일','') or row.get('이메일','')

        records.append({
            '기업명':        row['기업명'],
            '사업자등록번호': str(row['사업자등록번호']).replace('.0',''),
            '선정구분':      row['선정구분'],          # 선정 / 예비
            '순위':          row['순위'],              # 평가 순위 (1~70)
            '평가_정량점수': row['평가_정량점수'],      # 정량평가 30점 만점
            '평가_총점':     row['평가_총점'],          # 정량+주관식 합산 100점 만점
            '평가_내부논의': row['평가_내부논의'],      # 평가위원회 내부 협의 메모
            '평가_서비스요청': row['평가_서비스요청'],  # 신청 시 작성한 희망 서비스(주관식)
            '소재지':        row['소재지'],
            '이메일':        email.strip(),
            '관심사업분야':  row['관심사업분야'],
            '기술키워드':    row['기술키워드'],
            '제품분야':      row['제품분야'],
            '수출실적':      row['수출실적'],
            '수출국가':      merge_countries(row),
            'TRL단계':       parse_trl(row['TRL원본']),
            '매출규모':      row['매출규모'],
            '기업유형':      row['기업유형'],
            '핵심수요태그':  extract_demand_tags(row),
            '희망서비스요약': str(row['희망서비스'])[:100] if row['희망서비스'] else '',
            '구글계정':      '',
            '키워드보완':    '',
            '수신거부':      '',
            '메모':          '',
        })

    df_out = pd.DataFrame(records).sort_values('순위').reset_index(drop=True)

    # 엑셀 저장 (색상 구분)
    save_excel(df_out, OUTPUT_FILE)
    print(f"\n✅ 완료: {OUTPUT_FILE} ({len(df_out)}개사)")
    print("\n다음 단계:")
    print("  1. 생성된 파일을 구글 드라이브에 업로드")
    print("  2. 앱 기업 관리 메뉴에서 확인")
    print("  3. 핵심수요태그 보완 후 매칭 실행")


def save_excel(df, filename):
    NAVY  = "1F4E79"; BLUE  = "2E75B6"; LBLUE = "BDD7EE"
    TEAL  = "00897B"; LTEAL = "E0F2F1"
    PURP  = "5C2D91"; LPURP = "EDE7F6"
    ORNG  = "E65100"; LORNG = "FFE0B2"

    col_groups = {
        '선정구분':      ('선정정보', ORNG, LORNG),
        '순위':          ('선정정보', ORNG, LORNG),
        '평가_정량점수': ('선정정보', ORNG, LORNG),
        '평가_총점':     ('선정정보', ORNG, LORNG),
        '평가_내부논의': ('선정정보', ORNG, LORNG),
        '평가_서비스요청': ('선정정보', ORNG, LORNG),
        '기업명':        ('필수', BLUE, LBLUE),
        '사업자등록번호': ('필수', BLUE, LBLUE),
        '소재지':        ('필수', BLUE, LBLUE),
        '이메일':        ('필수', BLUE, LBLUE),
        '관심사업분야':  ('필수', BLUE, LBLUE),
        '기술키워드':    ('필수', BLUE, LBLUE),
        '제품분야':      ('필수', BLUE, LBLUE),
        '수출실적':      ('필수', BLUE, LBLUE),
        '수출국가':      ('필수', BLUE, LBLUE),
        'TRL단계':       ('매칭고도화', TEAL, LTEAL),
        '매출규모':      ('매칭고도화', TEAL, LTEAL),
        '기업유형':      ('매칭고도화', TEAL, LTEAL),
        '핵심수요태그':  ('매칭고도화', TEAL, LTEAL),
        '희망서비스요약': ('매칭고도화', TEAL, LTEAL),
        '구글계정':      ('운영관리', PURP, LPURP),
        '키워드보완':    ('운영관리', PURP, LPURP),
        '수신거부':      ('운영관리', PURP, LPURP),
        '메모':          ('운영관리', PURP, LPURP),
    }

    wb = Workbook(); ws = wb.active
    ws.title = "선정기업명단"
    ws.sheet_view.showGridLines = False
    s   = Side(style="thin", color="BFBFBF")
    bdr = Border(left=s, right=s, top=s, bottom=s)

    # 1행: 구분
    ws.row_dimensions[1].height = 18
    for ci, col in enumerate(df.columns, 1):
        grp, hc, dc = col_groups.get(col, ('운영관리', PURP, LPURP))
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col)*2, 14)
        c = ws.cell(row=1, column=ci, value=grp)
        c.fill = PatternFill("solid", start_color=hc, end_color=hc)
        c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr

    # 2행: 헤더
    ws.row_dimensions[2].height = 24
    for ci, col in enumerate(df.columns, 1):
        grp, hc, dc = col_groups.get(col, ('운영관리', PURP, LPURP))
        c = ws.cell(row=2, column=ci, value=col)
        c.fill = PatternFill("solid", start_color=hc, end_color=hc)
        c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr

    # 데이터 행
    grp_bg = {"필수": "FFFFFF", "매칭고도화": "F5FFFE", "운영관리": "FAF5FF", "선정정보": "FFF8F0"}
    for ri, row in enumerate(df.itertuples(index=False), 3):
        ws.row_dimensions[ri].height = 18
        for ci, (col, val) in enumerate(zip(df.columns, row), 1):
            grp, hc, dc = col_groups.get(col, ('운영관리', PURP, LPURP))
            c = ws.cell(row=ri, column=ci, value=str(val) if val else '')
            c.fill = PatternFill("solid", start_color=grp_bg[grp], end_color=grp_bg[grp])
            c.font = Font(name="맑은 고딕", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = bdr

    wb.save(filename)


if __name__ == "__main__":
    main()
